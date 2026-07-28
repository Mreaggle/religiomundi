#!/usr/bin/env python3
"""Build the revised comparative-religion workbook.

The workbook is generated from structured data so that its taxonomy, mappings,
and editorial caveats can be reviewed independently from Excel formatting.
"""

from __future__ import annotations

import csv
from copy import copy
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "UNO.xlsx"
OUTPUT = ROOT / "public" / "data" / "UNO_reformulado.xlsx"
CSV_OUTPUT = ROOT / "data" / "UNO_reformulado.csv"
REVISION_DATE = date(2026, 7, 28)


COLORS = {
    "navy": "17253A",
    "blue": "244A6B",
    "teal": "147D7E",
    "green": "4F7D55",
    "gold": "C69C46",
    "sand": "EFE6D5",
    "cream": "F8F5EE",
    "light_blue": "DDEAF3",
    "light_green": "E2F0D9",
    "light_gold": "FFF2CC",
    "light_red": "F4CCCC",
    "light_gray": "E7E6E6",
    "mid_gray": "B7B7B7",
    "dark": "1F1F1F",
    "white": "FFFFFF",
}

THIN_GRAY = Side(style="thin", color="C9C9C9")
MEDIUM_NAVY = Side(style="medium", color=COLORS["navy"])
GRID = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


@dataclass(frozen=True)
class Archetype:
    code: str
    name: str
    definition: str
    exclusion: str


@dataclass
class Tradition:
    name: str
    family: str
    region: str
    period: str
    kind: str
    profile: str
    sources: str
    status: str = "Viva"
    coverage: str = "Perfil de família"
    note: str = ""
    overrides: dict[str, str] = field(default_factory=dict)
    tradition_id: str = ""


ARCHETYPES = [
    Archetype("A01", "Origem / absoluto", "Estado, princípio ou realidade primordial de onde procede o cosmos.", "Não equivale automaticamente a um deus criador."),
    Archetype("A02", "Criador(a) / demiurgo", "Agente que cria, forma ou organiza o mundo ou a humanidade.", "Distinguir criação, geração e ordenação."),
    Archetype("A03", "Soberania celeste / divindade suprema", "Polo de supremacia divina, céu ou autoridade máxima.", "Não presumir monoteísmo."),
    Archetype("A04", "Ordem cósmica / verdade", "Princípio de ordem, verdade, lei ou harmonia que estrutura o real.", "Não reduzir a moral individual."),
    Archetype("A05", "Terra-mãe / fertilidade", "Terra, fecundidade, maternidade cósmica ou potência geradora.", "Evitar a hipótese universal da 'Grande Deusa'."),
    Archetype("A06", "Céu / tempestade / trovão", "Céu atmosférico, chuva, tempestade, raio ou trovão.", "Separar do sol e do céu soberano quando possível."),
    Archetype("A07", "Sol / luz", "Sol, claridade, luz divina ou função solar.", "Metáforas de luz são correlações parciais."),
    Archetype("A08", "Lua / noite", "Lua, noite e ritmos lunares.", "Não inferir gênero universal da Lua."),
    Archetype("A09", "Estrelas / aurora", "Astros, constelações, planeta Vênus ou aurora.", "Astrologia posterior requer recorte próprio."),
    Archetype("A10", "Mar / rios / águas", "Oceano, águas primordiais, rios, fontes ou chuva personificada.", "Distinguir água cósmica de divindade marítima."),
    Archetype("A11", "Fogo / lar / forja", "Fogo ritual, doméstico, solar, vulcânico ou metalúrgico.", "A forja pode pertencer também à técnica."),
    Archetype("A12", "Vento / ar / sopro", "Vento, ar, respiração ou sopro vital.", "Metáforas espirituais recebem correlação parcial."),
    Archetype("A13", "Agricultura / vegetação", "Cultivo, grãos, florestas, plantas e ciclos vegetais.", "Não confundir toda fertilidade com agricultura."),
    Archetype("A14", "Caça / animais / mundo selvagem", "Animais tutelares, caça, pastoreio ou domínio do não domesticado.", "Totemismo não é categoria universal."),
    Archetype("A15", "Amor / sexualidade / beleza", "Desejo, erotismo, beleza, atração e vínculo amoroso.", "Separar de matrimônio e parto."),
    Archetype("A16", "Matrimônio / parto / família", "Casamento, gestação, nascimento, infância e proteção familiar.", "Não presumir um único modelo familiar."),
    Archetype("A17", "Guerra / força / heroísmo", "Combate, força, vitória e figura heroica.", "Distinguir guerra, defesa e soberania."),
    Archetype("A18", "Realeza / lei / autoridade", "Realeza sagrada, governo, lei, mandato ou legitimidade.", "Não confundir com justiça pós-morte."),
    Archetype("A19", "Sabedoria / escrita / conhecimento", "Sabedoria, memória, escrita, ensino, adivinhação ou conhecimento.", "Ciência moderna não é panteão."),
    Archetype("A20", "Artesanato / técnica / invenção", "Ofícios, construção, tecelagem, metalurgia e invenção.", "Tecnologia moderna fica fora da matriz religiosa."),
    Archetype("A21", "Mensageiro / limiar / trickster", "Comunicação, caminhos, comércio de mensagens, transgressão ou liminaridade.", "O termo trickster é comparativo, não identidade nativa."),
    Archetype("A22", "Comércio / riqueza / fortuna", "Troca, prosperidade, sorte, tesouros e abundância.", "Não reduzir prosperidade a riqueza monetária."),
    Archetype("A23", "Cura / medicina", "Cura, saúde, plantas medicinais ou restauração corporal.", "Distinguir cura ritual de medicina científica."),
    Archetype("A24", "Doença / peste", "Doenças, epidemias, aflição e suas personificações.", "Não patologizar espíritos ou experiências religiosas."),
    Archetype("A25", "Morte / submundo", "Morte, reino dos mortos, sepultura ou dissolução.", "Não equivale necessariamente a mal."),
    Archetype("A26", "Ancestrais / mortos venerados", "Ancestrais, linhagens, memória ritual e mortos socialmente presentes.", "Distinguir ancestral de divindade."),
    Archetype("A27", "Psicopompo / guia", "Condução entre mundos, sobretudo de mortos, iniciandos ou viajantes.", "Mensageiro não é sempre psicopompo."),
    Archetype("A28", "Justiça / juramento / julgamento", "Justiça, juramento, retribuição e julgamento dos mortos.", "Distinguir lei política de julgamento cósmico."),
    Archetype("A29", "Destino / tempo", "Destino, sorte atribuída, tempo, ciclos e duração.", "Livre-arbítrio e determinismo são debates, não panteões."),
    Archetype("A30", "Caos / destruição", "Desordem primordial, catástrofe, ruptura ou destruição cósmica.", "Caos não implica mal moral."),
    Archetype("A31", "Adversário / mal personificado", "Figura ou princípio que personifica oposição, engano ou mal.", "Não impor dualismo a sistemas não dualistas."),
    Archetype("A32", "Renascimento / ressurreição / renovação", "Retorno à vida, renascimento, regeneração ou renovação do cosmos.", "Distinguir ciclo sazonal, reencarnação e ressurreição."),
    Archetype("A33", "Proteção / guardiões", "Proteção de pessoas, casas, cidades, caminhos ou doutrina.", "Inclui espíritos tutelares apenas quando documentados."),
    Archetype("A34", "Música / poesia / dança / arte", "Criação estética, canto, dança, poesia e performance ritual.", "Beleza visual pode também aparecer em A15."),
    Archetype("A35", "Êxtase / intoxicação / transgressão ritual", "Transe, possessão, substâncias, frenesi ou inversão ritual.", "Uso de enteógenos exige evidência específica; não inferir da arte."),
    Archetype("A36", "Ascese / renúncia", "Disciplina, celibato, jejum, monasticismo e renúncia.", "Não confundir privação involuntária com ascese."),
    Archetype("A37", "Compaixão / misericórdia", "Misericórdia, cuidado universal, caridade e compaixão.", "Virtude ética pode não ser personificada."),
    Archetype("A38", "Revelação / profecia / enviado", "Revelação textual ou oral, profeta, mensageiro divino ou mestre fundador.", "Não chamar todo herói cultural de profeta."),
    Archetype("A39", "Alma / consciência / princípio vital", "Alma, espírito, sopro, mente ou potência vital.", "Termos não são equivalentes; registrar a terminologia nativa."),
    Archetype("A40", "Equilíbrio / dualidade / complementaridade", "Pares, equilíbrio, reciprocidade ou complementaridade cosmológica.", "Dualidade não significa necessariamente conflito."),
    Archetype("A41", "Salvação / libertação / iluminação", "Objetivo último: salvação, libertação, iluminação, união ou vida plena.", "Não forçar soteriologia onde não existe."),
    Archetype("A42", "Escatologia / renovação final", "Fim dos tempos, juízo final, renovação ou consumação histórica.", "Distinguir ciclos cósmicos de fim linear."),
    Archetype("A43", "Sacrifício / oferenda / mediação", "Oferenda, sacrifício, pacto, sacerdócio ou mediação ritual.", "Não assumir sacrifício humano sem evidência."),
    Archetype("A44", "Transformação / iniciação / metamorfose", "Mudança de estado, iniciação, metamorfose ou passagem ritual.", "Distinguir conversão, rito de passagem e mito."),
]


def mapped(**items: str) -> dict[str, str]:
    """Return a profile map while making profile declarations compact."""
    return dict(items)


PROFILES: dict[str, dict[str, str]] = {
    "fragmentary": mapped(
        A01="? Cosmogonia não recuperável com segurança",
        A26="≈ Práticas mortuárias ou memória ancestral; interpretação limitada",
        A43="≈ Depósitos/ofertas arqueológicas; função exata incerta",
    ),
    "archaeological": mapped(
        A01="? Crenças não recuperáveis diretamente da cultura material",
        A05="? Figuras femininas/fertilidade: interpretação contestada",
        A14="≈ Relações humano-animal sugeridas pela iconografia",
        A25="≈ Sepultamentos intencionais e tratamento dos mortos",
        A34="≈ Arte parietal, objetos e possível performance",
        A35="? Hipóteses xamânicas/enteogênicas sem consenso geral",
        A39="? Comportamento simbólico; conceito de alma desconhecido",
        A43="≈ Deposição ritual e bens funerários",
        A44="≈ Ritos de passagem são hipótese, não leitura direta",
    ),
    "indigenous_local": mapped(
        A01="≈ Narrativas locais de origem; não há corpus único",
        A05="≈ Terra/território como relação sagrada; varia por povo",
        A14="≈ Relações com animais e seres-mais-que-humanos; varia",
        A25="≈ Concepções locais de morte; terminologia específica necessária",
        A26="≈ Ancestrais e memória ritual; varia",
        A33="≈ Seres tutelares/protetores locais; varia",
        A35="≈ Especialistas e estados rituais; não presumir substâncias",
        A39="≈ Princípio vital/pessoa relacional; terminologia local necessária",
        A43="≈ Reciprocidade, oferenda ou partilha ritual; varia",
        A44="≈ Ritos de passagem e iniciação; varia",
    ),
    "north_american": mapped(
        A01="≈ Narrativas de emergência, criação ou ordenação; varia por nação",
        A03="≈ Potência sagrada suprema; tradução 'Grande Espírito' não é universal",
        A05="≈ Terra como parente/território vivo; formulações locais",
        A06="≈ Seres do trovão; variam por nação",
        A07="≈ Sol em funções rituais e cosmológicas",
        A14="≈ Parentes animais, caça e seres tutelares",
        A21="≈ Coiote/Corvo ou outros transformadores; não universais",
        A23="≈ Cerimônias e especialistas de cura",
        A26="≈ Ancestrais e parentes mortos",
        A33="≈ Seres/objetos protetores",
        A38="≈ Visões e transmissão cerimonial",
        A39="≈ Sopro, sombra ou alma; terminologia varia",
        A43="≈ Oferenda, dança, cachimbo ou reciprocidade; conforme a nação",
        A44="≈ Busca de visão/iniciação; não universal",
    ),
    "australian": mapped(
        A01="◇ Dreaming/Dreamings: ordem ancestral contínua; termos variam",
        A02="≈ Seres ancestrais formadores; não um criador pan-australiano",
        A03="? Baiame em alguns povos do sudeste; não universal",
        A05="◇ Country como relação viva, ancestral e normativa",
        A14="≈ Ancestrais animais/territoriais e espécies totêmicas; local",
        A18="◇ Lei/Lore transmitida pelos seres ancestrais",
        A25="≈ Morte e retorno/continuidade com Country; varia",
        A26="● Ancestrais e presença ancestral no território",
        A33="≈ Seres e lugares tutelares locais",
        A34="● Canto, dança, pintura e narrativa cerimonial",
        A35="≈ Estados rituais; não inferir enteógenos",
        A39="≈ Espírito/princípio vital; vocabulários locais",
        A43="● Cerimônia e cuidado recíproco de Country",
        A44="● Iniciação e transmissão graduada de conhecimento",
    ),
    "polynesian": mapped(
        A01="≈ Te Pō/Po e genealogias de origem; versões insulares",
        A02="≈ Atua ancestrais e pares geradores; varia",
        A03="≈ Tangaroa/Tagaloa ou divindades celestes; varia",
        A05="≈ Papa/Papahānaumoku: terra geradora",
        A06="≈ Tāwhirimātea e potências atmosféricas; varia",
        A07="≈ Tama-nui-te-rā/Māui e tradições solares",
        A08="≈ Hina/Sina e tradições lunares",
        A10="● Tangaroa/Kanaloa: mar e pesca",
        A11="≈ Mahuika/Pele: fogo; tradições distintas",
        A13="≈ Rongo/Lono: cultivo e paz",
        A14="≈ Tāne/Kāne: floresta, vida e aves",
        A17="≈ Tū/Kū: guerra e ação",
        A21="● Māui: transformador e herói liminar",
        A25="≈ Hine-nui-te-pō/Milu: morte; varia",
        A26="● Genealogia, ancestrais e chefia",
        A33="≈ Atua tutelares de linhagem e lugar",
        A34="● Canto, dança, genealogia e poesia",
        A39="◇ Mana, mauri/wairua ou conceitos cognatos; não idênticos",
        A40="◇ Tapu/noa e reciprocidade; formulações locais",
        A43="● Karakia, oferendas e mediação sacerdotal",
    ),
    "sami": mapped(
        A03="≈ Rádien/Radien-attje em tradições históricas",
        A05="● Máttaráhkká e filhas: maternidade, gestação e terra",
        A06="● Horagalles/Dierpmis: trovão",
        A07="● Beaivi: Sol e restauração da vida",
        A08="≈ Mánnu: Lua",
        A12="≈ Biegga-almmái: vento",
        A14="● Leib-Olmai e potências da caça; urso ritual",
        A23="● Noaidi e práticas de cura/divinação",
        A25="≈ Ruohtta e domínios de doença/morte em registros históricos",
        A26="● Ancestrais e mundos sáiva/saivo",
        A27="≈ Noaidi como viajante entre domínios",
        A34="● Joik e tambor em contextos rituais",
        A35="● Transe do noaidi; não implica uso de Amanita",
        A39="≈ Conceitos múltiplos de alma/espírito",
        A43="● Oferendas em sieidi",
    ),
    "guarani": mapped(
        A01="◇ Yvy Tenonde e ciclos de criação; versões Mbyá/Kaiowá variam",
        A02="● Nhanderu Ete/Tenondé: origem e criação em tradições Mbyá",
        A03="● Nhanderu: 'nosso pai'; terminologia varia",
        A05="≈ Nhandesy: princípio materno; versões locais",
        A06="≈ Tupã como classe/potência do trovão em recortes Guarani",
        A07="● Kuaray: Sol",
        A08="● Jaxy: Lua",
        A13="≈ Jakairá e a vitalidade das plantas/cultivos em alguns relatos",
        A21="≈ heróis gêmeos/transformadores; versões locais",
        A25="≈ Angue e regiões pós-morte; vocabulário varia",
        A26="● Nhe'ẽ e continuidade ancestral/comunitária",
        A34="● Canto-reza, dança e palavra inspirada",
        A38="● Palavra-alma e transmissão pelos xamãs/rezadores",
        A39="● Nhe'ẽ: palavra-alma/princípio pessoal",
        A41="◇ Yvy Marã E'ỹ: 'terra sem mal', interpretações históricas diversas",
        A43="● Reza, canto, fumaça e reciprocidade ritual",
    ),
    "tengrism": mapped(
        A01="≈ Céu eterno e ordem cosmológica; mitos variam",
        A02="≈ Kayra/Ülgen em tradições altaicas posteriores; não pan-túrquico",
        A03="● Tengri: Céu soberano",
        A05="● Etügen/Yer-Sub: terra e águas",
        A06="● Tengri e potências atmosféricas",
        A07="≈ Sol como potência celeste",
        A08="≈ Lua como potência celeste",
        A10="≈ Yer-Sub: águas e lugares",
        A11="≈ Od Ana: fogo doméstico em tradições túrquicas",
        A16="● Umay: parto, crianças e linhagem",
        A17="≈ Sulde e fortuna guerreira; contexto mongol",
        A18="● Tengri legitima o qaghan/cã",
        A25="● Erlik: mortos/submundo em tradições altaicas",
        A26="● Culto ancestral e espíritos de linhagem",
        A33="≈ Umay, sulde e espíritos tutelares",
        A35="● Kam/böö e viagem ritual; tradições diversas",
        A39="≈ Kut/süne/süld: conceitos vitais distintos",
        A43="● Oferendas, libações e mediação xamânica",
    ),
    "yoruba": mapped(
        A01="◇ Àṣẹ e ordem de Ọ̀run/Ayé",
        A02="● Olódùmarè; Obàtálá molda corpos em muitos relatos",
        A03="● Olódùmarè/Ọlọ́run",
        A04="◇ Ìwà-pẹ̀lẹ́ e ordem de Ifá",
        A05="● Onílẹ̀/Ilẹ̀; Odùduwà em tradições específicas",
        A06="● Ṣàngó: trovão e realeza",
        A07="≈ Ọ̀rún: Sol",
        A08="≈ Oṣù: Lua",
        A10="● Yemoja e Olókun: águas, oceano e profundidade",
        A11="● Ògún: ferro, forja, caminhos técnicos",
        A12="≈ Ọya: ventos e tempestades",
        A13="● Ọ̀rìṣà Oko: agricultura",
        A14="● Ọ̀ṣọ́ọ̀sì: caça e floresta",
        A15="● Ọ̀ṣun: beleza, amor, águas doces e fecundidade",
        A16="● Yemoja/Ọ̀ṣun: maternidade e parto",
        A17="● Ògún e Ṣàngó",
        A18="● Ṣàngó e Odùduwà: realeza e legitimidade",
        A19="● Ọ̀rúnmìlà/Ifá: sabedoria e divinação",
        A20="● Ògún: técnica, metal e trabalho",
        A21="● Èṣù: comunicação, encruzilhada e ambiguidade",
        A22="● Ajé e Ọ̀ṣun: riqueza/prosperidade",
        A23="● Ọ̀sányìn e Ọbaluayé: folhas e cura",
        A24="● Ọbaluayé/Ṣọ̀pọ̀na: varíola, doença e cura",
        A25="● Ikú; Ọya e passagem ao mundo ancestral",
        A26="● Egúngún e ancestrais",
        A27="≈ Èṣù e Ọya em funções de passagem",
        A28="● Ṣàngó/Ògún e juramento; Ifá para ordem ética",
        A29="● Orí e ìpín: destino pessoal",
        A30="≈ Àjàlá/forças de ruptura; sem 'caos' único",
        A32="● Retorno ancestral e continuidade por linhagem",
        A33="● Òrìṣà tutelares e Orí",
        A34="● Ọ̀ṣun, Ṣàngó e culto: música, dança e poesia",
        A35="● Possessão ritual e transe; substâncias não são universais",
        A38="● Ifá, odù e sacerdócio",
        A39="● Orí, ẹ̀mí e componentes da pessoa",
        A40="◇ Ọ̀run/Ayé e equilíbrio relacional",
        A43="● Ẹbọ, oferenda e mediação",
        A44="● Iniciação e incorporação ao culto",
    ),
    "akan": mapped(
        A01="◇ Nyame/Nyankopɔn e ordem do universo",
        A02="● Nyame/Odomankoma: criação",
        A03="● Nyame/Nyankopɔn",
        A05="● Asase Yaa: terra e fertilidade",
        A06="≈ Tano e abosom ligados a rios/tempo",
        A07="≈ Nyame com associações celestes/solares",
        A10="● Tano e espíritos de rios",
        A13="● Asase Yaa e agricultura",
        A14="≈ Espíritos de floresta e caça",
        A21="● Ananse: trickster, narrativa e sabedoria ambígua",
        A23="● Abosom e especialistas akom em cura",
        A25="≈ Asamando: mundo dos mortos",
        A26="● Nsamanfo: ancestrais",
        A28="● Asase Yaa em juramentos; ordem ancestral",
        A33="● Abosom tutelares",
        A34="● Tambores, dança e narrativa",
        A35="● Possessão akom",
        A39="● Kra, sunsum e componentes da pessoa",
        A43="● Libação, oferenda e sacerdócio",
    ),
    "vodun": mapped(
        A01="● Nana Buluku: anterior ao par Mawu-Lisa em tradições Fon",
        A02="● Mawu-Lisa: criação/ordenação",
        A03="● Mawu-Lisa",
        A04="● Fá: ordem/divinação",
        A05="● Sakpata e divindades da terra",
        A06="● Hevioso: trovão",
        A07="● Lisa: polo solar",
        A08="● Mawu: polo lunar",
        A10="● Agbe e voduns das águas",
        A11="● Gu: ferro e guerra",
        A13="≈ Sakpata e voduns agrários",
        A15="≈ Mawu e voduns locais de fecundidade",
        A17="● Gu e Hevioso",
        A19="● Fá: divinação e conhecimento",
        A21="● Legba: limiar e comunicação",
        A23="● Sakpata: doença e cura",
        A24="● Sakpata: varíola/doença",
        A25="≈ Avlekete e entidades funerárias; varia",
        A26="● Tohossou e ancestrais de linhagem",
        A27="≈ Legba: mediação entre domínios",
        A33="● Voduns tutelares de família/localidade",
        A34="● Tambores, canto e dança de culto",
        A35="● Possessão ritual",
        A39="≈ Sè e componentes espirituais; terminologia varia",
        A40="● Mawu/Lisa; Dan/Ayido Hwedo: complementaridade",
        A43="● Oferenda, iniciação e sacerdócio",
    ),
    "sumerian": mapped(
        A01="● Namma/Nammu e o mar primordial; tradições variam",
        A02="● Enki, Namma e Ninmaḫ em relatos antropogônicos",
        A03="● An e Enlil: céu e soberania",
        A04="● Me e ordem distribuída pelos deuses",
        A05="● Ki/Ninḫursaĝa: terra, nascimento e montanha",
        A06="● Iškur: tempestade",
        A07="● Utu: Sol e justiça",
        A08="● Nanna-Suen: Lua",
        A09="● Inana: Vênus",
        A10="● Enki: águas subterrâneas; Namma: águas primordiais",
        A11="● Gibil: fogo; Nuska: fogo/luz",
        A12="≈ Enlil: vento/ar e autoridade",
        A13="● Ezina/Ašnan e Dumuzi: grão e pastoreio",
        A14="● Šakkan e Ninurta: animais, caça e lavoura",
        A15="● Inana: desejo, sexualidade e poder",
        A16="● Ninḫursaĝa, Nintu e Bau: nascimento/maternidade",
        A17="● Inana e Ninurta: guerra",
        A18="● Enlil e Inana: realeza e legitimação",
        A19="● Enki, Nisaba e Nabu tardio: sabedoria/escrita",
        A20="● Enki e Kulla: ofícios e construção",
        A21="● Ninšubur e Papsukkal: mensageiros",
        A22="≈ Nanshe: prosperidade, pesca e justiça",
        A23="● Gula/Ninisina/Bau: cura",
        A24="● Nergal e Erra: peste e guerra",
        A25="● Ereškigal e Nergal: submundo",
        A26="● Ki-sikil-líl-lá? Culto aos mortos e ancestrais familiares",
        A27="● Ninšubur e Namtar: passagem/mensagem entre mundos",
        A28="● Utu: justiça; julgamento no submundo",
        A29="● Namtar: destino; Uttu: tempo/tecelagem em alguns textos",
        A30="● Asag, Anzû e forças de ruptura",
        A31="≈ Demônios hostis; não há princípio único do mal",
        A32="● Dumuzi e Inana: ciclo, descida e retorno parcial",
        A33="● Lamma e deidades tutelares pessoais",
        A34="● Inana e Nisaba: canto, poesia e escrita",
        A35="● Inana: inversão e êxtase ritual; cerveja ligada a Ninkasi",
        A38="● Enmeduranki e tradição divinatória",
        A39="≈ Zi e gidim: vida e espírito do morto",
        A40="◇ Céu/terra, cidade/estepe e redes de complementaridade",
        A43="● Culto templário, oferendas e sacerdócio",
        A44="● Inana e ritos de mudança de status",
    ),
    "akkadian": mapped(
        A01="● Apsû e Tiāmat: águas primordiais no Enūma Eliš",
        A02="● Marduk organiza o cosmos; Ea participa da criação humana",
        A03="● Anu, Enlil e depois Marduk/Aššur",
        A04="● Šamaš e os decretos divinos",
        A05="● Ki/Bēlet-ilī e Ištar em funções de fertilidade",
        A06="● Adad: tempestade",
        A07="● Šamaš: Sol e justiça",
        A08="● Sîn: Lua",
        A09="● Ištar: Vênus",
        A10="● Ea: Apsû e sabedoria; Tiāmat: mar primordial",
        A11="● Girra/Gibil: fogo; Nusku: fogo/luz",
        A12="≈ Enlil: vento e soberania",
        A13="● Dagan e Tammuz: grão/pastoreio",
        A15="● Ištar: amor, sexualidade e guerra",
        A16="● Bēlet-ilī/Mami: parto",
        A17="● Ištar, Nergal e Ninurta",
        A18="● Marduk/Aššur e Šamaš: realeza/lei",
        A19="● Ea e Nabû: sabedoria e escrita",
        A20="● Ea e Kulla: ofícios",
        A21="● Papsukkal: mensageiro; Ea em astúcia",
        A22="≈ Marduk e Ištar em prosperidade",
        A23="● Gula: cura",
        A24="● Nergal/Erra: peste",
        A25="● Ereškigal e Nergal",
        A27="● Namtar e mensageiros do submundo",
        A28="● Šamaš: justiça e juramento",
        A29="● Šīmtu/Namtar e Tábuas do Destino",
        A30="● Tiāmat e forças do caos no Enūma Eliš",
        A31="≈ Demônios como Lamaštu/Pazuzu; não mal absoluto",
        A32="● Tammuz e descidas de Ištar",
        A33="● Lamassu, apkallu e deuses pessoais",
        A34="● Ištar e Nabû: artes/escrita",
        A35="≈ Ištar, cerveja e festivais; especificar o rito",
        A38="● Adivinhação, profecia e revelação régia",
        A39="≈ Napīštu/eṭemmu: vida e espírito do morto",
        A43="● Templo, oferenda e sacerdócio",
    ),
    "egyptian": mapped(
        A01="● Nun/Nunet: águas primordiais; múltiplas cosmogonias",
        A02="● Atum, Ptah, Amon-Rá ou Khnum conforme o centro cultual",
        A03="● Rá/Amon-Rá; Hórus na realeza",
        A04="● Maat: verdade, ordem e justiça",
        A05="● Ísis, Hathor e Renenutet; Geb é a terra masculina",
        A06="● Nut (céu) e Seth (tempestade/deserto)",
        A07="● Rá, Aten e Khepri: Sol",
        A08="● Khonsu, Thoth e Iah: Lua",
        A09="● Sopdet: Sírius e cheia; Nut: estrelas",
        A10="● Nun, Hapi e Sobek: águas/Nilo",
        A11="● Sekhmet e Rá: fogo solar; Ptah: ofícios",
        A12="● Shu: ar",
        A13="● Osíris, Renenutet e Min: vegetação/fecundidade",
        A14="● Neith e Wepwawet: caça; deuses animais múltiplos",
        A15="● Hathor e Ísis: amor, sexualidade e beleza",
        A16="● Taweret, Bes, Hathor e Ísis: parto/família",
        A17="● Sekhmet, Montu e Seth: guerra/força",
        A18="● Hórus e Rá: realeza divina",
        A19="● Thoth e Seshat: escrita, cálculo e sabedoria",
        A20="● Ptah e Khnum: artesanato e modelagem",
        A21="● Thoth: mensageiro/mediador; Bes é liminar",
        A22="● Renenutet e Hapi: abundância",
        A23="● Sekhmet, Ísis, Thoth e Serqet: cura",
        A24="● Sekhmet e demônios da doença",
        A25="● Osíris e Anúbis: morte/necrópole",
        A26="● Akhu e culto funerário aos mortos eficazes",
        A27="● Anúbis e Wepwawet: guias",
        A28="● Osíris, Maat, Anúbis e Thoth: julgamento",
        A29="● Shai: destino; Heh: eternidade",
        A30="● Apófis e Seth em funções destrutivas distintas",
        A31="≈ Apófis: inimigo da ordem; não equivalente a diabo",
        A32="● Osíris, Khepri e ciclo solar: renovação",
        A33="● Bes, Taweret, Bastet e amuletos tutelares",
        A34="● Hathor, Bes e Ihy: música/dança",
        A35="≈ Hathor e festivais de embriaguez; ritual específico",
        A36="≈ Pureza e disciplina sacerdotal; não monasticismo",
        A38="● Oráculos e fala divina; faraó como mediador",
        A39="● Ka, ba, akh, ib e ren: componentes da pessoa",
        A40="◇ Maat/isfet: ordem versus desordem",
        A41="● Vida transfigurada entre os akhu e com Osíris/Rá",
        A42="≈ Renovação cíclica; não fim linear universal",
        A43="● Culto templário, oferendas e rito funerário",
        A44="● Mumificação e transfiguração em akh",
    ),
    "ugaritic": mapped(
        A01="● El e Asherah/Athirat no fundo cosmogônico; dados fragmentários",
        A02="≈ El como progenitor/criador",
        A03="● El: chefe; Baal conquista soberania ativa",
        A04="≈ El e assembleia divina",
        A05="● Athirat/Asherah: maternidade divina",
        A06="● Baal Hadad: tempestade e chuva",
        A07="● Shapash: Sol",
        A08="● Yarikh: Lua",
        A09="● Athtar: Vênus",
        A10="● Yamm: mar; Baal combate Yamm",
        A13="● Baal: chuva/fertilidade; Dagon: grão",
        A15="● Astarte e Anat: sexualidade/guerra em funções distintas",
        A16="● Athirat: mãe dos deuses",
        A17="● Anat e Baal",
        A18="● El e Baal: realeza",
        A21="≈ Kothar-wa-Khasis viaja e comunica; mensageiros divinos",
        A20="● Kothar-wa-Khasis: artesanato e técnica",
        A25="● Mot: morte/submundo",
        A28="≈ El como juiz; juramentos cultuais",
        A30="● Yamm e Lotan; Mot como forças adversas",
        A31="≈ Lotan/Yamm/Mot são oponentes, não mal absoluto",
        A32="≈ Retorno de Baal: ciclo de ausência e restauração",
        A43="● Sacrifícios e culto templário",
    ),
    "hittite": mapped(
        A01="≈ Alalu/Anu e sucessões do ciclo Kumarbi (tradição hurrita)",
        A02="≈ Deuses formadores múltiplos; corpus sincrético",
        A03="● Deus da Tempestade de Hatti/Tarḫunna/Teššub",
        A04="● Divindades solares e juramentos",
        A05="● Deusa-Sol de Arinna e Ḫannahanna",
        A06="● Tarḫunna/Teššub: tempestade",
        A07="● Deusa-Sol de Arinna; Ištanu",
        A08="● Kušu/Arma: Lua",
        A10="≈ Aruna: mar",
        A11="≈ Ḫaššamili e divindades do fogo/lar",
        A13="● Telipinu: vegetação e desaparecimento sazonal",
        A15="● Šauška: amor e guerra",
        A17="● Teššub e Šauška",
        A18="● Deuses solares e da tempestade legitimam o rei",
        A19="≈ Ea e Kamrušepa: sabedoria/magia ritual",
        A25="● Allani e Lelwani: mundo inferior",
        A28="● Išḫara e deuses do juramento",
        A30="● Illuyanka e Ḫedammu: monstros de ruptura",
        A32="● Telipinu: retorno e restauração",
        A33="● Deidades tutelares (LAMMA)",
        A43="● Festivais, oferendas e tratados juramentados",
        A44="● Ritos de purificação/substituição",
    ),
    "greek": mapped(
        A01="● Chaos, Gaia, Tartaros e Eros em Hesíodo",
        A02="≈ Gaia e gerações divinas; Prometeu em antropogonia",
        A03="● Zeus",
        A04="● Themis, Dike e Moirai",
        A05="● Gaia e Deméter",
        A06="● Zeus: céu, raio e tempestade",
        A07="● Hélios e Apolo em associações posteriores",
        A08="● Selene e Ártemis em associações lunares posteriores",
        A09="● Eos e Afrodite/Héspero em associações planetárias",
        A10="● Poseidon, Okeanos e divindades fluviais",
        A11="● Héstia e Hefesto",
        A12="● Éolo e Anemoi",
        A13="● Deméter, Perséfone e Dioniso",
        A14="● Ártemis e Pan",
        A15="● Afrodite e Eros",
        A16="● Hera, Ilítia e Héstia",
        A17="● Ares e Atena; heróis",
        A18="● Zeus, Themis e Atena Polias",
        A19="● Atena, Apolo e Musas",
        A20="● Hefesto, Atena Ergane e Dédalo",
        A21="● Hermes e Prometeu",
        A22="● Hermes, Ploutos e Tyche",
        A23="● Asclépio, Higieia, Apolo e Quíron",
        A24="● Apolo envia/cessa peste; Nosoi personificam doenças",
        A25="● Hades e Perséfone",
        A26="● Heróis e mortos cultuados",
        A27="● Hermes Psicopompo e Caronte",
        A28="● Zeus Horkios, Themis e juízes do Hades",
        A29="● Moirai, Ananke e Chronos (tradições distintas)",
        A30="● Chaos, Typhon e Gigantes",
        A31="≈ Eris e monstros adversos; sem mal absoluto",
        A32="● Perséfone, Dioniso e mistérios; sentidos distintos",
        A33="● Atena, Apolo e daimon tutelar conforme a pólis",
        A34="● Apolo, Musas e Dioniso",
        A35="● Dioniso e êxtase báquico",
        A36="≈ Órficos, cínicos e filósofos ascéticos; não pan-helênico",
        A37="≈ Eleos: misericórdia personificada; ética varia",
        A38="● Apolo/oráculos, mantis e poetas inspirados",
        A39="● Psychē, thymos e pneuma: conceitos históricos distintos",
        A40="◇ Apolíneo/dionisíaco é leitura moderna; pares divinos variam",
        A41="≈ Mistérios e filosofias oferecem destinos pós-morte diversos",
        A42="≈ Ciclos cósmicos filosóficos; não doutrina única",
        A43="● Sacrifício, libação, mistérios e sacerdócio cívico",
        A44="● Dioniso, mistérios e metamorfoses míticas",
    ),
    "roman": mapped(
        A01="≈ Caos/Janus em poetas; religião romana não tem uma cosmogonia única",
        A02="≈ Jupiter/Janus e ordenação; empréstimos gregos",
        A03="● Jupiter Optimus Maximus",
        A04="● Fides, Iustitia, Pax e ordem cívica",
        A05="● Terra Mater, Ceres e Ops",
        A06="● Jupiter",
        A07="● Sol/​Sol Invictus",
        A08="● Luna e Diana",
        A09="● Aurora e Venus",
        A10="● Neptune, Oceanus e divindades fluviais",
        A11="● Vesta e Vulcan",
        A12="● Venti; Jupiter como tempo atmosférico",
        A13="● Ceres, Liber e Saturn",
        A14="● Diana e Silvanus",
        A15="● Venus, Cupid e Flora",
        A16="● Juno, Lucina e Lares",
        A17="● Mars, Bellona e Victoria",
        A18="● Jupiter, Mars, Quirinus e culto imperial",
        A19="● Minerva, Apollo e Mercury",
        A20="● Vulcan e Minerva",
        A21="● Mercury, Janus e Terminus",
        A22="● Mercury, Fortuna e Abundantia",
        A23="● Aesculapius, Salus e Apollo",
        A24="● Apollo e Febris",
        A25="● Dis Pater, Orcus e Proserpina",
        A26="● Manes, Lares e culto ancestral",
        A27="● Mercury; Charon por recepção grega",
        A28="● Jupiter, Fides e Iustitia",
        A29="● Parcae, Fortuna e Saturn",
        A30="≈ Discordia, monstros e forças destrutivas",
        A32="≈ Ciclos agrários e cultos de mistério",
        A33="● Lares, Penates e Genius",
        A34="● Apollo, Muses e Liber",
        A35="● Liber/Bacchus e Saturnalia",
        A38="● Sibila, augúrio e oráculos",
        A39="● Anima, animus, genius e manes",
        A41="≈ Mistérios e filosofias; não soteriologia romana única",
        A43="● Sacrifício cívico, voto e sacerdócios",
        A44="● Ritos de passagem e metamorfoses literárias",
    ),
    "celtic": mapped(
        A01="? Cosmogonia pré-cristã não preservada de modo sistemático",
        A02="≈ Dagda e figuras ancestrais; não criador universal",
        A03="≈ Lugus/Lugh, Taranis e deuses tribais; sem chefe pan-céltico",
        A05="● Matres/Matronae, Danu/Anu e deusas territoriais",
        A06="● Taranis e divindades da tempestade",
        A07="● Belenos e Grannus em associações solares",
        A10="● Manannán, Nodens e divindades de rios/fontes",
        A11="● Brigid/Brigantia: fogo, ofício e poesia; Sucellos em funções",
        A13="● Cernunnos, Sucellos e deusas da terra; funções regionais",
        A14="● Cernunnos e deuses locais de caça/animais",
        A15="● Áine e divindades locais; documentação tardia",
        A17="● Morrígan, Camulos e heroes",
        A18="● Lug/Lugus e soberania feminina em narrativas irlandesas",
        A19="● Ogma/Ogmios, Brigid e druidas",
        A20="● Goibniu/Gobannos: forja",
        A21="● Lug e figuras metamórficas; trickster não é categoria nativa",
        A23="● Brigid, Grannus, Belenos e fontes de cura",
        A25="● Donn e Outro Mundo; tradições insulares",
        A26="● Ancestrais tribais e sídhe em recepção medieval",
        A27="≈ Manannán e figuras de travessia",
        A29="≈ Morrígan/profecia e ciclos sazonais",
        A32="● Ciclos sazonais e narrativas de retorno",
        A33="● Deidades locais, Matres e gênios tutelares",
        A34="● Brigid, Ogma e bardos",
        A35="≈ Inspiração poética e festas; enteógenos não demonstrados",
        A38="● Druidas, vates e profecia",
        A39="≈ Alma/Outro Mundo em fontes clássicas e medievais",
        A43="● Oferendas votivas e sacerdócio druídico; fontes externas",
        A44="● Metamorfose em ciclos insulares",
    ),
    "norse": mapped(
        A01="● Ginnungagap, fogo e gelo",
        A02="● Óðinn, Vili e Vé formam o mundo e humanos",
        A03="● Óðinn; Þórr como protetor popular",
        A04="● Týr, Forseti e ordem das coisas/þing",
        A05="● Jörð, Nerthus e Freyr/Freyja",
        A06="● Þórr: trovão",
        A07="● Sól",
        A08="● Máni",
        A09="● Dagr, Nótt e estrelas; funções limitadas",
        A10="● Njörðr, Ægir e Rán",
        A11="● Logi (personificação narrativa); Loki não é deus do fogo",
        A12="≈ Óðinn: sopro; Kári em tradição tardia",
        A13="● Freyr e Gefjon",
        A14="● Skaði e Ullr: caça; Freyr: animais/fecundidade",
        A15="● Freyja e Freyr",
        A16="● Frigg e Eir/​dísir em proteção familiar",
        A17="● Týr, Þórr, Óðinn e valquírias",
        A18="● Óðinn, Freyr e sacralidade régia",
        A19="● Óðinn, Mímir e Kvasir",
        A20="● Völundr e anões ferreiros",
        A21="● Loki e Hermóðr em funções distintas",
        A22="● Njörðr e Freyr: prosperidade",
        A23="● Eir: cura",
        A24="≈ Hel e seres de doença; fontes escassas",
        A25="● Hel; Óðinn/Valhöll para alguns mortos",
        A26="● Dísir, álfar e túmulos ancestrais",
        A27="● Valquírias e Hermóðr",
        A28="● Týr e Forseti; julgamento pós-morte não central",
        A29="● Nornir e Urðr",
        A30="● Jötnar, Surtr e forças de ruptura",
        A31="≈ Loki/Fenrir como adversários no Ragnarök; não diabo",
        A32="● Baldr e mundo renovado após Ragnarök",
        A33="● Þórr, dísir e landvættir",
        A34="● Bragi, Óðinn e poesia",
        A35="● Seiðr e fúria guerreira; substâncias não demonstradas",
        A38="● Völur e runas/poesia inspirada",
        A39="● Hugr, fylgja, hamingja e hamr: componentes distintos",
        A42="● Ragnarök e renovação",
        A43="● Blót, sumbl e mediação cultual",
        A44="● Seiðr, iniciação poética e metamorfose",
    ),
    "finnic": mapped(
        A01="≈ Ovo cósmico e águas no Kalevala; compilação do séc. XIX",
        A02="≈ Ilmatar/Väinämöinen em versões kaleválicas",
        A03="● Ukko: céu e autoridade",
        A05="● Akka/Maa-emo: terra e maternidade",
        A06="● Ukko: trovão",
        A07="● Päivätär: Sol",
        A08="● Kuutar: Lua",
        A10="● Ahti/Vellamo: águas",
        A11="● Ilmarinen: forja; fogo em narrativas de origem",
        A12="● Ilmatar: ar",
        A13="● Pellonpekko: cultivo",
        A14="● Tapio e Mielikki: floresta/caça",
        A15="≈ Lemminkäinen e divindades femininas; corpus tardio",
        A17="≈ Ukko e heróis épicos",
        A19="● Väinämöinen: sabedoria/canto",
        A20="● Ilmarinen: artesanato",
        A21="≈ Lemminkäinen e figuras liminares",
        A22="● Ahti e espíritos domésticos em prosperidade",
        A23="● Väinämöinen e tietäjä: cura por canto",
        A25="● Tuoni/Tuonela",
        A26="● Väki, haltija e ancestrais",
        A27="≈ Väinämöinen/Lemminkäinen em viagens a Tuonela",
        A29="≈ Louhi e destino em épica compilada",
        A30="● Louhi e forças de Pohjola; não mal absoluto",
        A33="● Haltija e espíritos tutelares",
        A34="● Väinämöinen, kantele e poesia rúnica",
        A35="● Tietäjä e êxtase/canto ritual",
        A38="● Runos, encantamentos e especialistas",
        A39="● Henki, itse, luonto e väki: conceitos distintos",
        A43="● Oferendas a haltija e ancestrais",
        A44="● Metamorfose épica e iniciação do especialista",
    ),
    "slavic": mapped(
        A01="? Cosmogonias reconstruídas de fontes tardias e folclore",
        A02="≈ Rod/Svarog em reconstruções; alcance histórico discutido",
        A03="● Perun em cultos principescos; não universal",
        A05="● Mokoš: terra, trabalho feminino e fecundidade",
        A06="● Perun: trovão",
        A07="● Dažbog/Khors em associações solares discutidas",
        A10="● Veles e espíritos das águas; varia",
        A11="● Svarog/Svarožič: fogo em fontes medievais",
        A13="● Jarilo e Mokoš em folclore sazonal",
        A14="● Veles: gado, floresta e riqueza",
        A17="● Perun",
        A18="● Perun e culto dinástico de Kyiv",
        A19="≈ Veles e especialistas mágicos",
        A21="● Veles: limiar/oposição a Perun; não 'trickster' simples",
        A22="● Veles: gado/riqueza",
        A25="● Veles e Nav: mortos",
        A26="● Dziady e culto ancestral",
        A27="≈ Veles e guias folclóricos",
        A29="● Rodzanice/Sudjenice: destino em folclore",
        A30="≈ Chernobog em fonte única; não princípio seguro do caos",
        A31="? Chernobog: historicidade/função muito debatida",
        A32="● Jarilo e ciclos sazonais no folclore",
        A33="● Domovoi e espíritos domésticos",
        A34="● Canto ritual e épica",
        A35="≈ Ritos sazonais; enteógenos não demonstrados",
        A38="≈ Volkhvy e adivinhação em fontes cristãs",
        A39="≈ Duša e múltiplas concepções folclóricas",
        A40="≈ Perun/Veles como oposição estrutural: reconstrução acadêmica",
        A43="● Oferendas e festas sazonais; documentação desigual",
    ),
    "baltic": mapped(
        A01="● Dievas/Dievs e destino cósmico; dainas e fontes tardias",
        A02="≈ Dievas/Dievs; cosmogonias variam",
        A03="● Dievas/Dievs",
        A05="● Žemyna/Zemes māte: terra",
        A06="● Perkūnas/Pērkons: trovão",
        A07="● Saulė: Sol",
        A08="● Mėnulis/Mēness: Lua",
        A09="● Aušrinė/Auseklis: aurora/Vênus",
        A10="● Jūratė/Ūsiņš e mães das águas; varia",
        A11="● Gabija/Uguns māte: fogo doméstico",
        A13="● Žemyna e Jumis: colheita",
        A14="● Medeina/Meža māte: floresta/caça",
        A15="● Laima e Saule em casamento/fecundidade; funções distintas",
        A16="● Laima: parto e destino",
        A17="● Perkūnas",
        A19="≈ Dievas e Laima; sabedoria em dainas",
        A21="≈ Velnias: limiar e astúcia no folclore cristianizado",
        A22="≈ Jumis e Aitvaras: prosperidade",
        A23="≈ Laima e deusas locais; cura folclórica",
        A25="● Velnias/Veles e mundo dos mortos; camadas tardias",
        A26="● Vėlės e culto ancestral",
        A29="● Laima/Dekla/Kārta: destino",
        A30="≈ Velnias e forças ctônicas; não diabo original simples",
        A32="● Saulė e festivais sazonais",
        A33="● Mātes e espíritos domésticos",
        A34="● Dainas e canto ritual",
        A38="≈ Vaidelutės/krīvi em fontes históricas problemáticas",
        A39="≈ Siela/vėlė e dvēsele: alma/morto",
        A43="● Oferendas domésticas e sazonais",
    ),
    "zoroastrian": mapped(
        A01="● Ahura Mazdā e criação boa; cosmogonia sistemática é sobretudo pahlavi",
        A02="● Ahura Mazdā/Ohrmazd",
        A03="● Ahura Mazdā",
        A04="● Aša: verdade e ordem",
        A05="● Spənta Ārmaiti: terra, devoção e fecundidade",
        A06="● Vayu e Tištrya; tempestade/chuva em funções distintas",
        A07="● Hvar Xšaēta e luz de Ahura Mazdā",
        A08="● Māh: Lua",
        A09="● Tištrya: estrela Sírius/chuva",
        A10="● Arədvī Sūrā Anāhitā e Apąm Napāt",
        A11="● Ātar: fogo",
        A12="● Vayu: vento/atmosfera",
        A13="● Amərətāt: plantas; Tištrya: chuva",
        A14="● Vohu Manah e criação animal",
        A15="≈ Anāhitā: fecundidade; amor não é seu único domínio",
        A17="● Vərəθraγna: vitória",
        A18="● Xšaθra Vairya e xwarrah: poder legítimo",
        A19="● Vohu Manah e Ahura Mazdā: boa mente/sabedoria",
        A21="● Sraoša: mensagem, obediência e guia",
        A22="≈ Aši: recompensa/fortuna",
        A23="● Airyaman e plantas de cura",
        A24="● Angra Mainyu e daēvas produzem doença",
        A25="● Astō.wiδātu e mundo dos mortos",
        A26="● Fravašis: espíritos protetores/ancestrais",
        A27="● Sraoša e Daēnā na ponte Činvat",
        A28="● Mithra, Rašnu e Sraoša julgam na Činvat",
        A29="● Zurvān em vertente zurvanita; não ortodoxia geral",
        A30="● Angra Mainyu e contracriação",
        A31="● Angra Mainyu/Ahriman",
        A32="● Frašō.kərəti e ressurreição",
        A33="● Fravašis e Sraoša",
        A36="● Pureza, disciplina e escolha ética",
        A37="◇ Spənta Mainyu e boa ação",
        A38="● Zaraθuštra e revelação gática",
        A39="● Urvan, daēnā e fravaši: aspectos distintos",
        A40="● Aša/druj; dualismo ético-cósmico com história complexa",
        A41="● Travessia da Činvat e participação na renovação",
        A42="● Frašō.kərəti, julgamento e renovação final",
        A43="● Yasna, fogo e sacerdócio",
        A44="● Purificação e renovação escatológica",
    ),
    "vedic": mapped(
        A01="● Nāsadīya Sūkta: indeterminação; Hiraṇyagarbha/Puruṣa em outros hinos",
        A02="● Prajāpati, Viśvakarman e Puruṣa em estratos distintos",
        A03="● Indra/Varuṇa; não há chefe imutável",
        A04="● Ṛta e Varuṇa",
        A05="● Pṛthivī e Aditi",
        A06="● Indra, Parjanya e Maruts",
        A07="● Sūrya, Savitṛ e Mitra",
        A08="● Soma/Candra em desenvolvimento posterior",
        A09="● Uṣas e Aśvins",
        A10="● Varuṇa, Āpas e Sarasvatī",
        A11="● Agni",
        A12="● Vāyu e prāṇa",
        A13="● Pṛthivī, Parjanya e Pūṣan",
        A14="● Rudra, Pūṣan e Aśvins; gado/selvagem",
        A15="● Kāma e Uṣas; funções eróticas diversas",
        A16="● Aditi, Sinīvālī e Aśvins",
        A17="● Indra e Maruts",
        A18="● Indra, Varuṇa e Rājasūya",
        A19="● Bṛhaspati, Sarasvatī e Soma",
        A20="● Tvaṣṭṛ e Ṛbhus",
        A21="● Agni como mensageiro; Pūṣan como guia",
        A22="● Bhaga e Pūṣan",
        A23="● Aśvins e Rudra",
        A24="● Rudra e Nirṛti em aflição; também curam/protegem",
        A25="● Yama e reino dos ancestrais",
        A26="● Pitṛs",
        A27="● Pūṣan e Agni",
        A28="● Varuṇa, Mitra e Yama",
        A29="≈ Ṛta, Kāla em textos posteriores e destino ritual",
        A30="● Vṛtra e forças que bloqueiam as águas",
        A31="≈ Dasyus/asuras em contextos históricos; não mal absoluto",
        A32="● Soma e ciclo sacrificial; renascimento se desenvolve nos Upaniṣads",
        A33="● Indra, Agni e deuses tutelares",
        A34="● Sarasvatī, Gandharvas e canto sāmico",
        A35="● Soma ritual; identidade botânica é debatida",
        A36="≈ Tapas e ascese em estratos védicos tardios",
        A38="● Ṛṣis e śruti",
        A39="● Ātman e prāṇa sobretudo em textos védicos tardios",
        A40="◇ Ṛta e pares complementares céu/terra",
        A41="≈ Svarga; mokṣa emerge nos Upaniṣads",
        A43="● Yajña, Agni e sacerdócio",
        A44="● Iniciação sacrificial e transformação pelo rito",
    ),
    "hindu": mapped(
        A01="● Brahman; Nārāyaṇa, Śiva ou Devī conforme a escola",
        A02="● Brahmā/Viṣṇu/Śiva ou Devī; cosmogonias purânicas múltiplas",
        A03="● Īśvara em formas sectárias: Viṣṇu, Śiva, Devī etc.",
        A04="● Dharma e ṛta",
        A05="● Devī, Pṛthivī, Lakṣmī e Pārvatī",
        A06="● Indra, Rudra/Śiva e divindades atmosféricas",
        A07="● Sūrya",
        A08="● Candra/Soma",
        A09="● Uṣas, Śukra e Navagraha",
        A10="● Varuṇa, Gaṅgā e divindades fluviais",
        A11="● Agni; Śiva e Kārttikeya em fogo",
        A12="● Vāyu e prāṇa",
        A13="● Annapūrṇā, Lakṣmī, Balarāma e deuses locais",
        A14="● Paśupati/Śiva, Durgā e divindades de aldeia",
        A15="● Kāma, Rati, Kṛṣṇa e Rādhā",
        A16="● Pārvatī, Gaurī, Śaṣṭhī e Gaṇeśa",
        A17="● Durgā, Skanda e Hanumān",
        A18="● Viṣṇu/Rāma e Śiva; rājadharma",
        A19="● Sarasvatī, Gaṇeśa e Dakṣiṇāmūrti",
        A20="● Viśvakarman e Tvaṣṭṛ",
        A21="● Nārada, Hanumān e Kṛṣṇa em funções distintas",
        A22="● Lakṣmī, Kubera e Gaṇeśa",
        A23="● Dhanvantari, Aśvins e Śītalā",
        A24="● Śītalā, Māriyamman e divindades de doença/cura",
        A25="● Yama, Kālī e Śiva",
        A26="● Pitṛs e culto ancestral",
        A27="● Yama-dūtas; Agni e Hanumān como mediadores",
        A28="● Yama, Dharma e Citragupta",
        A29="● Kāla, karma e daiva",
        A30="● Kālī, Rudra e pralaya; não 'caos' simples",
        A31="≈ Asuras e Māra-like tempters; não princípio único do mal",
        A32="● Saṃsāra, avatāras e renovação cíclica",
        A33="● Gaṇeśa, Hanumān, Durgā e divindades tutelares",
        A34="● Sarasvatī, Kṛṣṇa, Śiva Naṭarāja e Gandharvas",
        A35="● Śiva, Kālī e tradições tântricas/bhakti; contexto específico",
        A36="● Śiva, sādhus, yoga e saṃnyāsa",
        A37="● Kṛṣṇa, Rāma, Devī e ahiṃsā/bhakti",
        A38="● Ṛṣis, avatāras, gurus e śruti",
        A39="● Ātman, jīva, prāṇa e manas: escolas divergem",
        A40="● Śiva/Śakti, Puruṣa/Prakṛti e outras complementaridades",
        A41="● Mokṣa por conhecimento, devoção, ação ou yoga",
        A42="● Yugas, pralaya e Kalki em tradições vaiṣṇavas",
        A43="● Pūjā, yajña, darśan e guru/sacerdócio",
        A44="● Dīkṣā, yoga, tantra e metamorfoses míticas",
    ),
    "jain": mapped(
        A01="◇ Universo eterno e incriado: loka/aloka",
        A02="— Sem deus criador; cosmos opera por sua própria natureza",
        A03="— Sem soberano divino criador; siddhas são libertos",
        A04="● Dharma, ahiṃsā e ordem kármica",
        A05="≈ Ambikā e yakṣiṇīs como tutelares; não criadoras",
        A07="◇ Jyotiṣka devas: seres celestes, não Sol soberano",
        A08="◇ Divindades astrais subordinadas",
        A13="◇ Vida vegetal possui jīva; ética de não violência",
        A14="◇ Animais possuem jīva; ahiṃsā radical",
        A16="≈ Yakṣiṇīs e deusas locais em proteção familiar",
        A17="◇ Vitória é conquista das paixões (jina)",
        A18="● Tīrthaṅkaras como mestres exemplares, não deuses criadores",
        A19="● Tīrthaṅkaras, kevala-jñāna e tradição canônica",
        A22="≈ Lakṣmī cultual em alguns contextos jainas; subordinada",
        A23="≈ Dharaṇendra/Padmāvatī e ritos tutelares; doutrina central é ética",
        A25="● Saṃsāra e destinos infernais/celestes",
        A26="≈ Memória dos tīrthaṅkaras e mestres; não culto ancestral central",
        A28="● Karma impessoal produz consequências",
        A29="● Kāla e ciclos cósmicos ascendentes/descendentes",
        A30="◇ Entropia moral dos ciclos; não divindade destruidora",
        A31="◇ Kaṣāyas e ignorância, não mal personificado absoluto",
        A32="● Renascimento condicionado pelo karma",
        A33="≈ Yakṣas/yakṣiṇīs tutelares",
        A36="● Ascese, jejum, monasticismo e sallekhanā sob disciplina",
        A37="● Ahiṃsā, anukampā e cuidado por seres vivos",
        A38="● Tīrthaṅkaras e mestres",
        A39="● Jīva: alma individual eterna",
        A40="● Jīva/ajīva; anekāntavāda evita absolutização de perspectivas",
        A41="● Mokṣa/kevala-jñāna",
        A42="● Ciclos cósmicos sem criação/fim absolutos",
        A43="● Pūjā e oferendas devocionais; sem sacrifício animal",
        A44="● Dīkṣā monástica e purificação kármica",
    ),
    "buddhist": mapped(
        A01="◇ Originação dependente; questões de começo absoluto são deixadas de lado",
        A02="— Sem deus criador; Brahmā é ser condicionado nos textos",
        A03="— Devas existem, mas nenhum é soberano eterno",
        A04="● Dharma e causalidade",
        A05="≈ Pṛthivī/Bhūmi como testemunha; não mãe criadora",
        A06="≈ Sakka e devas do tempo atmosférico, subordinados",
        A07="≈ Buda como 'sol' metafórico; deuses solares são condicionados",
        A08="≈ Lua em calendário e metáfora; não centro soteriológico",
        A10="≈ Nāgas e divindades das águas, subordinados",
        A11="◇ Fogo como elemento e metáfora das paixões",
        A12="◇ Vāyo-dhātu e respiração meditativa",
        A13="≈ Espíritos/devas locais; ética monástica regula plantas",
        A14="● Jātakas animais e ahiṃsā/compaixão",
        A15="◇ Kāma é desejo vinculante; não ideal divino",
        A16="≈ Hārītī em proteção de crianças; adoção regional",
        A17="◇ Māra e conquista interior; reis protetores no plano histórico",
        A18="● Cakkavatti e Aśoka como modelos de governo, não deuses",
        A19="● Buda, Dharma e saṅgha",
        A21="≈ Devas/mensageiros e narrativas de upāya; sem trickster central",
        A22="≈ Vasudhārā/Jambhala em tradições posteriores",
        A23="● Buda e medicina monástica; Bhaiṣajyaguru é mahāyāna",
        A24="◇ Doença como condição; causas naturais/kármicas em contextos",
        A25="● Māra/Māraṇa e reinos de renascimento; sem alma permanente",
        A26="≈ Mérito dedicado aos mortos e ancestrais; formas regionais",
        A27="≈ Kṣitigarbha é mahāyāna; monges guiam por ensinamento",
        A28="● Karma e Yama em cosmologias populares; processo não é juiz supremo",
        A29="● Karma, saṃsāra e ciclos cósmicos",
        A30="◇ Impermanência e dissolução dos mundos",
        A31="● Māra: obstáculo/tentador, não mal absoluto",
        A32="● Renascimento sem ātman",
        A33="● Quatro Reis Celestes e devas protetores",
        A34="≈ Canto e arte devocional; disciplina varia",
        A35="● Jhāna, meditação e ritual; intoxicantes são preceitualmente evitados",
        A36="● Saṅgha, renúncia e disciplina",
        A37="● Karuṇā e mettā",
        A38="● Buda como mestre desperto; não profeta de criador",
        A39="● Anattā: rejeição de alma permanente; fluxo condicionado",
        A40="● Caminho do Meio",
        A41="● Nibbāna/nirvāṇa",
        A42="≈ Maitreya e ciclos de budas; não fim final único",
        A43="● Dāna, pūjā e mérito; sacrifício violento rejeitado",
        A44="● Ordenação, meditação e despertar",
    ),
    "mahayana": mapped(
        A01="◇ Śūnyatā, talidade e dharmadhātu; escolas divergem",
        A02="— Sem criador; budas não criam o cosmos",
        A03="≈ Budas cósmicos, mas sem soberania criadora",
        A04="● Dharma, śūnyatā e originação dependente",
        A05="● Prajñāpāramitā: 'mãe dos budas'; Tārā em devoções",
        A06="≈ Vajrapāṇi e reis celestes protetores",
        A07="● Amitābha/Vairocana em funções de luz",
        A08="≈ Candraprabha e simbolismo lunar",
        A09="≈ Akṣobhya e constelações rituais em tradições específicas",
        A10="● Avalokiteśvara/Guanyin e nāgas em funções aquáticas regionais",
        A11="● Acala/Vajrapāṇi e fogo ritual em correntes esotéricas",
        A12="◇ Sopro/mente em meditação; mais explícito no vajrayāna",
        A13="≈ Bhūmi-devī e bodhisattvas locais",
        A14="● Jātakas e bodhisattvas em formas animais",
        A15="◇ Apego transformado por upāya; não divindade do amor",
        A16="● Guanyin/Kannon e Hārītī em parto/família",
        A17="● Vajrapāṇi e reis guardiões",
        A18="● Cakravartin e bodhisattva-governante como modelos",
        A19="● Mañjuśrī: sabedoria",
        A20="≈ Viśvakarman e bodhisattvas de ofícios em cultos locais",
        A21="● Avalokiteśvara e upāya; mensageiros/transformações",
        A22="● Vasudhārā e Jambhala",
        A23="● Bhaiṣajyaguru e Guanyin",
        A24="≈ Protetores e budas de medicina; doença não é deus único",
        A25="● Kṣitigarbha e Amitābha em destinos pós-morte",
        A26="● Ritos memoriais, transferência de mérito e ancestrais",
        A27="● Kṣitigarbha/Guanyin como guias",
        A28="● Karma; Yama em cosmologias populares",
        A29="● Kalpas, karma e saṃsāra",
        A30="◇ Śūnyatā não é caos; dissoluções cósmicas são cíclicas",
        A31="● Māra e obstáculos internos",
        A32="● Renascimento; votos de bodhisattva e terras puras",
        A33="● Avalokiteśvara, Tārā e dharmapālas",
        A34="● Sarasvatī/Benzaiten e artes devocionais",
        A35="● Meditação e visualização; intoxicantes geralmente preceitualmente evitados",
        A36="● Saṅgha e bodhisattva-caminho",
        A37="● Avalokiteśvara/Guanyin e karuṇā",
        A38="● Budas e bodhisattvas ensinam por múltiplos meios",
        A39="● Anattā e natureza búdica: interpretações não equivalentes a alma",
        A40="● Não dualidade, duas verdades e Caminho do Meio",
        A41="● Buddhahood/nirvāṇa e liberação de todos os seres",
        A42="● Maitreya e ciclos de budas",
        A43="● Pūjā, dāna, votos e ritos",
        A44="● Voto de bodhisattva, iniciações e transformação da mente",
    ),
    "vajrayana": mapped(
        A01="● Dharmakāya/ādibuddha em sistemas específicos; não criador",
        A02="— Sem deus criador; emanação não é criação teísta",
        A03="≈ Vairocana/Samantabhadra em mandalas; não soberano universal",
        A04="● Dharma, vacuidade e mandala",
        A05="● Tārā, Prajñāpāramitā e dākinīs",
        A06="● Vajrapāṇi e protetores atmosféricos",
        A07="● Amitābha/Vairocana e luz clara",
        A08="≈ Candraprabha e simbolismo lunar tântrico",
        A09="≈ Astrologia ritual tibetana; não figura única",
        A10="● Nāgas, Tārā e divindades locais das águas",
        A11="● Agni-homa, Acala e tummo em contextos distintos",
        A12="● Lung/prāṇa e práticas de canais",
        A13="≈ Protetores locais e ritos de prosperidade",
        A14="● Mahākāla, Siṃhamukhā e divindades zoomórficas",
        A15="● Kurukullā e yab-yum como símbolo tântrico",
        A16="● Tārā e protetoras do parto",
        A17="● Vajrapāṇi, Palden Lhamo e protetores",
        A18="● Cakravartin, reis do Dharma e Dalai Lamas historicamente",
        A19="● Mañjuśrī e gurus",
        A20="≈ Viśvakarman e artes rituais",
        A21="● Dākinīs, gurus e mensageiros",
        A22="● Jambhala e Vasudhārā",
        A23="● Bhaiṣajyaguru e Yuthok em tradição médica",
        A24="≈ Protetores e ritos contra obstáculos/doenças",
        A25="● Yama, Kṣitigarbha e bardos",
        A26="● Lamas, ancestrais de linhagem e transferência de mérito",
        A27="● Tārā, Avalokiteśvara e instruções de bardo",
        A28="● Karma e Yama; julgamento é imagético/ritual",
        A29="● Kālacakra, karma e ciclos",
        A30="● Mahākāla e divindades iradas transformam obstáculos",
        A31="● Māra e demônios subjugados; não mal absoluto",
        A32="● Tulku/renascimento e realização do corpo de buda",
        A33="● Dharmapālas, Tārā e protetores de linhagem",
        A34="● Mantra, dança cham, thangka e música ritual",
        A35="● Yoga tântrico e experiência visionária sob iniciação",
        A36="● Monasticismo e retiros iogues",
        A37="● Avalokiteśvara/Tārā",
        A38="● Gurus, terma e linhagens de revelação",
        A39="● Consciência sutil/luz clara; não ātman",
        A40="● União de sabedoria e método; não dualidade",
        A41="● Buddhahood em benefício de todos",
        A42="● Maitreya e ciclos Kālacakra",
        A43="● Abhiṣeka, pūjā, gaṇacakra e oferendas",
        A44="● Deity yoga e transformação de corpo-fala-mente",
    ),
    "chinese": mapped(
        A01="● Hùndùn, qì primordial e Pangu em tradições distintas",
        A02="● Nüwa/Pangu em mitos; o Céu não é criador único",
        A03="● Tiān, Imperador de Jade e Shangdi em camadas históricas",
        A04="● Tiān, lǐ e Mandato do Céu",
        A05="● Hòutǔ e Nüwa",
        A06="● Léigōng e Reis Dragões",
        A07="● Tàiyáng xīngjūn",
        A08="● Cháng'é e Tàiyīn xīngjūn",
        A09="● Dǒumǔ, estrelas e imperadores estelares",
        A10="● Lóngwáng, Mazu e deuses fluviais",
        A11="● Zàojūn e divindades do fogo/ofícios",
        A12="≈ Fēngbó e qì",
        A13="● Shénnóng, Tǔdìgōng e deuses agrários",
        A14="● Senhores Tigre e divindades locais de animais",
        A15="● Yuè Lǎo e deusas locais de amor/fertilidade",
        A16="● Sòngzǐ Guānyīn, Zhùshēng Niángniáng e ancestrais",
        A17="● Guāndì/Guān Yǔ e Erlang",
        A18="● Imperador de Jade, Chénghuáng e Mandato do Céu",
        A19="● Wénchāng, Confúcio e Guānyīn em saber religioso",
        A20="● Lǔ Bān e deuses de ofícios",
        A21="● Sūn Wùkōng e mensageiros; Ménshén guardam limiares",
        A22="● Cáishén e Tǔdìgōng",
        A23="● Bǎoshēng Dàdì, Huà Tuó e Guānyīn",
        A24="● Wángyé e deuses de peste também protetores",
        A25="● Yánluó Wáng e tribunais do submundo",
        A26="● Culto ancestral e tábuas de linhagem",
        A27="● Hēi-Bái Wúcháng e Dìzàng",
        A28="● Tribunais do inferno, Chénghuáng e Tiān",
        A29="● Sīmìng, Tàisuì e destino calculado",
        A30="● Hùndùn e calamidades; não mal absoluto",
        A31="≈ Demônios/guǐ hostis; podem ser apaziguados/transformados",
        A32="● Renovação ancestral, ciclos e imortalidade em correntes",
        A33="● Ménshén, Tǔdìgōng, Chénghuáng e Mazu",
        A34="● Deuses do teatro, música e artes; cultos locais",
        A35="● Possessão/espiritismo ritual e festivais; varia",
        A36="≈ Monasticismos budista/daoista; religião popular não exige",
        A37="● Guānyīn e Mazu",
        A38="● Médiuns, oráculos e textos revelados",
        A39="● Hún/pò, qì e shén",
        A40="● Yīn-yáng e cinco fases",
        A41="≈ Imortalidade, terras puras e boa ancestralidade; metas plurais",
        A42="≈ Ciclos dinásticos e escatologias salvacionistas",
        A43="● Incenso, oferendas, festivais e mediação ritual",
        A44="● Possessão, iniciação e cultivo",
    ),
    "daoist": mapped(
        A01="● Dào e qì primordial",
        A02="≈ Dào gera; Três Puros revelam/ordenam, não criam ex nihilo",
        A03="● Três Puros; Imperador de Jade em religião compartilhada",
        A04="● Dào, dé e registros celestes",
        A05="● Hòutǔ e Mãe do Carro (Dǒumǔ)",
        A06="● Léishēng Pǔhuà Tiānzūn e ministério do trovão",
        A07="● Imperador Solar e deidades da luz",
        A08="● Imperatriz Lunar",
        A09="● Dǒumǔ e deuses estelares",
        A10="● Reis Dragões e imortais das águas",
        A11="● Senhores do Fogo e alquimia interna",
        A12="● Qì e deuses do vento",
        A13="≈ Deuses locais e imortais agrários",
        A14="● Xīwángmǔ, montanhas e animais de imortais",
        A15="≈ Héxiān Gū e cultos locais; sexualidade varia por escola",
        A16="≈ Dǒumǔ e deusas protetoras",
        A17="● Zhēnwǔ e generais celestes",
        A18="● Imperador de Jade e burocracia celeste",
        A19="● Lǎojūn, Wénchāng e imortais",
        A20="● Lǔ Bān e alquimistas/ofícios",
        A21="● Mensageiros e oficiais celestes",
        A22="● Cáishén e imortais de prosperidade",
        A23="● Lǚ Dòngbīn, medicina e talismãs",
        A24="● Ministérios rituais de peste/exorcismo",
        A25="● Fēngdū Dàdì e tribunais",
        A26="● Ancestrais, mestres de linhagem e imortais",
        A27="● Oficiais de Fēngdū e guias",
        A28="● Registros de mérito/demérito e tribunais",
        A29="● Tàiyǐ, estrelas e ciclos",
        A30="◇ Hùndùn como estado primordial; não mal",
        A31="≈ Demônios e desvios do qì; não princípio absoluto",
        A32="● Longevidade, imortalidade e renovação",
        A33="● Zhēnwǔ, generais e talismãs",
        A34="● Música ritual, caligrafia e artes de imortais",
        A35="● Meditação, visualização e rituais; escola específica",
        A36="● Quánzhēn monástico e disciplinas internas",
        A37="◇ Cí, compaixão, nos Três Tesouros do Dàodéjīng",
        A38="● Revelações Shàngqīng/Língbǎo e mestres celestes",
        A39="● Hún/pò, qì, shén e cultivo",
        A40="● Yīn-yáng e retorno ao Dào",
        A41="● Imortalidade/transcendência; sentidos diversos",
        A42="● Escatologias Língbǎo e ciclos cósmicos",
        A43="● Jiào, talismãs, oferendas e sacerdócio",
        A44="● Alquimia interna, iniciação e transformação",
    ),
    "confucian": mapped(
        A01="◇ Tiān e ordem do cosmos; metafísica varia por período",
        A02="— Não há criador pessoal central",
        A03="◇ Tiān/Shangdi como autoridade moral",
        A04="● Lǐ, yì, rén e Mandato do Céu",
        A05="◇ Terra no par Céu-Terra",
        A06="◇ Céu como ordem; fenômenos eram lidos ritualmente",
        A07="≈ Sol em ritual estatal, não arquétipo confuciano central",
        A08="≈ Lua em calendário/ritual",
        A10="≈ Rios e montanhas em culto estatal",
        A13="◇ Governo benevolente e agricultura",
        A16="● Xiào, família e ritos",
        A17="◇ Coragem regulada por retidão",
        A18="● Mandato do Céu, rei-sábio e governo ritual",
        A19="● Confúcio, sábios, estudo e clássicos",
        A20="◇ Educação, ritos e aperfeiçoamento humano",
        A21="— Sem trickster ou mensageiro central",
        A22="◇ Prosperidade subordinada à retidão",
        A23="◇ Cultivo moral/corporal; medicina é tradição correlata, não núcleo",
        A25="◇ Morte tratada por luto e ritual",
        A26="● Culto ancestral, xiào e continuidade familiar",
        A28="● Yì e ordem moral; retribuição não exige juiz divino",
        A29="◇ Mìng: mandato/destino",
        A31="◇ Egoísmo/desordem moral, não mal personificado",
        A32="◇ Renovação moral e continuidade da linhagem",
        A33="≈ Ancestrais e espíritos tutelares no contexto ritual chinês",
        A34="● Música e ritos cultivam a pessoa e o Estado",
        A36="◇ Autodisciplina; não renúncia monástica clássica",
        A37="● Rén e benevolência",
        A38="● Sábios transmitem o Caminho; não profetas de revelação exclusiva",
        A39="● Xīn, qì e natureza humana; escolas divergem",
        A40="● Harmonia, meio e complementaridade Céu-Terra-Humano",
        A41="◇ Tornar-se sábio e realizar a humanidade, não salvação teísta",
        A43="● Lǐ: rito, sacrifício ancestral e estatal",
        A44="● Autocultivo e transformação moral",
    ),
    "shinto": mapped(
        A01="● Kotoamatsukami e caos primordial em Kojiki/Nihon Shoki",
        A02="● Izanagi e Izanami formam ilhas/kami; versões múltiplas",
        A03="● Amaterasu no mito imperial; sem 'deus supremo' único",
        A04="● Musubi, makoto e ordem por rito",
        A05="● Izanami, Ōgetsuhime e kami da terra",
        A06="● Susanoo, Raijin e kami atmosféricos",
        A07="● Amaterasu",
        A08="● Tsukuyomi",
        A09="● Ame-no-Uzume e kami astrais/aurora em leituras específicas",
        A10="● Ryūjin/Watatsumi e Suijin",
        A11="● Kagutsuchi e kami do fogo/forja",
        A12="● Shinatsuhiko: vento",
        A13="● Inari, Ōgetsuhime e kami do arroz",
        A14="● Ōyamatsumi e kami de montanhas/animais",
        A15="● Konohanasakuya-hime e Ōkuninushi em casamento/beleza",
        A16="● Konohanasakuya-hime e kami tutelares do parto",
        A17="● Hachiman e Takemikazuchi",
        A18="● Amaterasu e Ninigi na legitimação imperial",
        A19="● Tenjin: aprendizagem; Omoikane: sabedoria",
        A20="● Ame-no-Mahitotsu e kami de ofícios",
        A21="● Sarutahiko e Ame-no-Uzume: limiar/mediação",
        A22="● Inari, Ebisu e Daikokuten (sincretismos)",
        A23="● Sukunahikona e Ōkuninushi: cura",
        A24="● Gozu Tennō/Susanoo em cultos de epidemia; sincretismo",
        A25="● Izanami/Yomi e kami dos mortos",
        A26="● Ujigami e ancestrais de clã",
        A27="● Sarutahiko como guia; função psicopômpica limitada",
        A28="◇ Pureza/impureza e consequência ritual; sem juiz pós-morte central",
        A29="≈ Kami de destino e calendário; sem figura única",
        A30="● Magatsuhi: calamidade/poluição; não mal absoluto",
        A31="≈ Oni e espíritos hostis; categoria porosa",
        A32="● Musubi, purificação e renovação sazonal",
        A33="● Ujigami, Hachiman e kami locais",
        A34="● Ame-no-Uzume, kagura e artes",
        A35="● Kagura, matsuri e possessão/oráculo em contextos",
        A36="≈ Misogi e disciplina; não ascese monástica central",
        A37="◇ Magokoro e cuidado comunitário; não personificação única",
        A38="● Oráculos, miko e mitos imperiais",
        A39="● Mitama e tama",
        A40="◇ Musubi e equilíbrio por purificação/rito",
        A41="◇ Vida harmoniosa com kami; sem soteriologia única",
        A43="● Oferendas, norito, matsuri e sacerdócio",
        A44="● Misogi, iniciação e apoteose de pessoas como kami",
    ),
    "maya": mapped(
        A01="● Mar/céu primordial e deuses criadores; versões regionais/temporais",
        A02="● Itzamnaaj em fontes yucatecas; Xmucane/Xpiyacoc no Popol Wuj",
        A03="● Itzamnaaj e K'awiil; autoridade varia por cidade/período",
        A04="● K'awiil e ordem dinástica/calendárica",
        A05="● Ix Chel/Ixik Kab' e deusas do milho/terra; nomes variam",
        A06="● Chaak: chuva e tempestade",
        A07="● K'inich Ajaw",
        A08="● Deusa lunar (Ix Chel em tradição yucateca tardia)",
        A09="● Vênus ligada a guerra e augúrios; personificações variam",
        A10="● Chaak, Bacabs e divindades aquáticas",
        A11="● K'awiil e deuses do fogo",
        A12="≈ Ik': vento/sopro personificado",
        A13="● Deus do Milho (Hun Hunahpu em leitura K'iche' relacionada)",
        A14="● Deuses jaguar e animais-companheiros/wahy",
        A15="≈ Ix Chel e divindades femininas; atribuições coloniais variam",
        A16="● Ix Chel e deusas do parto/tecelagem",
        A17="● K'awiil, deuses de Vênus e heróis gêmeos",
        A18="● K'awiil e ancestrais divinizados legitimam governantes",
        A19="● Itzamnaaj, deuses escribas e Itzam Ye",
        A20="● Deuses escribas/artesãos; macaco-escriba",
        A21="● Heróis Gêmeos e coati/macaco em narrativas",
        A22="● Ek Chuah: comércio/cacau em fontes yucatecas",
        A23="● Ix Chel e especialistas rituais",
        A24="● Deuses da morte/doença; ritos apotropaicos",
        A25="● Deuses da morte e Xibalba/Metnal",
        A26="● Ancestrais reais e domésticos",
        A27="● Cães/seres guias em algumas tradições",
        A28="≈ Julgamento não é sistema único; provações em Xibalba",
        A29="● Calendários, deuses de períodos e destino",
        A30="● Monstros cósmicos e deuses da morte",
        A31="≈ Senhores de Xibalba: adversários narrativos, não mal absoluto",
        A32="● Deus do Milho e ciclos agrícolas/dinásticos",
        A33="● Deuses patronos de cidades/linhagens",
        A34="● Deuses escribas, músicos e dança régia",
        A35="● Sangria, jejum e substâncias em ritos específicos; evidência contextual",
        A38="● Ajk'uhuun, adivinhação e livros calendáricos",
        A39="● Ch'ulel/k'uh e coessências; termos regionais",
        A40="● Direções, cores e ciclos complementares",
        A41="≈ Continuidade ancestral e renascimento do milho; não salvação única",
        A42="● Fins/inícios calendáricos; 2012 não era 'fim do mundo' maia",
        A43="● Oferendas, auto-sacrifício e sacrifício em contextos documentados",
        A44="● Ritos de entronização, visão e transformação",
    ),
    "mexica": mapped(
        A01="● Ometeotl em fontes coloniais é debatido; ciclos dos Sóis são centrais",
        A02="● Quetzalcóatl e Tezcatlipoca formam o mundo em relatos; deuses se sacrificam",
        A03="● Tezcatlipoca e Huitzilopochtli em soberania mexica",
        A04="● Teotl, tonalli e ordem calendárica; interpretações variam",
        A05="● Tlaltecuhtli, Coatlicue e Cihuacóatl",
        A06="● Tláloc",
        A07="● Tonatiuh e Huitzilopochtli em função solar",
        A08="● Metztli/Coyolxauhqui em tradições distintas",
        A09="● Tlahuizcalpantecuhtli: Vênus/aurora",
        A10="● Chalchiuhtlicue e Tláloc",
        A11="● Xiuhtecuhtli-Huehuetéotl",
        A12="● Ehécatl-Quetzalcóatl",
        A13="● Chicomecóatl, Centeotl e Xipe Tótec",
        A14="● Mixcóatl e Tepeyóllotl",
        A15="● Xochiquetzal e Tlazoltéotl",
        A16="● Cihuacóatl, Tlazoltéotl e divindades do parto",
        A17="● Huitzilopochtli, Mixcóatl e Tezcatlipoca",
        A18="● Tezcatlipoca, Huitzilopochtli e Xiuhtecuhtli",
        A19="● Quetzalcóatl e Tonacatecuhtli; calmécac",
        A20="● Quetzalcóatl e Xipe Tótec em artes/ofícios",
        A21="● Tezcatlipoca e Huehuecóyotl; funções liminares distintas",
        A22="● Yacatecuhtli e deuses dos pochteca",
        A23="● Ixtlilton e Toci; cura ritual",
        A24="● Tlazoltéotl e deuses de doença/cura",
        A25="● Mictlantecuhtli e Mictecacíhuatl",
        A26="● Ancestrais, guerreiros mortos e cihuateteo",
        A27="● Xólotl: guia e limiar",
        A28="◇ Destino pós-morte depende sobretudo do modo de morte; não julgamento moral único",
        A29="● Tonalpohualli, tonalli e Xiuhtecuhtli",
        A30="● Tezcatlipoca, Tzitzimime e destruições dos Sóis",
        A31="≈ Tzitzimime e seres hostis; não diabo",
        A32="● Xipe Tótec e ciclos dos Sóis",
        A33="● Patronos de calpulli/cidades e Huitzilopochtli",
        A34="● Xochipilli, Xochiquetzal e Huehuecóyotl",
        A35="● Pulque, teonanácatl e outras substâncias em contextos específicos",
        A36="● Jejum, sangria e disciplina sacerdotal",
        A38="● Tonalpouhque, sacerdotes e presságios",
        A39="● Tonalli, teyolía e ihíyotl: componentes distintos",
        A40="● Oposições/complementaridades calendáricas e espaciais",
        A41="≈ Destinos pós-morte plurais e continuidade cósmica",
        A42="● Ciclos dos Sóis e possível catástrofe/renovação",
        A43="● Oferenda e sacrifício, inclusive humano em contextos estatais",
        A44="● Ixiptla, iniciação e transformação ritual",
    ),
    "andean": mapped(
        A01="● Viracocha/Wiraqucha em relatos coloniais; tradições andinas diversas",
        A02="● Viracocha como ordenador/criador em narrativas incas",
        A03="● Inti e Viracocha; soberania imperial",
        A04="● Camac/ordenamento e reciprocidade; termos variam",
        A05="● Pachamama",
        A06="● Illapa: trovão e chuva",
        A07="● Inti",
        A08="● Mama Killa",
        A09="● Ch'aska: Vênus/estrela",
        A10="● Mama Qucha e apus/huacas de água",
        A11="● Nina e cultos domésticos; sem deus único pan-andino",
        A12="≈ Wayra: vento",
        A13="● Pachamama, Sara Mama e deidades de cultivos",
        A14="● Apus e senhores dos animais; tradições locais",
        A15="≈ Mama Ocllo e divindades locais de fecundidade",
        A16="● Mama Killa/Pachamama em proteção familiar",
        A17="● Illapa e ancestrais imperiais",
        A18="● Inti e ancestrais incas legitimam o Sapa Inca",
        A19="● Amautas, quipucamayoc e Viracocha em sabedoria",
        A20="● Huari/Con-Tici em ofícios; atribuições variam",
        A21="● Ekeko é prosperidade colonial/andina tardia, não mensageiro",
        A22="● Ekeko e huacas de abundância em tradições posteriores",
        A23="● Kallawaya, huacas, apus e plantas",
        A24="≈ Supay/huacas e desequilíbrio; não causa única",
        A25="● Supay/Ukhu Pacha e deuses ctônicos",
        A26="● Mallquis, wak'as e ancestrais",
        A27="≈ Cães e especialistas conduzem em alguns relatos",
        A28="◇ Reciprocidade/ordem e destinos; não juiz universal",
        A29="● Pachakuti: reversão de tempo-espaço; calendários",
        A30="● Pachakuti e forças sísmicas/catastróficas",
        A31="≈ Supay foi demonizado no período colonial; não diabo original simples",
        A32="● Ciclos agrícolas, ancestrais e pachakuti",
        A33="● Apus, huacas e ancestrais tutelares",
        A34="● Música, dança e têxteis rituais",
        A35="● Chicha, coca e vilca em contextos específicos e documentados",
        A38="● Oráculos, wak'as e especialistas",
        A39="● Camay/ánimu e componentes da pessoa; termos regionais",
        A40="● Yanantin/masintin e reciprocidade complementar",
        A41="≈ Continuidade ancestral; sincretismos cristãos posteriores",
        A42="● Pachakuti como mudança de era em interpretações andinas",
        A43="● Despachos, libações, capac hucha em contexto imperial",
        A44="● Ritos de passagem e transformação xamânica",
    ),
    "judaism": mapped(
        A01="● Deus anterior e distinto da criação; leituras cabalísticas incluem Ein Sof",
        A02="● YHWH/Elohim",
        A03="● YHWH, Rei dos céus e da terra",
        A04="● Torá, emet e ordem da aliança",
        A05="≈ Shekhinah em linguagem feminina; não deusa-terra",
        A06="● YHWH domina tempestade; imagens poéticas não implicam deus separado",
        A07="● Sol/luz como criação e metáfora divina; não divindade",
        A08="● Lua como criação/calendário; não divindade",
        A09="● Estrelas como criação; proibição de culto astral",
        A10="● Deus domina mares/águas; Leviatã em imagens",
        A11="● Fogo divino, menorá e altar",
        A12="● Ruaḥ: vento/sopro/espírito",
        A13="● Deus, terra de Israel e leis agrícolas",
        A14="● Animais como criação; Leviatã/Beemote simbólicos",
        A15="● Amor humano/divino no Cântico e tradição; sem deidade do amor",
        A16="● Aliança, família, matriarcas e bênção",
        A17="● YHWH como guerreiro em textos antigos; ética interpretativa posterior",
        A18="● YHWH, Torá e realeza davídica",
        A19="● Ḥokhmah/Sabedoria, Torá e escribas",
        A20="● Bezalel como artesão inspirado; Deus como artífice metafórico",
        A21="● Anjos/mal'akhim; Jacó e figuras liminares",
        A22="● Bênção/prosperidade subordinada à aliança e justiça",
        A23="● Deus como curador; Rafael em tradição textual posterior",
        A24="● Pragas como narrativa/juízo; não divindade autônoma",
        A25="● Sheol, morte e tradição posterior de Gehinnom/Olam Ha-Ba",
        A26="● Patriarcas/matriarcas e memória; culto aos mortos é restringido",
        A27="● Anjos; função psicopômpica varia em textos tardios",
        A28="● Deus julga; justiça, teshuvah e tribunal celeste",
        A29="● Deus, tempo sagrado e providência; mazzal em cultura posterior",
        A30="● Tohu va-vohu/Leviatã: desordem dominada",
        A31="● Satan como acusador em textos antigos; desenvolvimentos posteriores",
        A32="● Ressurreição dos mortos em correntes rabínicas; teshuvah como retorno",
        A33="● Deus, anjos e mezuzah em proteção",
        A34="● Salmos, música levítica e arte ritual",
        A35="≈ Profecia/êxtase e alegria ritual; intoxicantes não são via normativa",
        A36="● Jejum, nazireado e disciplinas; monasticismo não central",
        A37="● Ḥesed e raḥamim",
        A38="● Profetas, Moisés e revelação da Torá",
        A39="● Nefesh, ruaḥ e neshamah; esquemas variam",
        A40="◇ Justiça/misericórdia e pares cabalísticos; não dualismo de deuses",
        A41="● Aliança, teshuvah e Olam Ha-Ba; concepções plurais",
        A42="● Era messiânica, ressurreição e juízo em correntes",
        A43="● Sacrifício do Templo historicamente; oração/mitzvot após 70 d.C.",
        A44="● Circuncisão, conversão, bar/bat mitzvah e teshuvah",
    ),
    "christian": mapped(
        A01="● Deus eterno; Trindade nas tradições nicenas",
        A02="● Deus cria por meio do Logos/Filho",
        A03="● Deus/Pai; Cristo como Senhor",
        A04="● Logos, vontade de Deus e lei divina",
        A05="≈ Maria como Theotokos/Mãe de Deus em muitas tradições; não deusa-terra",
        A06="● Deus domina céu/tempestade; sem divindade separada",
        A07="● Cristo como luz/sol metafórico; Sol é criação",
        A08="≈ Maria/lua em iconografia; não divindade lunar",
        A09="● Estrela de Belém e anjos; astros são criação",
        A10="● Batismo e Cristo domina as águas",
        A11="● Espírito Santo/fogo; Cristo como luz",
        A12="● Espírito Santo como sopro/vento",
        A13="● Deus/Cristo em parábolas agrárias; santos patronos são devoção",
        A14="● Criação animal e santos patronos; sem deus da caça normativo",
        A15="● Deus é amor; eros/agape distinguidos",
        A16="● Sagrada Família, Maria e santos; devoção não equivale a panteão",
        A17="● Cristo guerreiro em Apocalipse; Miguel e santos militares",
        A18="● Reino de Deus e Cristo Rei",
        A19="● Logos, Espírito Santo e tradição de Sophia",
        A20="● Deus criador; José e santos de ofícios em devoção",
        A21="● Gabriel, anjos e João Batista como mensageiros",
        A22="◇ Providência; crítica à idolatria da riqueza/Mammon",
        A23="● Cristo curador; Rafael e santos em tradições",
        A24="◇ Doença não é divindade; demônios não explicam toda doença na teologia atual",
        A25="● Morte, Hades/inferno e vitória de Cristo",
        A26="● Comunhão dos santos e memória dos mortos; práticas denominacionais divergem",
        A27="● Anjos e Cristo; psicopompo não é título universal",
        A28="● Deus/Cristo julga; justiça e graça",
        A29="● Providência e tempo da salvação; predestinação varia",
        A30="● Caos e poderes vencidos; Apocalipse usa imagens destrutivas",
        A31="● Satanás/Diabo",
        A32="● Ressurreição de Cristo e ressurreição final",
        A33="● Deus, anjos, santos e sacramentais; conforme a tradição",
        A34="● Salmos, música, ícones e artes litúrgicas",
        A35="● Pentecostes/mística; vinho sacramental não é intoxicação ritual",
        A36="● Jejum, monasticismo e disciplinas; variam",
        A37="● Cristo, graça, caridade e misericórdia",
        A38="● Cristo/Logos, profetas, apóstolos e Escritura",
        A39="● Alma/espírito e pessoa; antropologias denominacionais variam",
        A40="◇ Trindade não é dualidade; corpo/alma interpretados de modos diversos",
        A41="● Salvação, deificação ou justificação; modelos divergem",
        A42="● Segunda vinda, juízo, ressurreição e nova criação",
        A43="● Cristo como sacrifício; Eucaristia e sacerdócio variam",
        A44="● Batismo, conversão, santificação e transfiguração",
    ),
    "gnostic": mapped(
        A01="● Mônada/Profundidade e Pléroma; sistemas muito diversos",
        A02="● Demiurgo cria/organiza o mundo material em muitos sistemas",
        A03="● Pai invisível/Grande Espírito",
        A04="● Pléroma e ordem dos éons",
        A05="● Barbelo/Sophia em funções geradoras; não 'mãe-terra' simples",
        A07="● Luz do Pléroma e Cristo",
        A10="≈ Águas luminosas/abismo em textos específicos",
        A15="≈ Sizígias e Éros em alguns sistemas",
        A16="● Barbelo e Sophia como matrizes em sistemas",
        A17="≈ Arcontes e conflito cósmico",
        A18="● Demiurgo/arcontes governam o cosmos inferior",
        A19="● Sophia e gnōsis",
        A21="● Cristo/Salvador e reveladores",
        A23="≈ Cristo/gnose cura a ignorância",
        A25="● Mundo material/morte e regiões dos arcontes",
        A27="● Salvador e figuras reveladoras guiam a centelha",
        A28="≈ Ascensão através de poderes; sistemas divergem",
        A29="● Éons como emanações, não eras históricas comuns",
        A30="● Deficiência/paixão de Sophia em alguns mitos",
        A31="● Demiurgo/arcontes como adversários em muitos sistemas",
        A32="● Despertar da centelha e retorno ao Pléroma",
        A33="≈ Poderes/selos protetores em rituais",
        A35="● Experiência visionária e ritos; evidência desigual",
        A36="● Correntes ascéticas e outras libertinas: relatos polemizados",
        A38="● Cristo e revelações secretas",
        A39="● Pneuma/centelha divina; antropologias tripartidas em sistemas",
        A40="● Pléroma/deficiência e luz/matéria; dualismos variam",
        A41="● Gnose e retorno ao Pléroma",
        A42="≈ Restauração/consumação do Pléroma em sistemas",
        A43="≈ Batismo, unção, câmara nupcial; ritos variam",
        A44="● Despertar e ascensão",
    ),
    "manichaean": mapped(
        A01="● Pai da Grandeza e Reino da Luz; Matéria/Reino das Trevas coeterno",
        A02="● Mãe da Vida e Homem Primordial emanam para o conflito",
        A03="● Pai da Grandeza",
        A04="● Ordem da Luz",
        A05="● Mãe da Vida; não terra-mãe",
        A07="● Jesus Esplendor e luminários resgatam partículas de luz",
        A08="● Lua e Sol como navios de luz",
        A10="≈ Elementos da Luz aprisionados na mistura",
        A11="● Fogo como elemento da Luz e da mistura",
        A12="● Vento como elemento",
        A13="◇ Plantas concentram luz; dieta dos Eleitos",
        A14="◇ Animais contêm mistura; ética alimentar específica",
        A17="● Homem Primordial e batalha cósmica",
        A18="● Governantes da Luz e arcontes das Trevas",
        A19="● Grande Nous e Mani",
        A21="● Terceiro Enviado e Jesus Esplendor",
        A24="◇ Sofrimento decorre da mistura; não deus de doença",
        A25="● Matéria/corpo e reinos pós-morte",
        A27="● Jesus Esplendor e guias de luz",
        A28="● Julgamento e separação de luz",
        A29="● Três Tempos: antes, mistura e separação",
        A30="● Reino das Trevas e matéria",
        A31="● Príncipe das Trevas",
        A32="● Libertação das partículas de luz",
        A33="● Poderes da Luz",
        A35="≈ Hinos e experiência visionária; sem intoxicação normativa",
        A36="● Eleitos praticam ascese rigorosa",
        A37="● Compaixão ao libertar luz; ética dualista",
        A38="● Mani, Selo dos Profetas",
        A39="● Alma/luz aprisionada",
        A40="● Dualismo radical Luz/Trevas",
        A41="● Libertação e retorno ao Reino da Luz",
        A42="● Separação final e Grande Fogo",
        A43="● Confissão, refeições dos Eleitos e apoio dos Ouvintes",
        A44="● Purificação da luz",
    ),
    "islam": mapped(
        A01="● Allah eterno, absolutamente um e distinto da criação",
        A02="● Allah",
        A03="● Allah, al-Malik",
        A04="● Al-Ḥaqq, revelação e ordem divina",
        A05="— Sem deusa-terra; terra é criação/sinal",
        A06="● Allah governa chuva/trovão; anjos cumprem ordens",
        A07="● Sol/luz como criação e sinal; al-Nūr como nome/metáfora",
        A08="● Lua como criação e calendário; não divindade",
        A09="● Estrelas como sinais/criação; astrologia é contestada",
        A10="● Allah governa mares/águas",
        A11="● Fogo como criação, punição e símbolo; jinn criados de fogo",
        A12="● Rūḥ e ventos por ordem divina",
        A13="● Allah concede colheita; al-Khiḍr em tradições posteriores",
        A14="● Animais como comunidade/sinais; sem deus da caça",
        A15="● Al-Wadūd e amor ético/místico; sem deidade do amor",
        A16="● Família, Maryam e profetas como modelos; não divindades",
        A17="● Allah, ética de jihād e figuras históricas; não deus da guerra",
        A18="● Allah e sharīʿa; califado/imamato são instituições históricas",
        A19="● Al-ʿAlīm, revelação e ḥikma",
        A20="◇ Trabalho/ofício valorizado; sem divindade técnica",
        A21="● Jibrīl e anjos mensageiros",
        A22="● Al-Razzāq; riqueza é prova e responsabilidade",
        A23="● Al-Shāfī; medicina profética e prática médica histórica",
        A24="◇ Doença como prova/condição; não personificação divina",
        A25="● Malak al-mawt, barzakh e Jahannam",
        A26="● Respeito/memória dos mortos; culto varia e é debatido",
        A27="● Anjos, especialmente o anjo da morte; nomes variam por fonte",
        A28="● Allah julga; Mīzān e Dia do Juízo",
        A29="● Qadar e tempo criado; escolas debatem agência",
        A30="● Fim cósmico por decreto divino; não caos autônomo",
        A31="● Iblīs/al-Shayṭān",
        A32="● Ressurreição corporal",
        A33="● Allah e anjos; práticas de proteção corânicas",
        A34="≈ Recitação é central; música é avaliada de modos diversos",
        A35="≈ Dhikr/êxtase em correntes sufis; intoxicantes são proibidos",
        A36="● Ṣawm, disciplina e zuhd; sem monasticismo normativo",
        A37="● Al-Raḥmān/al-Raḥīm e raḥma",
        A38="● Muḥammad, profetas e Qurʾān",
        A39="● Nafs, rūḥ, qalb e ʿaql: não sinônimos",
        A40="● Tawḥīd; pares criação/ordem sem dualismo divino",
        A41="● Submissão a Deus, perdão e Paraíso",
        A42="● Qiyāma, juízo, ressurreição e nova ordem",
        A43="● Ṣalāt, zakāt, ḥajj e qurbān; sem sacerdócio sacramental",
        A44="● Shahāda, purificação e transformação ética",
    ),
    "sufi": mapped(
        A01="● Allah/al-Ḥaqq; doutrinas de unidade variam por escola",
        A02="● Allah; emanação não substitui criação no Islã normativo",
        A03="● Allah",
        A04="● Ḥaqīqa, sharīʿa e ṭarīqa",
        A07="● Nūr Muḥammadī em algumas tradições",
        A10="● Oceano como metáfora da unidade/amor; não divindade",
        A11="● Fogo do amor e luz interior como metáforas",
        A12="● Rūḥ e nafas al-Raḥmān em linguagem mística",
        A15="● ʿIshq/maḥabba: amor divino",
        A18="● Quṭb e hierarquias espirituais; não governo político universal",
        A19="● Maʿrifa e mestres",
        A21="● Khiḍr e mestres como guias",
        A22="◇ Baraka, faqr e generosidade; riqueza material é ambivalente",
        A23="● Baraka, oração e santos em cura popular",
        A25="● Morte do ego e barzakh",
        A26="● Linhagens de santos e visitação de túmulos; contestada por correntes",
        A27="● Khiḍr/murshid como guias",
        A28="● Allah julga; muḥāsaba interior",
        A29="● Qadar e momento espiritual (waqt)",
        A31="● Nafs al-ammāra e Shayṭān como adversários",
        A32="● Fanāʾ e baqāʾ; não ressurreição literal substituta",
        A33="● Awliyāʾ e baraka em tradições",
        A34="● Samāʿ, qawwali, poesia e dhikr; ordens divergem",
        A35="● Wajd/ḥāl em dhikr; intoxicantes são metáfora/proibidos",
        A36="● Zuhd, khalwa e disciplina da ṭarīqa",
        A37="● Raḥma e amor divino",
        A38="● Muḥammad, Khiḍr e inspiração dos santos sem nova profecia",
        A39="● Nafs, rūḥ, qalb e sirr",
        A40="● Jamāl/jalāl e integração sem dualismo de deuses",
        A41="● Maʿrifa, fanāʾ/baqāʾ e proximidade de Deus",
        A42="● Escatologia islâmica reinterpretada misticamente",
        A43="● Dhikr, bayʿa, suhba e serviço",
        A44="● Sulūk e purificação do ego",
    ),
    "sikh": mapped(
        A01="● Ik Oaṅkār: realidade una e sem forma",
        A02="● Kartā Purakh",
        A03="● Akāl Purakh",
        A04="● Hukam, Sat e dharam",
        A05="— Sem deusa-terra; natureza participa da criação",
        A06="● Fenômenos obedecem ao Hukam; sem deus do trovão",
        A07="● Luz divina (jot) como metáfora/presença",
        A08="● Lua como criação; não divindade",
        A10="● Água como criação e símbolo; Amrit em iniciação",
        A11="● Fogo como criação/metáfora",
        A12="● Pavan como ar/sopro na criação",
        A13="● Trabalho agrícola e honestidade; sem deidade da colheita",
        A14="◇ Criação é expressão divina; ética de cuidado",
        A15="● Amor por Vāhigurū; rejeição de erotização do divino",
        A16="● Gṛhasth: vida familiar como caminho",
        A17="● Sant-sipāhī e defesa da justiça",
        A18="● Mīrī-pīrī e Guru Panth; Deus é soberano",
        A19="● Gurū Granth Sāhib e Gurus",
        A20="◇ Kirat karō: trabalho honesto",
        A21="● Gurus e gurbāṇī comunicam o Nome",
        A22="◇ Vaṇḍ chhakō e desapego; riqueza não é bênção automática",
        A23="● Nām, oração e serviço; medicina não substituída",
        A24="◇ Doença como condição; sem entidade divina da doença",
        A25="● Morte e ciclo de transmigração",
        A26="● Memória dos Gurus e mártires; não culto ancestral",
        A27="● Guru como guia espiritual",
        A28="● Hukam, karma e justiça divina",
        A29="● Akāl e Hukam",
        A30="◇ Haumai e desordem moral, não caos divino",
        A31="◇ Cinco ladrões/haumai, não diabo soberano",
        A32="● Renascimento e mukti",
        A33="● Vāhigurū; Khalsa como comunidade protetora",
        A34="● Kīrtan e poesia do Gurū Granth",
        A35="● Nām simran; intoxicantes são proibidos no código Khalsa",
        A36="● Disciplina, mas rejeição do ascetismo como única via",
        A37="● Dayā, sevā e sarbat da bhala",
        A38="● Dez Gurus e Gurū Granth Sāhib",
        A39="● Ātma/jot em relação ao divino; linguagem não dual",
        A40="● Mīrī-pīrī e unidade na diversidade",
        A41="● Mukti por graça, Nām e vida ética",
        A42="≈ Sem escatologia final central; foco em libertação",
        A43="● Amrit, langar, sevā e culto sem sacrifício animal",
        A44="● Amrit sañcār e transformação pelo Nām",
    ),
    "afro_yoruba_diaspora": mapped(
        A03="● Olódùmarè/Olorun/Olofi; terminologia varia por linhagem",
        A10="● Yemanjá/Yemayá e Olokun; funções variam por nação",
        A15="● Oxum/Ochún e entidades afins",
        A21="● Exu/Elegguá: comunicação e encruzilhadas; não equivale ao Diabo",
        A23="● Ossaim/Osain e Omolu/Obaluaiê",
        A24="● Omolu/Obaluaiê/Babalú-Ayé: doença e cura",
        A26="● Eguns/egún e ancestrais",
        A34="● Toques, cantos, dança e línguas litúrgicas",
        A35="● Incorporação/possessão ritual; terminologia da casa",
        A43="● Ebó, oferenda, iniciação e sacerdócio",
        A44="● Feitura/iniciação e vínculo com orixá",
    ),
    "umbanda": mapped(
        A01="● Olorum/Zambi/Deus; formulações variam por corrente",
        A02="● Deus/Olorum; orixás não são criadores independentes",
        A03="● Olorum/Zambi",
        A04="◇ Lei divina, caridade e evolução espiritual",
        A05="● Nanã, Iemanjá e linhas da terra; teologias variam",
        A06="● Xangô e Iansã",
        A07="● Oxalá e linhas solares em algumas escolas",
        A08="≈ Iemanjá e linhas lunares em sistemas esotéricos específicos",
        A10="● Iemanjá, Oxum e Nanã",
        A11="● Ogum e Xangô; fogo/ferro",
        A12="● Iansã",
        A13="● Oxóssi e caboclos; mata/alimento",
        A14="● Oxóssi, caboclos e encantados",
        A15="● Oxum; Pombagiras em desejo/limiar, não 'deusas do amor'",
        A16="● Iemanjá, Oxum e linhas infantis",
        A17="● Ogum e Xangô",
        A18="● Xangô: justiça/autoridade",
        A19="● Pretos-velhos, caboclos e guias; saber espiritual",
        A20="● Ogum: trabalho e caminhos",
        A21="● Exus e Pombagiras: guardiões/comunicadores de esquerda",
        A22="≈ Ciganos, Exus e linhas de prosperidade; varia",
        A23="● Pretos-velhos, caboclos e orixás de cura",
        A24="◇ Desequilíbrio espiritual não substitui diagnóstico médico",
        A25="● Omolu/Obaluaiê e linhas de almas",
        A26="● Pretos-velhos, ancestrais e eguns em doutrinas distintas",
        A27="● Exus e guias espirituais",
        A28="● Xangô e lei de causa/efeito",
        A29="◇ Evolução/reencarnação em correntes kardecizadas",
        A30="≈ Quimbanda/esquerda não significa mal; categorias variam",
        A31="— Sem adversário absoluto comum; Exu não é Diabo",
        A32="● Reencarnação/evolução em muitas correntes",
        A33="● Guias, orixás e Exus guardiões",
        A34="● Pontos cantados, atabaques e dança",
        A35="● Incorporação mediúnica",
        A36="◇ Disciplina e caridade; sem ascetismo único",
        A37="● Caridade e atuação de guias",
        A38="● Guias, médiuns e revelação de terreiro",
        A39="● Espírito/perispírito em correntes; axé em outras",
        A40="● Direita/esquerda e linhas; não bem/mal simples",
        A41="● Evolução espiritual e caridade",
        A43="● Giras, oferendas e defumação; variam por casa",
        A44="● Desenvolvimento mediúnico e iniciação",
    ),
    "haitian_vodou": mapped(
        A01="● Bondye: fonte transcendente",
        A02="● Bondye",
        A03="● Bondye; lwa mediam relações",
        A04="◇ Regleman, reciprocidade e ordem ancestral",
        A05="● Ayizan e Ezili Dantò em funções maternas; não terra universal",
        A06="● Sogbo/Badè e Ogou em tempestade/força",
        A07="≈ Legba/Kalfou e espíritos solares em nações específicas",
        A08="≈ Ayida Wedo/Marassa em simbolismos; sem lua central",
        A10="● Agwe, Lasirèn e Simbi",
        A11="● Ogou e espíritos do fogo/ferro",
        A12="≈ Badè e lwa dos ventos",
        A13="● Azaka Mede e Kouzen Zaka",
        A14="● Gran Bwa e espíritos da mata",
        A15="● Ezili Freda e Ezili Dantò em aspectos distintos",
        A16="● Ezili Dantò e lwa familiares",
        A17="● Ogou",
        A18="● Ogou e líderes ancestrais; Bondye acima",
        A19="● Legba, Ayizan e sacerdócio",
        A20="● Ogou: ferro/trabalho",
        A21="● Papa Legba/Kalfou: portais e comunicação",
        A22="● Ezili Freda e lwa de prosperidade",
        A23="● Loko, Simbi, Ayizan e sacerdotes",
        A24="● Bawon Samdi/Gede e lwa podem causar/curar aflição",
        A25="● Bawon, Manman Brijit e Gede",
        A26="● Gede e ancestrais",
        A27="● Legba e Gede",
        A28="◇ Justiça por Bondye, lwa e ancestrais; não tribunal único",
        A29="≈ Marassa e ciclos ancestrais",
        A30="≈ Petro e Kalfou não equivalem a mal",
        A31="— Sem diabo central; leituras cristãs coexistem",
        A32="● Continuidade ancestral; concepções de retorno variam",
        A33="● Lwa mèt tèt e protetores",
        A34="● Tambores, canto, dança e vévé",
        A35="● Possessão pelos lwa; rum em ritos específicos",
        A38="● Lwa, sonhos, possessão e sacerdócio",
        A39="● Ti bon ange/gwo bon ange e componentes da pessoa",
        A40="● Rada/Petro e Marassa: diferenças não são bem/mal simples",
        A41="◇ Servir aos lwa e manter relações; não salvação única",
        A43="● Manje, libações, sèvis e iniciação",
        A44="● Kanzo e possessão",
    ),
    "spiritism": mapped(
        A01="● Deus: inteligência suprema e causa primeira",
        A02="● Deus",
        A03="● Deus",
        A04="● Leis divinas/naturais e progresso",
        A05="— Sem deusa-terra",
        A06="◇ Fenômenos naturais seguem leis; não divindade",
        A07="◇ Luz como metáfora de elevação",
        A08="— Sem divindade lunar",
        A10="— Sem divindade aquática",
        A11="◇ Fluido/energia em linguagem histórica; não fogo divino",
        A12="◇ Princípio vital/fluido em linguagem kardecista",
        A13="◇ Natureza como criação e campo de progresso",
        A14="◇ Animais possuem princípio inteligente em evolução",
        A15="● Amor como lei moral; não deidade",
        A16="● Família como espaço reencarnatório e moral",
        A17="◇ Coragem moral; rejeita deuses guerreiros",
        A18="● Deus e lei moral; Jesus como modelo/guia",
        A19="● Espíritos superiores, razão e estudo",
        A20="◇ Progresso intelectual/moral",
        A21="● Espíritos comunicantes e médiuns",
        A22="◇ Provas e uso ético da riqueza",
        A23="● Passe/prece como prática complementar; não substitui medicina",
        A24="◇ Obsessão é categoria religiosa; não explica toda doença mental",
        A25="● Desencarnação e erraticidade",
        A26="● Espíritos de mortos; evocação sob finalidade moral",
        A27="● Espíritos protetores/guias",
        A28="● Lei de causa e efeito; consciência como tribunal",
        A29="● Reencarnação e progresso",
        A30="◇ Ignorância moral; não caos personificado",
        A31="— Sem demônio criado para o mal; espíritos imperfeitos progridem",
        A32="● Reencarnação",
        A33="● Espírito protetor/anjo guardião",
        A34="≈ Arte e música no mundo espiritual em literatura doutrinária",
        A35="● Mediunidade; intoxicantes não são método doutrinário",
        A36="◇ Reforma íntima e disciplina, sem monasticismo",
        A37="● Caridade",
        A38="● Revelação espírita por médiuns; Jesus como guia",
        A39="● Espírito, perispírito e corpo",
        A40="◇ Progresso integra razão e fé",
        A41="● Evolução moral e felicidade dos espíritos",
        A42="≈ Renovação moral; não apocalipse final típico",
        A43="● Prece e caridade; sem sacrifício ou sacerdócio",
        A44="● Desobsessão, reforma íntima e reencarnação",
    ),
    "theosophy": mapped(
        A01="● Absoluto/Parabrahman e raiz sem raiz",
        A02="≈ Logos e hierarquias criadoras; não criador pessoal único",
        A03="≈ Logos e Mestres; estrutura esotérica",
        A04="● Karma e lei cósmica",
        A05="≈ Matéria primordial/Prakriti; apropriação comparativa",
        A06="≈ Devas elementais em sínteses teosóficas",
        A07="● Logos solar",
        A08="≈ Cadeias planetárias e simbolismo lunar",
        A09="● Hierarquias planetárias/astrais",
        A10="≈ Planos/elementais da água",
        A11="≈ Fohat e elementais do fogo",
        A12="≈ Fohat/prāṇa e planos",
        A13="≈ Espíritos da natureza",
        A14="≈ Devas/elementais e reinos da natureza",
        A15="◇ Amor-sabedoria como raio/princípio",
        A16="≈ Mãe do Mundo em correntes posteriores",
        A17="◇ Conflito evolutivo simbólico",
        A18="≈ Hierarquia espiritual/Mestres",
        A19="● Mahat, Mestres e sabedoria antiga",
        A20="● Fohat como energia organizadora; linguagem pré-científica",
        A21="● Mestres e mensageiros ocultos",
        A22="◇ Karma e uso evolutivo dos recursos",
        A23="≈ Corpos sutis e cura esotérica; não substitui medicina",
        A24="◇ Desequilíbrio kármico em doutrina histórica; evitar culpa médica",
        A25="● Kāmaloka/devachan e ciclos pós-morte",
        A26="≈ Mestres e cadeia de vidas; não culto ancestral central",
        A27="● Mestres/guias",
        A28="● Karma impessoal",
        A29="● Rondas, raças-raiz e ciclos; esquema especulativo",
        A30="≈ Caos primordial e involução",
        A31="◇ Ignorância/materialismo; sem adversário único",
        A32="● Reencarnação",
        A33="● Mestres e protetores",
        A34="≈ Arte como expressão de planos sutis",
        A35="≈ Clarividência/meditação; sem enteógeno doutrinário",
        A36="● Disciplina e caminho do discipulado",
        A37="● Fraternidade universal",
        A38="● Mestres, Blavatsky e revelação esotérica",
        A39="● Sete princípios/corpos; terminologia sincrética",
        A40="● Espírito/matéria e involução/evolução",
        A41="● Iniciação e evolução espiritual",
        A42="● Ciclos e humanidade futura",
        A43="≈ Ritual varia entre sociedades; sem sacrifício",
        A44="● Iniciação e evolução",
    ),
    "wicca": mapped(
        A01="≈ Mistério último/Uno em algumas tradições; não consenso",
        A02="≈ Deusa e Deus como fontes da natureza; teologias variam",
        A03="≈ Deusa/Deus Cornífero; sem soberania única",
        A04="◇ Reciprocidade, ética e ciclos naturais",
        A05="● Deusa/Deusa Tríplice e Terra; formas modernas",
        A06="≈ Deus Cornífero e divindades invocadas; varia",
        A07="● Deus solar em ciclos sazonais; tradição específica",
        A08="● Deusa lunar/Tríplice",
        A09="≈ Astrologia e deidades invocadas; não obrigatórias",
        A10="● Elemento Água e divindades escolhidas",
        A11="● Elemento Fogo, lareira e forja simbólica",
        A12="● Elemento Ar",
        A13="● Roda do Ano e fertilidade/colheita",
        A14="● Deus Cornífero e animais",
        A15="● Deusa/Deus e Grande Rito simbólico/ritual",
        A16="≈ Deusa e ritos de passagem",
        A17="≈ Divindades patronas escolhidas; não núcleo único",
        A18="◇ Sacerdócio de coven; sem realeza sagrada comum",
        A19="● Deusa, anciã(o), Livro das Sombras e iniciação",
        A20="● Magia, artes rituais e instrumentos",
        A21="≈ Divindades liminares invocadas conforme tradição",
        A22="≈ Magia de prosperidade; sem divindade universal",
        A23="● Cura ritual/herbalismo em correntes; não substitui medicina",
        A24="— Sem entidade universal da doença",
        A25="● Deusa Anciã/Deus e Summerland em algumas tradições",
        A26="● Mighty Dead/ancestrais em correntes tradicionais",
        A27="≈ Deidades psicopompas escolhidas; varia",
        A28="◇ Ética, reciprocidade e responsabilidade; 'lei tríplice' não universal",
        A29="● Roda do Ano, fases da Lua e destino mágico",
        A30="◇ Caos/banimento em magia; não deus universal",
        A31="— Sem Diabo cristão; Wicca não é satanismo",
        A32="● Ciclo sazonal, renascimento e reencarnação em muitas correntes",
        A33="● Círculo, guardiões dos quadrantes e divindades patronas",
        A34="● Canto, dança, poesia e artes mágicas",
        A35="● Elevação de energia/transe; intoxicantes não são requisito",
        A36="≈ Disciplina iniciática e jejum ocasional; não ascetismo central",
        A37="◇ Reverência à vida e ética de não causar dano; formulações variam",
        A38="● Sacerdotisa/sacerdote e experiência direta",
        A39="≈ Espírito/alma e reencarnação; teologias variam",
        A40="● Polaridades complementares em formas tradicionais; revisão contemporânea",
        A41="≈ Harmonia, autoconhecimento e Summerland; sem salvação única",
        A42="— Sem escatologia final normativa",
        A43="● Esbats, sabás, oferendas e iniciação",
        A44="● Iniciação e magia transformativa",
    ),
    "thelema": mapped(
        A01="● Nuit: continuidade infinita; Hadit: ponto de experiência",
        A02="≈ Nuit/Hadit e manifestação; não criação teísta simples",
        A03="● Ra-Hoor-Khuit no Aeon de Hórus; teologia aberta",
        A04="● Thelema/Verdadeira Vontade",
        A05="● Nuit e Babalon em funções distintas; não terra-mãe",
        A06="≈ Ra-Hoor-Khuit e forças marcianas/solares",
        A07="● Ra-Hoor-Khuit e Heru-ra-ha",
        A08="● Nuit e Babalon em simbolismo lunar/estelar",
        A09="● Nuit: estrelas; astrologia hermética",
        A10="≈ Babalon e 'mar' simbólico",
        A11="● Hadit e fogo/serpente; elementos rituais",
        A12="≈ Hadit/sopro e elemento Ar",
        A13="≈ Babalon/terra em correspondências, não agricultura",
        A14="● To Mega Therion e formas bestiais simbólicas",
        A15="● Babalon, Nuit e sexualidade sacramental em correntes",
        A16="≈ Criança Coroada e Conquistadora; não família",
        A17="● Ra-Hoor-Khuit",
        A18="● Verdadeira Vontade e Aeon de Hórus; não programa estatal único",
        A19="● Thoth, Aiwass e Conhecimento e Conversação",
        A20="● Magick e sistema de correspondências",
        A21="● Aiwass e Hermes/Thoth",
        A22="≈ Babalon e Júpiter em magia; não núcleo ético",
        A23="≈ Raphael e práticas herméticas; não substitui medicina",
        A24="— Sem entidade normativa da doença",
        A25="● Aniquilação no 'Abismo' e continuidade estelar; interpretações variam",
        A26="● Santos da Ecclesia Gnostica Catholica e linhagem mágica",
        A27="● Santo Anjo Guardião",
        A28="● Ma'at/ajuste e Verdadeira Vontade",
        A29="● Aeons de Ísis/Osíris/Hórus são esquema esotérico, não periodização histórica",
        A30="● Choronzon e Abismo",
        A31="● Choronzon como dispersão; não diabo absoluto",
        A32="● Morte/iniciação e novo Aeon",
        A33="● Santo Anjo Guardião",
        A34="● Rito, poesia, teatro e arte mágica",
        A35="● Êxtase, magia sexual e ritos; substâncias não são requisito universal",
        A36="● Disciplina mágica e A∴A∴",
        A37="◇ Amor sob vontade",
        A38="● Aiwass e Liber AL",
        A39="● Khabs/Khu e conceitos de estrela individual",
        A40="● Nuit/Hadit e amor/vontade",
        A41="● Descoberta e realização da Verdadeira Vontade",
        A42="● Aeon de Hórus como escatologia esotérica aberta",
        A43="● Missa Gnóstica, iniciação e sacramentos",
        A44="● Iniciação, travessia do Abismo e Grande Obra",
    ),
    "rastafari": mapped(
        A01="● Jah",
        A02="● Jah",
        A03="● Jah; Haile Selassie recebe interpretações diversas",
        A04="● Livity e justiça",
        A05="≈ Mama Africa como pátria/mãe simbólica",
        A06="● Jah governa criação; sem deus do trovão",
        A07="● Luz de Jah; símbolos solares/etíopes",
        A10="≈ Rivers of Babylon e água em linguagem bíblica",
        A11="≈ Fogo como purificação/julgamento em linguagem bíblica",
        A12="≈ Breath/livity",
        A13="● Ital e cultivo natural",
        A14="● Leão de Judá como símbolo",
        A15="● Amor/one love e livity",
        A16="● Família/comunidade; mansões divergem sobre gênero",
        A17="● Resistência a Babylon; não deus da guerra",
        A18="● Haile Selassie e linhagem salomônica em interpretação rastafari",
        A19="● Bíblia, reasoning e anciãos",
        A21="● Profetas bíblicos e Marcus Garvey como precursor",
        A22="◇ Rejeição do materialismo de Babylon",
        A23="● Ital, ervas e cuidado; não substitui medicina",
        A24="◇ Babylon como sistema opressivo, não doença",
        A25="◇ Ênfase em vida/livity; concepções de morte variam",
        A26="● Ancestralidade africana e memória",
        A28="● Justiça de Jah",
        A29="◇ Tempo vivido e repatriação; sem destino personificado",
        A30="● Babylon como ordem opressiva",
        A31="≈ Babylon/Diabo em linguagem bíblica",
        A32="● Repatriação e redenção; leituras sobre reencarnação variam",
        A33="● Jah",
        A34="● Nyabinghi, reggae e salmodia",
        A35="● Ganja como sacramento em muitas comunidades; não universal",
        A36="● Ital e disciplina",
        A37="● One love, justiça e comunidade",
        A38="● Bíblia, profetas, Garvey e Selassie em papéis distintos",
        A39="● I-and-I e livity",
        A40="● I-and-I supera separação indivíduo/comunidade",
        A41="● Libertação de Babylon e Zion/repatriação",
        A42="≈ Queda de Babylon e redenção",
        A43="● Reasoning, Nyabinghi e comunhão",
        A44="● Conversão de consciência e livity",
    ),
    "bahai": mapped(
        A01="● Deus incognoscível e eterno",
        A02="● Deus; criação contínua por sua vontade",
        A03="● Deus",
        A04="● Revelação progressiva, aliança e justiça",
        A05="— Sem deusa-terra",
        A06="● Natureza como criação/sinal; sem deus do trovão",
        A07="● Luz como metáfora da revelação",
        A08="● Lua como metáfora/criação; não divindade",
        A09="● Astres como criação e metáforas",
        A10="● Oceanos/água como metáfora da revelação",
        A11="● Fogo como metáfora de amor/provação",
        A12="● Espírito como sopro/metáfora",
        A13="● Natureza e agricultura sob responsabilidade ética",
        A14="● Unidade da criação e bondade aos animais",
        A15="● Amor de Deus e unidade humana",
        A16="● Família, igualdade e casamento",
        A17="◇ Luta espiritual/serviço; religião deve promover paz",
        A18="● Deus e Ordem Administrativa; não teocracia sacerdotal",
        A19="● Manifestações de Deus e investigação independente da verdade",
        A20="● Trabalho realizado em espírito de serviço como adoração",
        A21="● Manifestações de Deus",
        A22="◇ Justiça social, trabalho e moderação",
        A23="● Oração e medicina; ambas valorizadas",
        A24="◇ Prova e condição material; não entidade espiritual da doença",
        A25="● Continuidade da alma após a morte",
        A26="● Memória dos mortos; oração, não culto ancestral",
        A27="● Manifestações/ensinamentos como guia",
        A28="● Justiça e misericórdia divinas; progresso da alma",
        A29="● Revelação progressiva; ciclos proféticos",
        A30="◇ Desunião/ignorância, não caos personificado",
        A31="— Sem diabo pessoal necessário; 'Satanás' como ego inferior",
        A32="● Progresso da alma; ressurreição frequentemente espiritual",
        A33="● Deus e assistência espiritual",
        A34="● Música e artes como elevação",
        A35="● Oração/meditação; intoxicantes proibidos salvo uso médico",
        A36="● Jejum anual; não monasticismo",
        A37="● Unidade, serviço e misericórdia",
        A38="● Báb, Bahá'u'lláh e Manifestações de Deus",
        A39="● Alma racional imortal",
        A40="● Unidade da humanidade e harmonia ciência-religião",
        A41="● Aproximação de Deus e progresso da alma",
        A42="● Maturidade/unidade da humanidade; não fim destrutivo literal",
        A43="● Oração, jejum e serviço; sem sacerdócio",
        A44="● Transformação pessoal e social",
    ),
    "cao_dai": mapped(
        A01="● Cao Đài/Đức Chí Tôn: Ser Supremo",
        A02="● Cao Đài",
        A03="● Cao Đài; Olho Divino como símbolo",
        A04="● Đại Đạo e unidade dos Três Ensinamentos",
        A05="● Đức Phật Mẫu: Mãe Divina",
        A06="≈ Divindades taoistas no panteão sincrético",
        A07="● Olho Divino/luz; deidades solares sincréticas",
        A08="≈ Mãe Divina e cosmologia yin",
        A09="● Hierarquias astrais e santos",
        A10="≈ Deidades taoistas/budistas das águas",
        A11="≈ Fogo ritual e hierarquias espirituais",
        A12="≈ Khí e espírito",
        A13="◇ Vegetarianismo e respeito à vida",
        A15="● Amor universal",
        A16="● Mãe Divina e ética familiar",
        A17="◇ Combate espiritual; não deus da guerra central",
        A18="● Hierarquia religiosa e governo divino",
        A19="● Cao Đài, santos e médiuns",
        A21="● Espíritos/santos comunicantes",
        A22="◇ Mérito e serviço",
        A23="≈ Cura espiritual em contextos; não substitui medicina",
        A25="● Mundos espirituais e karma",
        A26="● Ancestrais e santos",
        A27="● Guias/espíritos e médiuns",
        A28="● Karma e julgamento",
        A29="● Ciclos de revelação e Terceira Anistia",
        A30="◇ Desordem da era, não caos divino",
        A31="≈ Espíritos inferiores; não adversário soberano",
        A32="● Reencarnação",
        A33="● Santos e espíritos tutelares",
        A34="● Música e ritual templário",
        A35="● Sessões mediúnicas históricas; disciplina atual institucional",
        A36="● Vegetarianismo, disciplina e clero",
        A37="● Amor universal",
        A38="● Revelação mediúnica e fundadores",
        A39="● Alma e evolução",
        A40="● Síntese yin-yang e Três Ensinamentos",
        A41="● Libertação do ciclo e união com Deus",
        A42="● Terceira Era/Anistia",
        A43="● Culto, oferendas e hierarquia clerical",
        A44="● Iniciação e progresso espiritual",
    ),
    "tenrikyo": mapped(
        A01="● Tenri-O-no-Mikoto/Oyagami: Deus-Pai/Mãe",
        A02="● Oyagami cria os seres humanos para a Vida Plena de Alegria",
        A03="● Oyagami",
        A04="● Ordem divina, causalidade e hinokishin",
        A05="● Oyagami como Pai-Mãe; não deusa-terra",
        A06="● Dez providências divinas governam funções naturais",
        A07="≈ Providências de luz/calor; não deidade solar separada",
        A08="≈ Lua/Sol aparecem nas providências; não panteão astral",
        A10="≈ Água como providência divina",
        A11="≈ Fogo/calor como providência",
        A12="≈ Respiração como empréstimo de Deus",
        A13="● Corpo/natureza como empréstimo; agricultura na prática histórica",
        A15="● Alegria e amor parental de Deus",
        A16="● Oyagami e criação parental; família humana",
        A17="◇ Superação da mente egocêntrica, não guerra divina",
        A18="● Oyagami; Oyasama como Santuário de Deus",
        A19="● Miki Nakayama/Oyasama e Ofudesaki",
        A21="● Oyasama como mediadora/reveladora",
        A22="◇ Hinokishin e contentamento",
        A23="● Salvação e cura no contexto fundacional; cuidado médico não excluído",
        A24="◇ Doença como orientação em doutrina histórica; evitar culpabilização",
        A25="● Retorno/partida e continuidade da alma",
        A26="● Memória de Oyasama e ancestrais",
        A27="● Oyasama e ensinamentos",
        A28="● Causalidade e purificação da mente",
        A29="● Causalidade e renascimento",
        A30="◇ Poeiras da mente, não caos personificado",
        A31="— Sem diabo central",
        A32="● Renascimento e restauração",
        A33="● Oyagami e proteção",
        A34="● Serviço (Tsutome), música e dança",
        A35="● Serviço e oração; sem intoxicação ritual",
        A36="◇ Disciplina alegre, não ascese extrema",
        A37="● Vida Plena de Alegria e ajuda mútua",
        A38="● Oyasama e escrituras",
        A39="● Alma; corpo como empréstimo",
        A40="● Deus-Pai/Mãe e ajuda mútua",
        A41="● Yōki gurashi: Vida Plena de Alegria",
        A42="≈ Restauração do mundo ideal, não apocalipse destrutivo",
        A43="● Serviço, oferendas e santuário",
        A44="● Purificação das poeiras da mente",
    ),
    "lds": mapped(
        A01="● Deus e inteligências/elementos eternos em cosmologia SUD",
        A02="● Pai Celestial por meio de Jesus Cristo",
        A03="● Pai Celestial; Divindade inclui Filho e Espírito Santo",
        A04="● Plano de salvação, lei e arbítrio",
        A05="≈ Mãe Celestial é afirmada, mas há pouca elaboração litúrgica",
        A06="● Deus governa natureza; sem deus do trovão",
        A07="● Cristo/luz de Cristo",
        A08="● Lua como criação e grau de glória em metáfora escritural",
        A09="● Estrelas como criação e metáfora de glórias",
        A10="● Batismo por imersão",
        A11="● Espírito Santo/fogo",
        A12="● Espírito e sopro; não divindade do vento",
        A13="◇ Mordomia da criação e autossuficiência",
        A14="◇ Animais na criação/ressurreição em ensinamentos",
        A15="● Amor divino e casamento",
        A16="● Família eterna e selamento",
        A17="● Guerra no céu e guerreiros escriturais; ética institucional",
        A18="● Deus, sacerdócio e reino",
        A19="● Cristo, Espírito Santo e revelação contínua",
        A20="◇ Trabalho e criação como parte da exaltação",
        A21="● Morôni, anjos e profetas",
        A22="◇ Providência, dízimo e mordomia",
        A23="● Bênção de saúde e medicina",
        A24="◇ Doença/provação; não entidade divina",
        A25="● Mundo dos espíritos e morte",
        A26="● Obra vicária e genealogia pelos mortos; não culto ancestral",
        A27="● Cristo, anjos e missionários no mundo espiritual",
        A28="● Cristo julga; graus de glória",
        A29="● Pré-existência, mortalidade e eternidade; arbítrio",
        A30="● Guerra no céu e fim da mortalidade",
        A31="● Satanás/Lúcifer",
        A32="● Ressurreição universal",
        A33="● Deus, anjos e sacerdócio",
        A34="● Hinos e artes devocionais",
        A35="● Espírito/revelação; álcool e drogas são proibidos",
        A36="● Jejum e disciplina; não monasticismo",
        A37="● Caridade e serviço",
        A38="● Joseph Smith e profetas vivos; Livro de Mórmon",
        A39="● Espírito como filho de Pais Celestiais; corpo e espírito",
        A40="● Unidade da Divindade em propósito; corpo/espírito",
        A41="● Salvação e exaltação",
        A42="● Segunda vinda, Milênio, julgamento e graus de glória",
        A43="● Expiação de Cristo, sacramento e ordenanças do templo",
        A44="● Batismo, investidura, selamento e santificação",
    ),
    "satanism_nontheist": mapped(
        A01="◇ Natureza/materialidade; sem absoluto sobrenatural obrigatório",
        A02="— Sem criador sobrenatural",
        A03="— Satanás é símbolo, não divindade, no satanismo LaVeyano/TST",
        A04="◇ Autonomia, razão e ética conforme a organização",
        A05="◇ Natureza/sexualidade sem deusa-terra normativa",
        A06="— Sem deus do trovão",
        A07="◇ Luz/conhecimento como metáfora",
        A08="◇ Noite como estética/símbolo",
        A11="◇ Fogo como símbolo ritual",
        A14="◇ Animalidade humana valorizada no LaVeyanismo",
        A15="● Sexualidade consentida e prazer; formulações variam",
        A17="◇ Assertividade/autodefesa",
        A18="◇ Autossoberania",
        A19="● Conhecimento, ceticismo e magia como psicodrama em correntes",
        A20="◇ Técnica/arte ritual humana",
        A21="● Satanás como arquétipo de oposição/transgressão",
        A22="◇ Sucesso e responsabilidade individual no LaVeyanismo",
        A23="◇ Cuidado baseado em evidências; ritual não substitui medicina",
        A24="— Sem demônio de doença",
        A25="◇ Morte como fim da consciência em correntes materialistas",
        A28="◇ Justiça e reciprocidade conforme códigos da organização",
        A29="◇ Finitude e agência",
        A30="◇ Caos/transgressão como símbolo",
        A31="● Satanás como símbolo adversarial; não mal sobrenatural",
        A32="— Sem ressurreição normativa",
        A33="◇ Autoproteção e comunidade",
        A34="● Estética, música e teatro ritual",
        A35="● Ritual psicodramático; não exige intoxicação",
        A36="◇ Autodisciplina, não ascese",
        A37="◇ Compaixão/empatia varia entre organizações",
        A38="● Textos/fundadores modernos; não revelação divina necessária",
        A39="◇ Mente corporal/material em correntes nontheístas",
        A40="◇ Razão/paixão integradas",
        A41="◇ Autonomia e vida plena, não salvação",
        A42="— Sem escatologia normativa",
        A43="● Ritual simbólico sem sacrifício real",
        A44="● Autotransformação ritual",
    ),
    "discordian": mapped(
        A01="● Caos e Eris em registro satírico",
        A02="≈ Eris cria/desvela desordem em mitopoese paródica",
        A03="● Eris, tratada simultaneamente com seriedade e sátira",
        A04="◇ Anerístico/erístico e crítica de categorias",
        A05="≈ Eris como deusa, não terra-mãe",
        A15="◇ Jogo e liberdade; sem deidade do amor",
        A17="≈ Maçã da Discórdia e conflito mítico",
        A18="◇ 'Todo homem, mulher e criança é Papa': anti-autoridade",
        A19="● Eris e Principia Discordia",
        A21="● Eris, Malaclypse e tricksterismo",
        A22="≈ Cinco toneladas de linho: símbolo absurdo",
        A25="◇ Morte tratada humoristicamente; sem doutrina única",
        A28="◇ Lei dos Cincos e contra-regras",
        A29="● Calendário discordiano",
        A30="● Eris/caos",
        A31="◇ Greyface como símbolo da rigidez, não mal absoluto",
        A32="◇ Recorrência criativa do caos",
        A33="≈ Santos discordianos e autoatribuição papal",
        A34="● Humor, colagem, performance e paródia",
        A35="● Jogo, absurdo e transgressão; intoxicação não é requisito",
        A36="— Ascese não é central",
        A37="◇ Liberação por humor; ética descentralizada",
        A38="● Principia Discordia como texto fundador/paródico",
        A39="◇ Mente que impõe grades à realidade",
        A40="● Erístico/anerístico como complementaridade crítica",
        A41="◇ Iluminação/liberação por descondicionamento humorístico",
        A42="— Sem escatologia normativa",
        A43="● Rituais auto-inventados/paródicos",
        A44="● Mudança de perspectiva",
    ),
    "new_age": mapped(
        A01="≈ Fonte/Uno/Consciência; vocabulário descentralizado",
        A02="≈ Fonte, mente cósmica ou cocriadores; sem doutrina comum",
        A03="≈ Eu Superior/Mestres; não autoridade única",
        A04="≈ Karma, vibração e 'leis' espirituais; afirmações variam",
        A05="≈ Gaia/Divino Feminino",
        A06="≈ Elementais e astrologia; sem consenso",
        A07="≈ Luz/energia solar como metáfora",
        A08="≈ Ciclos lunares e Divino Feminino",
        A09="≈ Astrologia, seres estelares e planetas",
        A10="≈ Elemento água e cura energética",
        A11="≈ Energia/fogo e transformação",
        A12="≈ Prāṇa/respiração apropriados de tradições diversas",
        A13="≈ Gaia, herbalismo e espiritualidade ecológica",
        A14="≈ Animais de poder; apropriação/neoshamanismo requer crítica",
        A15="≈ Amor universal e sexualidade sagrada",
        A16="≈ Criança interior/Divino Feminino",
        A17="≈ Guerreiro de luz; metáfora",
        A18="≈ Mestres ascensos e hierarquias canalizadas",
        A19="≈ Eu Superior, registros akáshicos e mestres",
        A20="≈ Cristais/tecnologias espirituais sem validação científica",
        A21="≈ Guias, anjos e seres canalizados",
        A22="≈ Manifestação/abundância; não consenso e risco de culpabilização",
        A23="≈ Cura energética como crença; não substitui medicina",
        A24="≈ 'Bloqueio energético' não é diagnóstico clínico",
        A25="≈ Planos sutis e pós-vida",
        A26="≈ Ancestrais/vidas passadas",
        A27="≈ Guias espirituais",
        A28="≈ Karma e 'lei da atração'; conceitos não equivalentes",
        A29="≈ Eras astrológicas e ciclos; não cronologia acadêmica",
        A30="≈ Caos/transição de frequência",
        A31="≈ Ego/energia negativa; sem adversário comum",
        A32="≈ Reencarnação e ascensão",
        A33="≈ Anjos, guias e cristais",
        A34="≈ Música meditativa, mandalas e arte visionária",
        A35="≈ Meditação, canalização e psicodélicos em subculturas; não universal",
        A36="≈ Detox/retreats; não disciplina comum",
        A37="≈ Amor universal/compaixão",
        A38="≈ Canalizadores, mestres e revelação pessoal",
        A39="≈ Eu Superior, alma e consciência cósmica",
        A40="≈ Masculino/feminino e luz/sombra; modelos frequentemente essencialistas",
        A41="≈ Ascensão, despertar e autorrealização",
        A42="≈ Mudança de era/ascensão planetária; crenças descentralizadas",
        A43="≈ Cerimônias e práticas ecléticas",
        A44="≈ Despertar, cura e transformação pessoal",
    ),
    "jurema": mapped(
        A01="● Deus/Olorum e Reino da Jurema; teologias sincréticas",
        A02="● Deus; Jurema é árvore/força/reino, não criadora única",
        A03="● Deus/Olorum",
        A04="◇ Ciência da Jurema, memória e reciprocidade",
        A05="● Jurema sagrada e Mães/Mestras; terra/mata",
        A06="≈ Caboclos e encantados; funções locais",
        A07="≈ Mestres/encantados solares em linhas específicas",
        A08="≈ Mestras e ciclos lunares; não universal",
        A10="● Encantados das águas e cidades da Jurema",
        A11="● Fumaça, vela e tronco/força vegetal",
        A12="● Fumaça/sopro ritual",
        A13="● Jurema (Mimosa tenuiflora) e plantas de poder",
        A14="● Caboclos, mata e animais encantados",
        A15="≈ Mestras e Pombagiras em correntes cruzadas; varia",
        A16="● Ancestrais, caboclos e comunidade",
        A17="● Caboclos/Mestres guerreiros",
        A18="● Reis, mestres e cidades encantadas",
        A19="● Mestres/Mestras e ciência da Jurema",
        A20="● Cachimbo, tronqueira e artes rituais",
        A21="● Mestres, Exus e mensageiros em tradições cruzadas",
        A22="≈ Mestres e linhas de prosperidade",
        A23="● Mestres, caboclos, ervas e cura",
        A24="◇ Aflição espiritual é categoria ritual; não substitui diagnóstico",
        A25="● Encantamento e mestres falecidos/encantados",
        A26="● Ancestrais indígenas/afro-indígenas e mestres",
        A27="● Mestres e caboclos guiam entre mundos",
        A28="◇ Justiça de mestres/encantados; varia",
        A29="◇ Encantamento suspende morte linear; sem destino único",
        A30="≈ Esquerda/forças bravas não equivalem a mal",
        A31="— Sem adversário absoluto comum",
        A32="● Encantamento e continuidade espiritual",
        A33="● Caboclos, mestres e cidades tutelares",
        A34="● Pontos, maracá, canto e dança",
        A35="● Transe e bebida da Jurema em linhas específicas",
        A38="● Mestres/Mestras, sonhos e incorporação",
        A39="● Espírito/encantado e força da Jurema",
        A40="● Catimbó, Umbanda e matrizes indígenas/africanas em combinações locais",
        A41="◇ Cura, proteção e relação com encantados",
        A43="● Mesa, cachimbo, bebida, oferendas e iniciação",
        A44="● Tombamento/iniciação e encantamento",
    ),
    "entheogenic_christian": mapped(
        A13="● Plantas sacramentais e floresta em teologias amazônicas",
        A23="≈ Cura espiritual comunitária; não substitui medicina",
        A35="● Ayahuasca/Daime/Vegetal como sacramento em contexto ritual",
        A38="● Bíblia, hinários/revelações e liderança da tradição",
        A39="● Alma/espírito e experiência visionária cristã",
        A43="● Sacramento enteogênico, canto e disciplina",
        A44="● Concentração, miração e transformação moral",
    ),
    "modern_nontheist": mapped(
        A01="◇ Natureza/cosmos sem sobrenatural",
        A02="— Sem criador",
        A03="— Sem divindade suprema",
        A04="● Razão, dignidade e ética humana",
        A05="◇ Terra/ecologia sem personificação necessária",
        A13="◇ Responsabilidade ecológica",
        A14="◇ Animais como sujeitos de consideração ética",
        A15="● Amor e sexualidade sob consentimento/dignidade",
        A16="● Família em formas plurais",
        A17="◇ Defesa de direitos, não guerra sagrada",
        A18="● Autoridade democrática e direitos humanos",
        A19="● Razão, ciência e crítica",
        A20="● Técnica humana",
        A21="◇ Comunicação e sátira",
        A22="◇ Justiça distributiva e bem-estar",
        A23="● Medicina baseada em evidências",
        A24="● Doença como fenômeno biopsicossocial",
        A25="● Morte como finitude; posições pessoais variam",
        A26="● Memória dos mortos sem culto obrigatório",
        A28="● Justiça humana e responsabilidade",
        A29="◇ Tempo natural e agência",
        A30="◇ Caos como conceito natural/matemático, não potência sagrada",
        A31="— Sem mal personificado necessário",
        A32="◇ Renovação social/biológica, não ressurreição normativa",
        A33="● Solidariedade e instituições",
        A34="● Arte como criação humana",
        A35="◇ Estados alterados estudados/éticos, não sacramento comum",
        A36="◇ Autodisciplina sem ascese religiosa",
        A37="● Compaixão, direitos e cuidado",
        A38="— Sem revelação sobrenatural",
        A39="◇ Consciência como questão filosófica/científica",
        A40="● Pluralismo e equilíbrio crítico",
        A41="◇ Florescimento humano, não salvação sobrenatural",
        A42="— Sem escatologia normativa",
        A43="— Sem sacrifício/mediação sobrenatural",
        A44="● Educação e transformação social/pessoal",
    ),
    "african_local": mapped(
        A01="≈ Fonte/origem conforme cosmologia local",
        A02="≈ Criador ou modelador local; documentação variável",
        A03="≈ Ser supremo/céu em muitas, não todas, as tradições",
        A04="◇ Ordem de parentesco, território e reciprocidade",
        A05="≈ Terra/fertilidade e maternidade; nomes locais necessários",
        A06="≈ Trovão/chuva em cultos locais",
        A07="≈ Sol como potência; varia",
        A08="≈ Lua e calendário; varia",
        A10="≈ Rios, lagoas e espíritos aquáticos",
        A11="≈ Fogo doméstico/forja e especialistas",
        A13="≈ Cultivos e fertilidade",
        A14="≈ Caça, animais e floresta",
        A16="≈ Linhagem, parto e família",
        A17="≈ Ancestrais/heróis e defesa",
        A18="≈ Chefia sagrada e ancestrais",
        A19="≈ Adivinhação e anciãos/especialistas",
        A20="≈ Ferreiros e corporações de ofício",
        A21="≈ Mensageiros/tricksters locais; não universais",
        A22="≈ Prosperidade, gado ou colheita",
        A23="● Especialistas, plantas e cultos de cura",
        A24="≈ Entidades de aflição também podem curar",
        A25="● Mundo dos mortos e ritos funerários",
        A26="● Ancestrais e linhagens",
        A27="≈ Especialistas/espíritos guias",
        A28="≈ Juramento, ancestrais e retribuição",
        A29="≈ Destino/nome/linhagem; varia",
        A31="— Não presumir diabo ou mal absoluto",
        A32="≈ Retorno ancestral/renascimento em algumas tradições",
        A33="● Espíritos tutelares de pessoa, linhagem e lugar",
        A34="● Música, dança e oralidade ritual",
        A35="● Possessão/transe em muitas tradições; não universal",
        A38="● Adivinhos, médiuns e transmissão oral",
        A39="≈ Componentes múltiplos da pessoa; terminologia local",
        A40="◇ Reciprocidade entre vivos, mortos e território",
        A41="◇ Bem-viver e continuidade relacional; não salvação única",
        A43="● Oferenda, libação e mediação",
        A44="● Iniciação e incorporação comunitária",
    ),
    "korean": mapped(
        A01="≈ Haneullim/Ch'ŏnju e cosmologias locais; tradição plural",
        A02="≈ Hwanin/Hwanung em mito de fundação; não criador universal",
        A03="● Hanŭnim/Haneullim e Imperador de Jade em camadas",
        A05="● Jowang, Samsin e deusas da terra/família",
        A06="● Yongwang e deuses do trovão/tempo",
        A07="≈ Haenim: Sol",
        A08="≈ Dalnim: Lua",
        A10="● Yongwang: Rei Dragão e águas",
        A11="● Jowangsin: fogo/cozinha",
        A13="● Teojusin e deuses agrários",
        A14="● Sansin: montanha e animais",
        A16="● Samsin Halmoni e espíritos domésticos",
        A17="● Chaksin/generais e heróis divinizados",
        A18="● Haneullim e ancestrais fundadores",
        A19="● Mudang e deuses de escrita/saber",
        A21="● Jeseok e mensageiros; tricksters em narrativas",
        A22="● Eopsin: prosperidade",
        A23="● Mudang e deuses de cura",
        A24="● Espíritos de doença em gut específicos",
        A25="● Daebyeol-wang/Jeoseung e mundo dos mortos",
        A26="● Josangsin: ancestrais",
        A27="● Bari-degi: guia/salvadora dos mortos",
        A28="● Reis do outro mundo e retribuição",
        A29="≈ Chilseong e destino",
        A30="≈ Espíritos ressentidos/caóticos; podem ser pacificados",
        A31="— Sem mal absoluto único",
        A32="● Pacificação, renovação e ciclos ancestrais",
        A33="● Seongju, Samsin, Sansin e espíritos domésticos",
        A34="● Gut, música e dança",
        A35="● Possessão e êxtase do mudang",
        A38="● Mudang e oráculos",
        A39="● Hon/baek e nexos espirituais; vocabulário sincrético",
        A40="● Yin-yang, vivos/mortos e pacificação",
        A43="● Gut e oferendas",
        A44="● Naerim-gut e iniciação",
    ),
    "vietnamese": mapped(
        A01="≈ Trời e cosmologias compartilhadas sino-vietnamitas",
        A02="≈ Ông Trời/Nữ Oa em camadas; sem relato único",
        A03="● Ngọc Hoàng/Ông Trời",
        A04="◇ Thiên mệnh, reciprocidade e đạo",
        A05="● Mẫu Địa e sistema das Mães",
        A06="● Mẫu Thượng Thiên e deuses do trovão",
        A07="● Deidade solar e culto celeste",
        A08="● Deidade lunar e calendários",
        A10="● Mẫu Thoải e Reis Dragões",
        A11="● Táo Quân e deuses do fogo/cozinha",
        A12="≈ Mẫu Thượng Ngàn/vento em tradições",
        A13="● Thần Nông e Mẫu Thượng Ngàn",
        A14="● Espíritos de montanha/floresta",
        A16="● Mẫu e culto familiar",
        A17="● Trần Hưng Đạo e heróis divinizados",
        A18="● Imperador de Jade e heróis nacionais",
        A19="● Văn Xương e ancestrais letrados",
        A21="● Mensageiros e espíritos médiuns",
        A22="● Thần Tài",
        A23="● Santos/médiuns e deuses de cura",
        A24="≈ Espíritos de aflição; culto apotropaico",
        A25="● Diêm Vương e mundos dos mortos",
        A26="● Culto ancestral",
        A27="● Quan Âm, Địa Tạng e guias",
        A28="● Karma/tribunais e ordem celeste",
        A29="● Táo Quân, estrelas e destino",
        A30="≈ Espíritos errantes e calamidade",
        A31="— Sem mal absoluto único",
        A32="● Retorno ancestral e renascimento em camadas budistas",
        A33="● Thành hoàng e deuses tutelares",
        A34="● Hát chầu văn e artes rituais",
        A35="● Lên đồng e possessão",
        A38="● Médiuns e revelação",
        A39="● Hồn/vía e linhagem",
        A40="● Três/Quatro Palácios e yin-yang",
        A41="≈ Harmonia ancestral, mérito e libertação budista",
        A43="● Culto ancestral, oferendas e lên đồng",
        A44="● Iniciação mediúnica",
    ),
    "indonesian": mapped(
        A01="≈ Origem/cosmos conforme povo e região",
        A02="≈ Criadores/ancestrais locais; sem sistema único",
        A03="≈ Divindade celeste suprema em muitas tradições",
        A04="◇ Adat e equilíbrio cósmico-social",
        A05="≈ Mãe Terra/Deusa do arroz em tradições locais",
        A06="≈ Deuses da tempestade e ancestrais",
        A07="≈ Sol em pares celestes",
        A08="≈ Lua/calendário",
        A10="● Espíritos do mar, rios e ancestrais aquáticos",
        A11="≈ Fogo doméstico e forja",
        A13="● Dewi Sri e equivalentes locais do arroz",
        A14="● Espíritos da floresta/animais e ancestrais",
        A16="● Ancestrais e casas",
        A17="≈ Heróis/ancestrais guerreiros",
        A18="≈ Chefia/ancestralidade sagrada",
        A19="● Anciãos, sacerdotes e adivinhação",
        A20="● Tecelagem, casas e ofícios rituais",
        A21="≈ Semar/tricksters em tradições javanesas; não pan-indonésio",
        A22="≈ Dewi Sri e ancestrais de prosperidade",
        A23="● Dukun/balian e plantas",
        A24="≈ Espíritos de aflição; também curados ritualmente",
        A25="● Mundos ancestrais e funerais",
        A26="● Ancestrais e casas de origem",
        A27="≈ Sacerdotes/ancestrais guiam",
        A28="◇ Adat, reciprocidade e sanções",
        A29="● Calendários agrícolas e destino",
        A30="≈ Espíritos perigosos e desequilíbrio",
        A31="— Sem mal absoluto comum",
        A32="● Renovação agrícola e continuidade ancestral",
        A33="● Espíritos de casa, aldeia e território",
        A34="● Gamelão, dança, teatro e tecelagem ritual",
        A35="● Possessão/transe em tradições específicas",
        A38="● Sacerdotes, médiuns e narrativas",
        A39="≈ Alma/força vital; terminologia local",
        A40="● Pares alto/baixo, masculino/feminino e adat",
        A41="◇ Equilíbrio e continuidade ancestral",
        A43="● Oferenda, festa e sacrifício conforme a tradição",
        A44="● Iniciação e ritos de passagem",
    ),
    "mandaean": mapped(
        A01="● Hayyi Rabbi: Grande Vida e Mundo da Luz",
        A02="● Hayyi Rabbi; Ptahil como demiurgo subordinado",
        A03="● Hayyi Rabbi",
        A04="● Kušṭa: verdade/aliança",
        A05="≈ Ruha em cosmologia, frequentemente ambivalente; não mãe-terra",
        A07="● Luz e uthras",
        A10="● Yardna: água viva/batismal",
        A11="≈ Fogo pertence ao mundo material/ritual funerário em contextos",
        A12="● Ruha: espírito/vento, categoria complexa",
        A19="● Manda d-Hayyi: Conhecimento da Vida",
        A21="● Uthras e Manda d-Hayyi",
        A25="● Mundo das Trevas e estações pós-morte",
        A27="● Uthras guiam a alma",
        A28="● Balanças/estações e purificação",
        A29="● Alma percorre ciclos/estações",
        A30="● Mundo das Trevas",
        A31="● Ruha e Ur em funções adversas; sistema complexo",
        A32="● Ascensão da alma; batismos repetidos",
        A33="● Uthras protetores",
        A36="● Disciplina sacerdotal e pureza",
        A37="● Vida, verdade e caridade",
        A38="● João Batista como profeta central; uthras reveladores",
        A39="● Nišimta e ruha: componentes distintos",
        A40="● Luz/Trevas e direita/esquerda",
        A41="● Retorno ao Mundo da Luz",
        A42="≈ Consumação e ascensão; textos variam",
        A43="● Masbuta, masiqta e sacerdócio",
        A44="● Batismo repetido e ascensão",
    ),
    "yazidi": mapped(
        A01="● Xwedê/Deus",
        A02="● Deus cria o mundo e o confia a sete seres",
        A03="● Deus; Tawûsê Melek é chefe dos Sete Anjos",
        A04="● Ordem divina, tradição oral e pacto comunitário",
        A05="≈ Terra e santos locais; sem deusa-terra central",
        A06="≈ Anjos governam aspectos do mundo; detalhes variam",
        A07="● Sol/luz em oração e simbolismo; não divindade independente",
        A08="≈ Lua em calendário e tradição",
        A10="● Kaniya Spî e Lalish: água sagrada",
        A11="● Fogo/luz sagrados",
        A12="≈ Sopro/espírito",
        A13="≈ Tawûsê Melek e santos em fertilidade/proteção",
        A14="● Pavão como símbolo de Tawûsê Melek",
        A16="● Linhagens, castas religiosas e família",
        A17="≈ Sheikh Adi e heróis/santos",
        A18="● Deus, Tawûsê Melek e autoridade religiosa",
        A19="● Qewwals, qewls e tradição oral",
        A21="● Sete Anjos e santos",
        A22="≈ Santos e bênção",
        A23="● Santuários, água e santos de cura",
        A25="● Morte, transmigração em algumas formulações e outro mundo",
        A26="● Sheikh Adi, santos e ancestrais",
        A27="● Anjos/santos",
        A28="● Deus e responsabilidade; detalhes escatológicos variam",
        A29="● Ciclos, calendário e destino",
        A30="— Tawûsê Melek não é Satanás; não impor mito de queda",
        A31="— Rejeitar equivalência de Tawûsê Melek com Diabo",
        A32="≈ Transmigração/retorno em tradições; formulações variam",
        A33="● Tawûsê Melek e santos",
        A34="● Qewls e música dos qewwals",
        A35="● Peregrinação e experiência ritual; sem enteógeno",
        A36="● Jejuns e disciplina comunitária",
        A37="● Hospitalidade e dever comunitário",
        A38="● Qewls, santos e tradição",
        A39="● Alma; detalhes internos variam",
        A40="● Sete Anjos em ordem unificada, não dualismo",
        A41="● Proximidade de Deus e continuidade comunitária",
        A43="● Peregrinação a Lalish, lâmpadas, água e oferendas",
        A44="● Batismo/ritos de passagem e peregrinação",
    ),
    "native_american_church": mapped(
        A03="● Deus/Grande Espírito em formulações cristãs-indígenas diversas",
        A11="● Fogo central e cedro",
        A23="● Peyote como medicina/sacramento; não substitui cuidados médicos",
        A35="● Peyote em cerimônia noturna regulamentada",
        A37="● Vida ética, família, sobriedade e comunidade",
        A38="● Bíblia e revelação/visão em combinações locais",
        A43="● Reunião de peyote, água, fogo, canto e oração",
        A44="● Cura e compromisso comunitário",
    ),
    "new_thought": mapped(
        A01="● Mente/Espírito infinito; formulações variam",
        A02="● Deus/Mente divina cria por lei mental",
        A03="● Deus como presença/princípio",
        A04="● Lei mental/espiritual e pensamento correto",
        A05="≈ Feminino divino em correntes posteriores; não núcleo histórico",
        A07="● Luz/consciência como metáfora",
        A12="≈ Espírito/respiração",
        A15="● Amor divino e afirmação",
        A18="◇ Autodomínio pela mente",
        A19="● Mente divina, Bíblia metafísica e professores",
        A20="◇ Cocriação/manifestação",
        A21="● Cristo/Eu-Cristo e professores",
        A22="● Prosperidade em algumas escolas; risco de culpabilização",
        A23="● Cura mental/espiritual como crença histórica; não substituir medicina",
        A24="≈ Pensamento não deve ser usado para culpar pessoas por doença",
        A25="≈ Continuidade espiritual; escolas variam",
        A28="● Lei de causa mental/espiritual",
        A29="◇ Presente e desenvolvimento",
        A31="◇ Erro/ignorância, não diabo necessário",
        A32="● Renovação da mente",
        A33="● Presença divina",
        A35="● Oração afirmativa/meditação",
        A36="◇ Disciplina mental",
        A37="● Amor, perdão e unidade",
        A38="● Jesus reinterpretado e professores modernos",
        A39="● Mente/consciência espiritual",
        A40="● Espírito/matéria como expressão, conforme escola",
        A41="● Realização espiritual e vida abundante",
        A42="≈ Nova era de consciência, não consenso",
        A43="● Oração e culto; sem sacrifício",
        A44="● Transformação do pensamento",
    ),
    "scientology": mapped(
        A01="◇ Existência e theta; sem criador obrigatório na prática",
        A02="— Questão de Ser Supremo é deixada à descoberta individual",
        A03="≈ Oitava dinâmica/Ser Supremo, não doutrina teísta detalhada",
        A04="● Dinâmicas, ética e sobrevivência",
        A07="◇ Theta como vida/consciência; não luz solar",
        A15="◇ Relações e dinâmicas",
        A16="● Família como dinâmica",
        A18="◇ Autodeterminação e organização",
        A19="● L. Ron Hubbard e tecnologia de estudo/auditing",
        A20="● 'Tecnologia espiritual' em autodescrição; alegações contestadas",
        A21="● Auditor como facilitador",
        A22="◇ Sobrevivência e prosperidade",
        A23="≈ Dianética/auditing como crença; não substitui medicina",
        A24="≈ Engramas são conceito doutrinário, não diagnóstico clínico",
        A25="● Thetan continua após a morte",
        A26="≈ Vidas passadas; não culto ancestral",
        A27="● Auditor/treinamento",
        A28="● Ética e condições",
        A29="● Vidas passadas e linha do tempo",
        A30="≈ Mente reativa/engramas",
        A31="≈ Pessoas supressivas em doutrina; não mal sobrenatural",
        A32="● Reencarnação/retorno do thetan",
        A33="◇ Autodeterminação e organização",
        A35="● Auditing e estados alterados de atenção; drogas são rejeitadas",
        A36="● Treinamento e disciplina organizacional",
        A37="◇ Ajuda/sobrevivência; avaliação ética externa necessária",
        A38="● Escritos de Hubbard",
        A39="● Thetan",
        A40="◇ Mente analítica/reativa",
        A41="● Clear e Operating Thetan",
        A42="— Sem escatologia final central",
        A43="● Auditing, cursos e cerimônias",
        A44="● Ponte para a Liberdade Total",
    ),
    "ufo_religion": mapped(
        A01="≈ Inteligência cósmica/fonte; varia por movimento",
        A02="≈ Extraterrestres criadores/engenheiros em alguns grupos",
        A03="≈ Hierarquias cósmicas",
        A04="≈ Evolução cósmica e lei universal",
        A05="≈ Terra/Gaia como planeta vivo em alguns grupos",
        A07="≈ Seres de luz e simbolismo solar",
        A09="● Extraterrestres, planetas e civilizações estelares",
        A15="≈ Amor cósmico",
        A17="≈ Conflito cósmico em narrativas específicas",
        A18="≈ Conselhos/irmandades espaciais",
        A19="● Conhecimento extraterrestre/canalizado",
        A20="● Tecnologia avançada como elemento salvífico",
        A21="● Contatados/canalizadores",
        A22="≈ Abundância pós-tecnológica",
        A23="≈ Cura energética/extraterrestre alegada; não substitui medicina",
        A25="≈ Ascensão/transferência; varia",
        A27="● Guias extraterrestres",
        A28="≈ Seleção/evolução conforme movimento",
        A29="● Eras cósmicas e datas proféticas",
        A30="≈ Catástrofe/transição planetária",
        A31="≈ Extraterrestres hostis em alguns grupos",
        A32="≈ Ascensão, clonagem ou evacuação; varia",
        A33="≈ Irmandades/seres protetores",
        A34="≈ Arte e música cósmica",
        A35="● Canalização/contato; substâncias não são universais",
        A36="≈ Disciplina de preparação",
        A37="≈ Amor universal",
        A38="● Contatados e mensagens",
        A39="≈ Consciência/alma cósmica",
        A40="≈ Humano/cósmico e matéria/consciência",
        A41="● Ascensão/evolução/contato conforme movimento",
        A42="● Transformação planetária ou resgate",
        A43="≈ Meditação, transmissão e encontros",
        A44="● Preparação para mudança cósmica",
    ),
    "esoteric_order": mapped(
        A01="≈ Absoluto/Uno em sínteses herméticas",
        A02="≈ Logos/demiurgo e emanação",
        A03="≈ Grande Arquiteto/Logos conforme ordem; não consenso entre membros",
        A04="● Correspondência, ordem e lei hermética",
        A05="≈ Shekhinah/Ísis/terra em simbolismo",
        A06="≈ Júpiter/forças elementais em correspondências",
        A07="● Sol/Tiphareth em simbolismo",
        A08="● Lua/Yesod em simbolismo",
        A09="● Astrologia e planetas",
        A10="● Elemento Água",
        A11="● Elemento Fogo e alquimia",
        A12="● Elemento Ar",
        A13="● Elemento Terra e ciclos",
        A14="≈ Simbolismo animal",
        A15="● Vênus e união simbólica",
        A16="≈ Masculino/feminino simbólicos; modelos históricos",
        A17="● Marte e disciplina",
        A18="● Adeptos/hierarquias iniciáticas",
        A19="● Hermes/Thoth, Cabala e gnose",
        A20="● Alquimia e magia cerimonial",
        A21="● Hermes e guardiões de limiar",
        A22="● Júpiter/Vênus em prosperidade ritual",
        A23="≈ Raphael e cura hermética; não substitui medicina",
        A24="— Sem entidade universal de doença",
        A25="● Saturno e morte simbólica",
        A26="● Cadeia iniciática/ancestrais da ordem",
        A27="● Hermes/Anúbis em correspondências",
        A28="● Maat/balança e juramentos iniciáticos",
        A29="● Saturno, astrologia e ciclos",
        A30="● Caos/alquimia e dissolução",
        A31="≈ Qliphoth/adversários em sistemas; não uniformes",
        A32="● Morte e renascimento iniciáticos",
        A33="● Pentagramas, anjos e círculos",
        A34="● Arte, arquitetura, música e rito",
        A35="● Visão, magia e êxtase ritual; substâncias não necessárias",
        A36="● Disciplina e graus",
        A37="◇ Serviço/fraternidade conforme ordem",
        A38="● Textos, mestres e iniciação",
        A39="● Alma/corpos sutis em sínteses",
        A40="● Solve/coagula, pilares e polaridades",
        A41="● Grande Obra/iluminação",
        A42="≈ Ciclos e nova era; varia",
        A43="● Iniciação, juramento e ritual",
        A44="● Transmutação e iniciação",
    ),
}

# Profiles declared above can inherit a well-documented parent and then replace
# only the cells whose terminology or ritual context changes.
PROFILES["afro_yoruba_diaspora"] = PROFILES["yoruba"] | PROFILES["afro_yoruba_diaspora"]
PROFILES["entheogenic_christian"] = PROFILES["christian"] | PROFILES["entheogenic_christian"]
PROFILES["native_american_church"] = PROFILES["north_american"] | PROFILES["native_american_church"]


SOURCES = [
    ("M01", "Método", "Comparação exige categorias explícitas e não prova origem comum", "Nota editorial desta revisão; consultar também obras de história das religiões", "", "Princípio metodológico"),
    ("M02", "Pré-história", "Re-animating Hunter-gatherer Rock-art Research", "Cambridge Archaeological Journal", "https://www.cambridge.org/core/journals/cambridge-archaeological-journal/article/reanimating-huntergatherer-rockart-research/FE6F4E068D776C186C10A5848B8264E8", "Crítica ao uso automático de 'xamanismo' na arte rupestre"),
    ("M03", "Pré-história", "Psychoactive Substances in Prehistoric Times", "Time and Mind / Taylor & Francis", "https://www.tandfonline.com/doi/full/10.1080/1751696X.2014.993244", "Revisão de evidências arqueológicas; não autoriza universalização"),
    ("A01", "Mesopotâmia", "Electronic Text Corpus of Sumerian Literature", "University of Oxford", "https://etcsl.orinst.ox.ac.uk/", "Textos sumérios, traduções e bibliografia"),
    ("A02", "Mesopotâmia", "Open Richly Annotated Cuneiform Corpus", "ORACC", "https://oracc.museum.upenn.edu/", "Corpora cuneiformes anotados"),
    ("A03", "Egito", "The Art of Ancient Egypt: A Resource for Educators", "Metropolitan Museum of Art", "https://www.metmuseum.org/-/media/files/learn/for-educators/publications-for-educators/the-art-of-ancient-egypt.pdf", "Visão geral de religião, arte e deidades"),
    ("A04", "Grécia/Roma", "Perseus Digital Library: Hesiod, Theogony", "Tufts University", "https://atlas.perseus.tufts.edu/library/urn:cts:greekLit:tlg0020.tlg001/", "Texto e traduções da Teogonia"),
    ("A05", "Irã", "Encyclopaedia Iranica: Zoroastrianism", "Encyclopaedia Iranica Foundation", "https://www.iranicaonline.org/articles/zoroastrianism-i-historical-review/", "História e teologia zoroastrianas"),
    ("A06", "Índia", "GRETIL — Göttingen Register of Electronic Texts in Indian Languages", "University of Göttingen", "https://gretil.sub.uni-goettingen.de/gretil.html", "Textos indianos em línguas originais"),
    ("A07", "Budismo antigo", "SuttaCentral", "SuttaCentral Development Trust", "https://suttacentral.net/about", "Textos budistas antigos, paralelos e traduções"),
    ("A08", "Budismo tibetano", "84000: Translating the Words of the Buddha", "84000", "https://84000.co/", "Traduções do cânone tibetano"),
    ("A09", "China", "Chinese Text Project", "Chinese Text Project", "https://ctext.org/introduction", "Base aberta de textos chineses pré-modernos"),
    ("A10", "Daoismo", "Religious Daoism", "Stanford Encyclopedia of Philosophy", "https://plato.stanford.edu/entries/daoism-religion/", "História, rituais e panteões daoistas"),
    ("A11", "Japão", "Encyclopedia of Shinto", "Kokugakuin University", "https://d-museum.kokugakuin.ac.jp/eos/", "Referência acadêmica sobre kami, ritos e escolas"),
    ("I01", "Māori", "Traditional Māori religion — ngā karakia a te Māori", "Te Ara Encyclopedia of New Zealand", "https://teara.govt.nz/en/traditional-maori-religion-nga-karakia-a-te-maori", "Atua, Te Kore, Rangi/Papa, mana, tapu e mauri"),
    ("I02", "Austrália", "AIATSIS Education resources", "Australian Institute of Aboriginal and Torres Strait Islander Studies", "https://aiatsis.gov.au/education", "Diversidade e protocolos de conhecimento indígena"),
    ("I03", "Yorùbá/diáspora", "Ifá divination system", "UNESCO Intangible Cultural Heritage", "https://ich.unesco.org/en/RL/ifa-divination-system-00146", "Sistema de Ifá e corpus oral"),
    ("I04", "Yorùbá/diáspora", "Music and African Spiritual Traditions", "Smithsonian Institution", "https://www.si.edu/spotlight/music-and-spirituality/spirituality-and-african-religious-traditions", "Objetos, música e tradições africanas"),
    ("I05", "Santería/Lukumí", "Santeria", "Smithsonian Folklife Festival", "https://festival.si.edu/articles/1989/santeria", "Formação afro-cubana e orichas"),
    ("I06", "Mexica", "Zona Arqueológica Templo Mayor", "Instituto Nacional de Antropología e Historia", "https://inah.gob.mx/zonas/zona-arqueologica-templo-mayor", "Culto estatal mexica e deidades do Templo Mayor"),
    ("I07", "Mexica", "Religión mexica, resultado de la fusión de dioses", "Instituto Nacional de Antropología e Historia", "https://inah.gob.mx/boletines/religion-mexica-resultado-de-la-fusion-de-dioses-de-otras-culturas", "Sincretismo e complexidade do panteão mexica"),
    ("B01", "Judaísmo", "Sefaria Library", "Sefaria", "https://www.sefaria.org/texts", "Tanakh, textos rabínicos e comentários"),
    ("B02", "Cristianismo", "Bíblia e documentos", "Vatican", "https://www.vatican.va/archive/index.htm", "Textos e documentos católicos; comparar com fontes confessionais próprias"),
    ("B03", "Islã", "Corpus Coranicum", "Berlin-Brandenburg Academy of Sciences", "https://corpuscoranicum.de/", "Texto, contexto e documentação do Qurʾān"),
    ("B04", "Siquismo", "Sri Guru Granth Sahib", "Shiromani Gurdwara Parbandhak Committee", "https://old.sgpc.net/CDN/Siri%20Guru%20Granth%20Sahib%2C%20Romanized.pdf", "Texto do Guru Granth Sahib em transliteração"),
    ("B05", "Bahá'í", "Bahá'í Reference Library", "Bahá'í International Community", "https://www.bahai.org/library/", "Fonte oficial de escritos bahá'ís"),
    ("C01", "Espiritismo", "Obras básicas e estudo", "Federação Espírita Brasileira", "https://www.febnet.org.br/portal/", "Fonte confessional; confrontar com pesquisa histórica"),
    ("C02", "Tenrikyo", "The Teachings and History of Tenrikyo", "Tenrikyo Church Headquarters", "https://www.tenrikyo.or.jp/eng/teaching/", "Fonte oficial sobre doutrina e história"),
    ("C03", "Santos dos Últimos Dias", "Gospel Library", "The Church of Jesus Christ of Latter-day Saints", "https://www.churchofjesuschrist.org/study?lang=eng", "Fonte oficial; confrontar com pesquisa histórica"),
    ("C04", "Thelema", "Liber AL vel Legis", "Ordo Templi Orientis USA", "https://oto-usa.org/thelema/liber-al/", "Texto primário do movimento"),
    ("C05", "Satanismo LaVeyano", "The Church of Satan", "Church of Satan", "https://www.churchofsatan.com/", "Fonte confessional; distingue Satanás simbólico"),
    ("C06", "Novos movimentos", "World Religions and Spirituality Project", "Virginia Commonwealth University", "https://wrldrels.org/", "Perfis acadêmicos de movimentos religiosos novos"),
]


TRADITIONS: list[Tradition] = []


def add(
    name: str,
    family: str,
    region: str,
    period: str,
    kind: str,
    profile: str,
    sources: str,
    *,
    status: str = "Viva",
    coverage: str = "Perfil de família",
    note: str = "",
    overrides: dict[str, str] | None = None,
) -> None:
    TRADITIONS.append(
        Tradition(
            name=name,
            family=family,
            region=region,
            period=period,
            kind=kind,
            profile=profile,
            sources=sources,
            status=status,
            coverage=coverage,
            note=note,
            overrides=overrides or {},
        )
    )


def add_many(
    names: Iterable[str],
    family: str,
    region: str,
    period: str,
    kind: str,
    profile: str,
    sources: str,
    *,
    status: str = "Viva",
    coverage: str = "Perfil de família",
    note: str = "",
) -> None:
    for name in names:
        add(name, family, region, period, kind, profile, sources, status=status, coverage=coverage, note=note)


# Evidência arqueológica e religiões históricas. Datas indicam atestação,
# não "nascimento" de povos nem duração total de tradições.
add("Evidências paleolíticas de comportamento ritual", "Arqueologia da religião", "África/Eurásia", "c. 100.000–10.000 a.C.", "Evidência material, não religião nomeada", "archaeological", "M02; M03", status="Arqueológica", coverage="Crítico", note="Nenhum panteão ou cosmovisão específica pode ser recuperado com segurança.")
add("Complexos neolíticos da Europa e Anatólia", "Arqueologia da religião", "Europa/Anatólia", "c. 10.000–3.000 a.C.", "Evidência material", "archaeological", "M02; M03", status="Arqueológica", coverage="Crítico")
add("Göbekli Tepe", "Arqueologia da religião", "Anatólia", "c. 9.600–8.200 a.C.", "Sítio/complexo ritual", "archaeological", "M02", status="Arqueológica", coverage="Crítico")
add("Çatalhöyük", "Arqueologia da religião", "Anatólia", "c. 7.100–5.950 a.C.", "Sítio/complexo ritual", "archaeological", "M02", status="Arqueológica", coverage="Crítico")
add("Religião suméria", "Mesopotâmica", "Mesopotâmia meridional", "c. 2.600–1.800 a.C. (corpus)", "Politeísta histórica", "sumerian", "A01; A02", status="Histórica", coverage="Detalhado")
add("Religião acádia", "Mesopotâmica", "Mesopotâmia", "c. 2.300–2.000 a.C.", "Politeísta histórica", "akkadian", "A02", status="Histórica", coverage="Detalhado")
add("Religião babilônica", "Mesopotâmica", "Mesopotâmia", "c. 1.900–539 a.C.", "Politeísta histórica", "akkadian", "A02", status="Histórica", coverage="Detalhado", overrides={"A03": "● Marduk: soberania babilônica", "A18": "● Marduk, Šamaš e ideologia régia"})
add("Religião assíria", "Mesopotâmica", "Mesopotâmia setentrional", "c. 2.000–609 a.C.", "Politeísta histórica", "akkadian", "A02", status="Histórica", coverage="Detalhado", overrides={"A03": "● Aššur", "A18": "● Aššur legitima a monarquia assíria"})
add("Religião egípcia antiga", "Egípcia", "Vale do Nilo", "c. 3.000 a.C.–séculos IV–VI d.C.", "Politeísta histórica", "egyptian", "A03", status="Histórica", coverage="Detalhado")
add("Atonismo de Akhenaton", "Egípcia", "Egito", "c. 1353–1336 a.C.", "Culto real henoteísta/monolátrico", "egyptian", "A03", status="Histórica", coverage="Parcial", note="Reforma curta; chamar de monoteísmo pleno é debatido.", overrides={"A03": "● Aten como foco exclusivo do culto real", "A07": "● Aten: disco solar e raios vivificantes", "A18": "● Akhenaton e família real como mediadores privilegiados"})
add("Religião cananeia/ugarítica", "Semítica ocidental antiga", "Levante", "c. 1.500–1.200 a.C. (Ugarit)", "Politeísta histórica", "ugaritic", "A02", status="Histórica", coverage="Detalhado")
add("Religião fenícia", "Semítica ocidental antiga", "Levante/Mediterrâneo", "c. 1.200–300 a.C.", "Politeísta histórica", "ugaritic", "A02", status="Histórica", coverage="Perfil de família", note="Cidades possuíam panteões e patronos distintos.")
add("Religião púnica/cartaginesa", "Semítica ocidental antiga", "Norte da África/Mediterrâneo", "c. 800 a.C.–século II d.C.", "Politeísta histórica", "ugaritic", "A02", status="Histórica", coverage="Perfil de família", overrides={"A03": "● Baal Hammon e Tanit no culto cartaginês", "A05": "● Tanit; funções exatas e iconografia requerem contexto"})
add("Religião hitita-hurrita", "Anatólia antiga", "Anatólia/Síria", "c. 1.650–1.200 a.C.", "Politeísta histórica", "hittite", "A02", status="Histórica", coverage="Detalhado")
add_many(["Religião luvita", "Religião urartiana", "Religião elamita"], "Religiões do Oriente Próximo antigo", "Anatólia/Cáucaso/Irã", "II–I milênio a.C.", "Politeísta histórica", "fragmentary", "A02", status="Histórica", coverage="Fragmentário")
add("Religião minoica", "Egeia antiga", "Creta", "c. 2.000–1.450 a.C.", "Religião arqueológica", "fragmentary", "A04", status="Histórica", coverage="Fragmentário", note="Linear A não foi decifrada; nomes e mitos são inseguros.")
add("Religião micênica", "Grega antiga", "Grécia e Egeu", "c. 1.600–1.100 a.C.", "Politeísta histórica", "greek", "A04", status="Histórica", coverage="Perfil de família", note="Tábuas Linear B atestam nomes divinos, mas não uma mitologia completa.")
add("Religião grega arcaica e clássica", "Grega antiga", "Mediterrâneo", "c. 800–146 a.C.", "Politeísta histórica", "greek", "A04", status="Histórica", coverage="Detalhado")
add("Religiões helenísticas", "Helenística", "Mediterrâneo/Ásia", "323–30 a.C. e além", "Politeísmos e cultos sincréticos", "greek", "A04", status="Histórica", coverage="Perfil de família")
add("Orfismo", "Mistérios gregos", "Mediterrâneo", "c. século VI a.C.–Antiguidade tardia", "Movimentos iniciáticos", "greek", "A04", status="Histórica", coverage="Parcial", overrides={"A01": "● Noite/Chronos/Ovo conforme teogonias órficas", "A36": "● Regras de pureza e vida órfica em testemunhos", "A41": "● Libertação do ciclo e destino melhor para iniciados"})
add("Mistérios eleusinos", "Mistérios gregos", "Eleusis/Mediterrâneo", "c. século VII a.C.–392 d.C.", "Culto iniciático", "greek", "A04", status="Histórica", coverage="Parcial", overrides={"A13": "● Deméter e Perséfone", "A32": "● Retorno de Perséfone e esperança iniciática", "A35": "? Kykeon é atestado; conteúdo psicoativo permanece hipótese", "A44": "● Iniciação eleusina"})
add("Religião romana", "Romana antiga", "Mediterrâneo/Império Romano", "c. 500 a.C.–século V d.C.", "Politeísta cívica", "roman", "A04", status="Histórica", coverage="Detalhado")
add("Religião etrusca", "Itálica antiga", "Itália", "c. 800–100 a.C.", "Politeísta histórica", "roman", "A04", status="Histórica", coverage="Perfil de família", note="Fontes são fragmentárias e mediadas por autores romanos.")
add("Culto de Mithras", "Mistérios romanos", "Império Romano", "c. século I–IV d.C.", "Culto iniciático", "fragmentary", "A04; A05", status="Histórica", coverage="Fragmentário", note="Não confundir automaticamente com Mithra iraniano; não há 'Bíblia mitraica' preservada.")
add_many(["Religião trácia", "Religião dácica", "Religião ilíria", "Religião antiga dos albaneses", "Religião ibérica", "Religião celtibera"], "Paleobalcânica/europeia", "Europa", "I milênio a.C.", "Politeísta histórica", "fragmentary", "A04", status="Histórica", coverage="Fragmentário")
add("Religiões célticas continentais", "Céltica", "Europa ocidental", "I milênio a.C.–séculos I–IV d.C.", "Politeísmos históricos", "celtic", "A04", status="Histórica", coverage="Perfil de família")
add("Tradições mitológicas irlandesas e galesas", "Céltica insular", "Irlanda/Grã-Bretanha", "manuscritos medievais; materiais anteriores", "Corpus literário cristianizado", "celtic", "A04", status="Histórica/Revival", coverage="Detalhado")
add("Religião nórdica antiga", "Germânica", "Escandinávia/Atlântico Norte", "c. 750–1100 d.C.; fontes séc. XIII", "Politeísta histórica", "norse", "A04", status="Histórica/Revival", coverage="Detalhado")
add("Religião germânica continental/anglo-saxã", "Germânica", "Europa setentrional", "Antiguidade tardia–Idade Média", "Politeísmos históricos", "norse", "A04", status="Histórica/Revival", coverage="Perfil de família")
add("Religião fino-careliana", "Urálica", "Finlândia/Carélia", "oral; registros sobretudo modernos", "Tradição oral histórica", "finnic", "A04", status="Viva/Revival", coverage="Detalhado")
add("Religião eslava pré-cristã", "Eslava", "Europa oriental", "fontes sobretudo medievais e folclóricas", "Politeísmos históricos", "slavic", "A04", status="Histórica/Revival", coverage="Detalhado")
add("Religiões bálticas históricas", "Báltica", "Lituânia/Letônia/Prússia", "fontes medievais e dainas posteriores", "Politeísmos históricos", "baltic", "A04", status="Histórica/Revival", coverage="Detalhado")
add_many(["Religião armênia pré-cristã", "Religião georgiana pré-cristã", "Religião vainakh", "Religião circassiana", "Religião osseta tradicional"], "Caucasiana/iraniana", "Cáucaso", "Antiguidade–presente/avivamento", "Tradições históricas e vivas", "fragmentary", "A05", status="Histórica/Revival", coverage="Fragmentário")
add("Religião védica", "Indo-ariana", "Sul da Ásia", "c. 1.500–500 a.C. (composição)", "Religião sacrificial histórica", "vedic", "A06", status="Histórica/continuidade", coverage="Detalhado")
add("Zoroastrismo/Mazdeísmo", "Iraniana", "Irã/Índia/diáspora", "Gāthās: data debatida; tradição viva", "Religião viva", "zoroastrian", "A05", coverage="Detalhado")
add("Zurvanismo", "Iraniana", "Irã", "atestado sobretudo período sassânida", "Corrente histórica", "zoroastrian", "A05", status="Histórica", coverage="Parcial", overrides={"A01": "● Zurvān/Tempo ilimitado em fontes zurvanitas", "A29": "● Zurvān", "A40": "● Ohrmazd/Ahriman como gêmeos em versões zurvanitas"})
add_many(["Religião cita", "Religião sármata", "Religião sogdiana"], "Iraniana antiga", "Estepe/Ásia Central", "I milênio a.C.–I milênio d.C.", "Politeísmos históricos", "fragmentary", "A05", status="Histórica", coverage="Fragmentário")
add("Religião árabe pré-islâmica", "Semítica árabe", "Península Arábica", "até o século VII d.C.", "Politeísmos históricos", "fragmentary", "B03", status="Histórica", coverage="Fragmentário")
add_many(["Religião nabateia", "Religião palmirena"], "Semítica antiga", "Arábia/Levante", "I milênio a.C.–Antiguidade tardia", "Politeísmos históricos", "fragmentary", "B03", status="Histórica", coverage="Fragmentário")
add("Gnosticismos da Antiguidade tardia", "Gnóstica", "Mediterrâneo/Oriente Próximo", "séculos I–IV d.C.", "Família de movimentos", "gnostic", "B01; B02", status="Histórica/Revival", coverage="Detalhado", note="Não houve uma única 'religião gnóstica'.")
add("Maniqueísmo", "Maniqueia", "Eurásia", "século III–XIV d.C.", "Religião universal histórica", "manichaean", "A05", status="Histórica", coverage="Detalhado")
add("Mandaísmo", "Mandaica", "Iraque/Irã/diáspora", "Antiguidade tardia–presente", "Religião viva", "mandaean", "C06", coverage="Detalhado")
add("Yazidismo", "Curda/Mesopotâmica", "Iraque/Síria/Armênia/diáspora", "formação medieval; tradição viva", "Religião viva", "yazidi", "C06", coverage="Detalhado")

# Tradições africanas e afro-diaspóricas. Entradas de família permanecem
# separadas para evitar o rótulo homogeneizante "subsaariano".
add("Religião Yorùbá e Ifá", "Yorùbá", "Nigéria/Benim/diáspora", "oral e documentação moderna; tradição viva", "Religião tradicional viva", "yoruba", "I03; I04", coverage="Detalhado")
add("Religião Akan", "Akan", "Gana/Costa do Marfim", "oral e documentação moderna", "Religião tradicional viva", "akan", "I04", coverage="Detalhado")
add("Vodun Fon-Ewe", "Fon/Ewe", "Benim/Togo/Gana", "oral e documentação moderna", "Religião tradicional viva", "vodun", "I04", coverage="Detalhado")
add_many(
    ["Odinani (Igbo)", "Religião Kongo", "Religião Serer", "Religião Dogon", "Religião Bambara/Bamana", "Religião Mossi", "Religião Dagara", "Religião Edo/Benin", "Religião Nupe", "Religião Tiv"],
    "África Ocidental/Central",
    "África Ocidental/Central",
    "tradições orais vivas",
    "Religiões tradicionais vivas",
    "african_local",
    "I04",
    coverage="Perfil de família",
    note="Necessita revisão por corpus, povo e linhagem; nomes locais não são intercambiáveis.",
)
add_many(
    ["Religião Dinka", "Religião Nuer", "Religião Maasai", "Waaqeffanna (Oromo)", "Religião Somali tradicional/Waaq", "Religião Sidama", "Religião Kalenjin", "Religião Acholi", "Religião Baganda"],
    "África Oriental",
    "África Oriental",
    "tradições orais vivas",
    "Religiões tradicionais vivas",
    "african_local",
    "I04",
    coverage="Perfil de família",
)
add_many(
    ["Religiões San", "Religião Zulu tradicional", "Religião Xhosa tradicional", "Religião Shona", "Religião Lozi", "Religião Bemba", "Religião Chewa", "Religião Tswana", "Religião Ovambo", "Religião Herero"],
    "África Austral",
    "África Austral",
    "tradições orais vivas",
    "Religiões tradicionais vivas",
    "african_local",
    "I04",
    coverage="Perfil de família",
)
add_many(
    ["Religião tradicional malgaxe", "Religião Amazigh/Berbere tradicional", "Religião Guanche", "Religião núbia/kushita", "Cultos Mami Wata", "Bwiti", "Bori Hausa", "Culto Zar"],
    "África/Índico",
    "África e Madagascar",
    "históricas e vivas",
    "Tradições locais e cultos transregionais",
    "african_local",
    "I04",
    status="Viva/Histórica",
    coverage="Perfil de família",
)
add("Candomblé Ketu/Nagô", "Afro-brasileira Yorùbá", "Brasil", "século XIX–presente", "Religião afro-diaspórica", "afro_yoruba_diaspora", "I03; I04", coverage="Detalhado")
add("Candomblé Jeje", "Afro-brasileira Fon-Ewe", "Brasil", "século XIX–presente", "Religião afro-diaspórica", "vodun", "I03; I04", coverage="Perfil de família")
add("Candomblé Angola/Congo", "Afro-brasileira Bantu", "Brasil", "século XIX–presente", "Religião afro-diaspórica", "african_local", "I04", coverage="Perfil de família", note="Inquices e terminologia bantu não devem ser substituídos automaticamente por orixás.")
add("Umbanda", "Afro-brasileira/sincrética", "Brasil", "c. 1908–presente; origens plurais", "Religião brasileira", "umbanda", "C01; I04", coverage="Detalhado")
add("Quimbanda", "Afro-brasileira", "Brasil", "séculos XIX–XXI; múltiplas linhagens", "Religião/culto de entidades", "umbanda", "I04", coverage="Parcial", note="Não é sinônimo de 'magia negra' nem simples subdivisão uniforme da Umbanda.")
add_many(
    ["Batuque do Rio Grande do Sul", "Tambor de Mina", "Xangô do Recife", "Terecô", "Omolocô", "Jarê da Chapada Diamantina"],
    "Afro-brasileira",
    "Brasil",
    "séculos XIX–XXI",
    "Religiões afro-brasileiras",
    "afro_yoruba_diaspora",
    "I03; I04",
    coverage="Perfil de família",
    note="Cada tradição possui nações, casas e vocabulários próprios.",
)
add("Jurema Sagrada/Catimbó-Jurema", "Afro-indígena brasileira", "Nordeste do Brasil", "séculos XVIII–XXI", "Religião brasileira", "jurema", "I04", coverage="Detalhado")
add("Vodou haitiano", "Afro-haitiana", "Haiti/diáspora", "séculos XVII–XXI", "Religião afro-diaspórica", "haitian_vodou", "I04", coverage="Detalhado")
add("Louisiana Voodoo", "Afro-norte-americana", "Louisiana/EUA", "séculos XVIII–XXI", "Tradição afro-diaspórica", "haitian_vodou", "I04", coverage="Perfil de família", note="Não é idêntico ao Vodou haitiano.")
add("Regla de Ocha/Lukumí (Santería)", "Afro-cubana Yorùbá", "Cuba/diáspora", "séculos XIX–XXI", "Religião afro-diaspórica", "afro_yoruba_diaspora", "I05", coverage="Detalhado")
add_many(
    ["Palo Monte/Palo Mayombe", "Abakuá", "Ifá cubano", "Orisha/Shango de Trinidad", "Kumina jamaicana", "Obeah (complexo de práticas)", "Spiritual Baptist/Shouter", "Revival Zion jamaicano"],
    "Afro-caribenha",
    "Caribe/diáspora",
    "séculos XVIII–XXI",
    "Religiões e complexos rituais",
    "african_local",
    "I04; I05",
    coverage="Perfil de família",
)

# Povos e tradições indígenas das Américas. As entradas genéricas são
# deliberadamente marcadas como perfis de família, não panteões completos.
add_many(
    ["Religiões Inuit", "Religiões Yupik", "Religião Unangan/Aleúte", "Tradições Tlingit", "Tradições Haida", "Tradições Coast Salish"],
    "Indígena do Ártico/Pacífico Norte",
    "América do Norte",
    "tradições orais vivas",
    "Religiões indígenas",
    "north_american",
    "I02",
    coverage="Perfil de família",
)
add_many(
    ["Tradição Lakota", "Tradição Cheyenne", "Tradição Blackfoot/Niitsitapi", "Tradições Anishinaabe", "Religião Haudenosaunee", "Tradição Diné/Navajo", "Religião Hopi", "Religião Zuni", "Tradições Pueblo", "Religião Cherokee", "Religião Muscogee/Creek", "Religião Choctaw", "Religião Pawnee", "Religião Caddo"],
    "Indígena norte-americana",
    "América do Norte",
    "tradições orais vivas",
    "Religiões indígenas",
    "north_american",
    "I02",
    coverage="Perfil de família",
    note="Não usar 'Grande Espírito' como tradução universal; revisar terminologia na língua da nação.",
)
add("Native American Church/Peyotismo", "Indígena-cristã", "América do Norte", "século XIX–presente", "Religião enteogênica", "native_american_church", "M03", coverage="Detalhado")
add("Religião maia clássica", "Mesoamericana", "Mesoamérica", "c. 250–900 d.C. (período clássico)", "Politeísta histórica", "maya", "I06", status="Histórica/continuidade", coverage="Detalhado")
add("Tradições K'iche' e Popol Wuj", "Maia", "Guatemala", "oral; manuscrito colonial do séc. XVI", "Tradição indígena viva", "maya", "I06", coverage="Detalhado")
add("Religião mexica", "Nahua/Mesoamericana", "México central", "c. 1325–1521; fontes coloniais", "Politeísta histórica", "mexica", "I06; I07", status="Histórica/Revival", coverage="Detalhado")
add_many(
    ["Religião mixteca", "Religião zapoteca", "Religião purépecha", "Religião totonaca", "Religião huasteca", "Religiões nahuas vivas", "Religião otomí"],
    "Mesoamericana",
    "México/Mesoamérica",
    "pré-colonial–presente",
    "Religiões indígenas",
    "indigenous_local",
    "I06; I07",
    status="Viva/Histórica",
    coverage="Perfil de família",
)
add_many(
    ["Religião olmeca (reconstrução arqueológica)", "Religião de Teotihuacan (reconstrução arqueológica)", "Religião tolteca (fontes tardias)"],
    "Mesoamericana arqueológica",
    "Mesoamérica",
    "II milênio a.C.–século XII d.C.",
    "Evidência arqueológica/histórica",
    "archaeological",
    "I06; I07",
    status="Arqueológica/Histórica",
    coverage="Fragmentário",
)
add_many(["Religião Taíno", "Religião Kalinago/Carib", "Religiões indígenas das Guianas"], "Indígena caribenha", "Caribe/Guianas", "pré-colonial–presente/avivamento", "Religiões indígenas", "indigenous_local", "I02", status="Viva/Revival", coverage="Perfil de família")
add("Religião inca estatal", "Andina", "Andes", "séculos XV–XVI; tradições anteriores", "Politeísta histórica", "andean", "I06", status="Histórica/continuidade", coverage="Detalhado")
add_many(
    ["Tradições Quechua", "Tradições Aymara", "Religião Mapuche", "Religião Muisca", "Religião Kogi/Arhuaco", "Tradição Warao", "Tradição Yanomami", "Tradições Tukano", "Tradição Shipibo-Konibo", "Tradições Asháninka", "Tradições Shuar/Achuar", "Tradições Kayapó/Mebêngôkre", "Tradições Xavante/A'uwe", "Tradições Guajajara/Tenetehara"],
    "Indígena sul-americana",
    "América do Sul",
    "tradições orais vivas",
    "Religiões indígenas",
    "indigenous_local",
    "I02",
    coverage="Perfil de família",
)
add("Tradições Guarani (ênfase Mbyá/Kaiowá)", "Guarani", "Brasil/Paraguai/Argentina/Bolívia", "tradição oral viva", "Religião indígena", "guarani", "I02", coverage="Detalhado")
add("Tradições Tupi históricas", "Tupi", "Brasil", "fontes coloniais e continuidades indígenas", "Religiões indígenas", "indigenous_local", "I02", status="Histórica/continuidade", coverage="Perfil de família", note="Não fundir Tupinambá, Tupiniquim e povos atuais em um único panteão.")

# Oceania e Ásia setentrional.
add("Tradições aborígenes australianas (visão agregada)", "Aborígene australiana", "Austrália", "tradições orais vivas; registro colonial moderno", "Famílias religiosas indígenas", "australian", "I02", coverage="Detalhado com ressalvas", note="Entrada-guia; há centenas de nações, línguas e Dreamings.")
add_many(["Tradições Yolŋu", "Tradições Arrernte", "Tradições Warlpiri", "Tradições Noongar", "Tradições Tiwi", "Tradições dos povos do Estreito de Torres"], "Indígena australiana", "Austrália/Estreito de Torres", "tradições orais vivas", "Religiões indígenas", "australian", "I02", coverage="Perfil de família")
add("Religião Māori", "Polinésia", "Aotearoa/Nova Zelândia", "tradição oral; registro escrito séc. XIX", "Religião indígena viva", "polynesian", "I01", coverage="Detalhado")
add_many(
    ["Religião havaiana", "Religião samoana", "Religião tonganesa", "Religião taitiana", "Religião Rapa Nui", "Religião marquesana", "Religião das Ilhas Cook"],
    "Polinésia",
    "Oceania",
    "tradições orais; períodos pré-colonial–presente",
    "Religiões indígenas",
    "polynesian",
    "I01",
    status="Viva/Revival",
    coverage="Perfil de família",
)
add_many(
    ["Religião fijiana tradicional", "Kastom de Vanuatu", "Religião Kanak", "Tradições das Terras Altas da Papua", "Religião Asmat", "Religião Tolai", "Tradições Sepik", "Religião Chamorro", "Religião palauana", "Religiões carolínias"],
    "Melanésia/Micronésia",
    "Oceania",
    "tradições orais vivas",
    "Religiões indígenas",
    "indigenous_local",
    "I02",
    coverage="Perfil de família",
)
add_many(["Movimento John Frum", "Movimento Prince Philip", "Cultos cargo melanésios (categoria histórica)"], "Novos movimentos melanésios", "Melanésia", "séculos XIX–XXI", "Movimentos religiosos", "indigenous_local", "C06", coverage="Perfil de família", note="'Cargo cult' é rótulo externo e deve ser usado com cautela.")
add("Religião Sámi histórica e contemporânea", "Urálica/indígena", "Sápmi", "tradição oral; fontes modernas", "Religião indígena/Revival", "sami", "A04", status="Viva/Revival", coverage="Detalhado")
add("Tengrismo", "Túrquica/mongólica", "Ásia Central/Estepe", "fontes antigas e tradições vivas/revivalistas", "Religião indígena/Revival", "tengrism", "A05", status="Viva/Revival", coverage="Detalhado")
add_many(["Xamanismo mongol/Böö mörgöl", "Tradições altaicas", "Religião Sakha/Aiyy", "Tradições Buriates", "Tradições Evenki", "Tradições Chukchi", "Tradições Nenet"], "Siberiana/centro-asiática", "Sibéria/Ásia Central", "tradições orais vivas", "Religiões indígenas", "tengrism", "A05", coverage="Perfil de família", note="Evitar tratar 'xamanismo' como uma religião universal única.")

# Sul, Sudeste e Leste da Ásia.
add("Hinduísmo (visão agregada)", "Dhármica", "Sul da Ásia/diáspora", "raízes védicas; formas atuais plurais", "Família religiosa viva", "hindu", "A06", coverage="Detalhado com ressalvas", note="Não existe um único panteão, cânone ou teologia hindu.")
add_many(["Vaiṣṇavismo", "Śaivismo", "Śāktismo", "Smārta/Advaita devocional", "Tradições tântricas hindus", "Tradições de bhakti"], "Hindu", "Sul da Ásia/diáspora", "Antiguidade–presente", "Famílias hindus", "hindu", "A06", coverage="Perfil de família")
add_many(["Liṅgāyat/Vīraśaiva", "Ayyavazhi", "Tradição Swaminarayan", "ISKCON/Gauḍīya Vaiṣṇava", "Ramakrishna/Vedanta moderno", "Brahma Kumaris"], "Hindu/novo movimento", "Sul da Ásia/diáspora", "medieval–presente", "Tradições religiosas vivas", "hindu", "A06; C06", coverage="Perfil de família")
add("Jainismo (visão agregada)", "Dhármica", "Índia/diáspora", "c. século VI a.C.–presente", "Religião viva", "jain", "A06", coverage="Detalhado")
add_many(["Jainismo Digambara", "Jainismo Śvetāmbara", "Sthānakavāsī/Terāpanth jaina"], "Jain", "Índia/diáspora", "Antiguidade–presente", "Tradições jainas", "jain", "A06", coverage="Perfil de família")
add("Budismo antigo/Theravāda (perfil doutrinário)", "Budista", "Sul/Sudeste da Ásia", "c. século V a.C.–presente", "Religião viva", "buddhist", "A07", coverage="Detalhado")
add_many(["Theravāda do Sri Lanka", "Budismo tailandês", "Budismo birmanês", "Budismo cambojano", "Budismo lao"], "Budista Theravāda", "Sul/Sudeste da Ásia", "Antiguidade–presente", "Tradições budistas", "buddhist", "A07", coverage="Perfil de família")
add("Budismo Mahāyāna (perfil agregado)", "Budista", "Ásia/diáspora", "c. século I a.C.–presente", "Família religiosa viva", "mahayana", "A07; A08", coverage="Detalhado com ressalvas")
add_many(["Budismo Chan/Zen", "Budismo Terra Pura", "Tiantai/Tendai", "Huayan/Kegon", "Budismo Nichiren", "Sōka Gakkai", "Risshō Kōsei Kai"], "Budista Mahāyāna", "Leste Asiático/diáspora", "séculos VI–XXI", "Escolas/movimentos budistas", "mahayana", "A07; A09", coverage="Perfil de família")
add("Budismo Vajrayāna tibetano (perfil agregado)", "Budista Vajrayāna", "Tibete/Himalaia/diáspora", "século VII–presente", "Família religiosa viva", "vajrayana", "A08", coverage="Detalhado com ressalvas")
add_many(["Nyingma", "Kagyu", "Sakya", "Gelug", "Jonang", "Budismo Newar", "Budismo Shingon"], "Budista Vajrayāna/esotérico", "Himalaia/Leste Asiático", "séculos VII–presente", "Escolas budistas", "vajrayana", "A08; A11", coverage="Perfil de família")
add("Bön", "Tibetana", "Tibete/Himalaia", "raízes pré-budistas; forma institucional medieval–presente", "Religião viva", "vajrayana", "A08", coverage="Perfil de família", note="Não é apenas 'xamanismo tibetano'; possui cânone e escolas próprios.")
add_many(["Sarna/Dharma tribal (Adivasi)", "Sanamahismo Meitei", "Donyi-Polo", "Bathou", "Niam Khasi", "Kirat Mundhum", "Religião Kalash", "Mun/Bongthing Lepcha"], "Indígena sul-asiática", "Índia/Nepal/Paquistão", "tradições orais vivas", "Religiões indígenas", "indigenous_local", "A06", coverage="Perfil de família")
add("Siquismo", "Dhármica/monoteísta", "Punjab/diáspora", "século XV–presente", "Religião viva", "sikh", "B04", coverage="Detalhado")
add("Religião popular chinesa", "Chinesa", "China/Taiwan/diáspora", "Antiguidade–presente", "Complexo religioso vivo", "chinese", "A09; A10", coverage="Detalhado")
add("Religião Shang-Zhou", "Chinesa antiga", "China", "c. 1.600–256 a.C.", "Religião estatal/ancestral histórica", "chinese", "A09", status="Histórica/continuidade", coverage="Perfil de família", overrides={"A03": "● Shangdi/Tian conforme o período", "A26": "● Ancestrais reais e linhagem"})
add("Daoismo religioso (perfil agregado)", "Daoista", "China/diáspora", "século II–presente; raízes anteriores", "Família religiosa viva", "daoist", "A09; A10", coverage="Detalhado")
add_many(["Daoismo Zhengyi", "Daoismo Quanzhen", "Tradição Shangqing", "Tradição Lingbao"], "Daoista", "China/diáspora", "séculos IV–presente", "Escolas daoistas", "daoist", "A09; A10", coverage="Perfil de família")
add("Confucionismo", "Confuciana", "Leste Asiático", "século V a.C.–presente", "Tradição ético-ritual", "confucian", "A09", coverage="Detalhado")
add_many(["Neoconfucionismo", "Religião estatal confuciana", "Caodaísmo confuciano-budista (ver Cao Đài)"], "Confuciana", "Leste/Sudeste Asiático", "século X–presente", "Tradições ético-rituais", "confucian", "A09", coverage="Perfil de família")
add_many(["Yiguandao", "Xiantiandao", "Luoísmo", "Religião do Lótus Branco (família histórica)", "Falun Dafa/Falun Gong"], "Salvacionista/chinesa", "China/Taiwan/diáspora", "séculos XV–XXI", "Movimentos religiosos", "chinese", "A09; C06", status="Viva/Histórica", coverage="Perfil de família")
add("Xintoísmo (perfil agregado)", "Japonesa", "Japão/diáspora", "Antiguidade–presente; termo e instituição históricos", "Família religiosa viva", "shinto", "A11", coverage="Detalhado")
add_many(["Xintoísmo de santuário", "Xintoísmo sectário", "Shugendō", "Religião Ryūkyū", "Religião Ainu"], "Japonesa/indígena", "Japão", "Antiguidade–presente", "Tradições religiosas vivas", "shinto", "A11", coverage="Perfil de família")
add_many(["Tenrikyō", "Konkōkyō", "Ōmoto", "Shinnyo-en", "Seichō-no-Ie"], "Novas religiões japonesas", "Japão/diáspora", "séculos XIX–XXI", "Novos movimentos religiosos", "tenrikyo", "C02; C06", coverage="Perfil de família")
add("Tenrikyō (perfil específico)", "Nova religião japonesa", "Japão/diáspora", "1838–presente", "Religião viva", "tenrikyo", "C02", coverage="Detalhado")
add("Muismo/Xamanismo coreano", "Coreana", "Coreia/diáspora", "tradição oral viva", "Religião indígena viva", "korean", "C06", coverage="Detalhado")
add_many(["Cheondoísmo", "Daejongismo", "Jeungsanismo", "Daesun Jinrihoe"], "Novas religiões coreanas", "Coreia", "séculos XIX–XXI", "Novos movimentos religiosos", "korean", "C06", coverage="Perfil de família")
add("Religião popular vietnamita", "Vietnamita", "Vietnã/diáspora", "Antiguidade–presente", "Complexo religioso vivo", "vietnamese", "A09", coverage="Detalhado")
add("Đạo Mẫu/Religião das Mães", "Vietnamita", "Vietnã/diáspora", "formação histórica; tradição viva", "Religião mediúnica viva", "vietnamese", "A09", coverage="Detalhado")
add("Cao Đài", "Sincrética vietnamita", "Vietnã/diáspora", "1926–presente", "Nova religião", "cao_dai", "C06", coverage="Detalhado")
add("Hòa Hảo", "Budista vietnamita", "Vietnã/diáspora", "1939–presente", "Nova religião", "buddhist", "C06", coverage="Perfil de família")
add_many(["Kejawen", "Hinduísmo balinês/Agama Hindu Dharma", "Sunda Wiwitan", "Kaharingan", "Aluk To Dolo", "Parmalim/Batak", "Marapu", "Tolotang"], "Indonésia tradicional", "Indonésia", "tradições históricas vivas", "Religiões indígenas/sincréticas", "indonesian", "C06", coverage="Perfil de família")
add_many(["Anito/Religiões filipinas indígenas", "Tradições Igorot", "Religião Ifugao", "Tradições Lumad", "Dayawismo/revival filipino"], "Indígena filipina", "Filipinas", "tradições orais vivas", "Religiões indígenas/Revival", "indigenous_local", "C06", status="Viva/Revival", coverage="Perfil de família")
add_many(["Religião popular tailandesa", "Culto dos nat birmaneses", "Satsana Phi lao", "Religião Khmer popular"], "Sudeste asiática popular", "Sudeste Asiático", "tradições vivas", "Complexos religiosos", "indonesian", "A07; C06", coverage="Perfil de família")

# Judaísmo, Cristianismo, Islã e famílias relacionadas.
add("Judaísmo bíblico/Israel antigo", "Israelita/Judaica", "Levante", "c. 1.000–200 a.C. (corpus em formação)", "Religião histórica", "judaism", "B01", status="Histórica/continuidade", coverage="Detalhado")
add("Judaísmo rabínico (perfil agregado)", "Judaica", "Global", "século I–presente", "Família religiosa viva", "judaism", "B01", coverage="Detalhado")
add_many(["Judaísmo Ortodoxo", "Judaísmo Conservador/Masorti", "Judaísmo Reformista/Progressista", "Judaísmo Reconstrucionista", "Judaísmo Humanista"], "Judaica", "Global", "séculos XIX–XXI", "Movimentos judaicos", "judaism", "B01", coverage="Perfil de família")
add_many(["Judaísmo caraíta", "Samaritanismo", "Judaísmo Beta Israel"], "Israelita/Judaica", "Oriente Médio/África/diáspora", "Antiguidade–presente", "Religiões vivas", "judaism", "B01", coverage="Perfil de família")
add_many(["Cabala judaica", "Hassidismo", "Misticismo Merkavah/Hekhalot"], "Mística judaica", "Europa/Levante/Global", "Antiguidade tardia–presente", "Correntes místicas", "judaism", "B01", status="Viva/Histórica", coverage="Perfil de família")
add("Cristianismo niceno (perfil agregado)", "Cristã", "Global", "século IV–presente; raízes séc. I", "Família religiosa viva", "christian", "B02", coverage="Detalhado com ressalvas")
add_many(["Igreja Católica", "Ortodoxia Oriental", "Igrejas Ortodoxas Orientais", "Igreja Assíria do Oriente", "Anglicanismo", "Luteranismo", "Tradições Reformadas/Presbiterianas", "Batistas", "Metodismo/Wesleyanismo", "Pentecostalismo", "Movimento Carismático", "Adventismo do Sétimo Dia", "Quakers/Sociedade dos Amigos", "Igrejas Menonitas/Anabatistas"], "Cristã", "Global", "séculos I–XXI", "Tradições cristãs", "christian", "B02", coverage="Perfil de família", note="Sacramentos, santos, autoridade e soteriologia diferem amplamente.")
add_many(["Cristianismo etíope/eritreu", "Cristianismo siríaco", "Cristianismo copta", "Cristianismo armênio"], "Cristã oriental", "África/Oriente Médio/diáspora", "Antiguidade–presente", "Tradições cristãs", "christian", "B02", coverage="Perfil de família")
add_many(["Igreja de Jesus Cristo dos Santos dos Últimos Dias", "Comunidade de Cristo e movimentos do Latter Day Saint"], "Restauracionista cristã", "Américas/Global", "1830–presente", "Novos movimentos cristãos", "lds", "C03", coverage="Detalhado")
add_many(["Testemunhas de Jeová", "Cristadelfianos", "Iglesia ni Cristo"], "Restauracionista/não trinitária", "Global", "séculos XIX–XXI", "Movimentos cristãos", "christian", "B02; C06", coverage="Perfil de família", note="Perfil cristão agregado; doutrinas trinitárias e escatológicas exigem ajuste específico.")
add("Ciência Cristã", "Novo Pensamento cristão", "Global", "1879–presente", "Nova religião cristã", "new_thought", "C06", coverage="Detalhado")
add("Unitarismo Universalista", "Liberal/pluralista", "América do Norte/Global", "1961–presente; raízes anteriores", "Religião pluralista", "modern_nontheist", "C06", coverage="Perfil de família")
add("Islamismo sunita (perfil agregado)", "Islâmica", "Global", "século VII–presente", "Família religiosa viva", "islam", "B03", coverage="Detalhado")
add_many(["Sunismo Hanafi", "Sunismo Maliki", "Sunismo Shafi'i", "Sunismo Hanbali"], "Islâmica sunita", "Global", "século VIII–presente", "Escolas jurídicas", "islam", "B03", coverage="Perfil de família")
add("Xiismo duodecimano", "Islâmica xiita", "Irã/Iraque/Global", "século VII–presente", "Religião viva", "islam", "B03", coverage="Perfil de família", overrides={"A26": "● Ahl al-Bayt e mártires, especialmente Husayn; veneração não divinização", "A38": "● Muḥammad e imames como guias; profecia encerra em Muḥammad"})
add_many(["Ismailismo", "Zaidismo", "Ibadismo"], "Islâmica", "Oriente Médio/África/Ásia", "séculos VIII–presente", "Tradições islâmicas", "islam", "B03", coverage="Perfil de família")
add("Sufismo/Taṣawwuf (perfil agregado)", "Mística islâmica", "Global", "século VIII–presente", "Família mística", "sufi", "B03", coverage="Detalhado")
add_many(["Qadiriyya", "Naqshbandiyya", "Chishtiyya", "Mevleviyya", "Tijaniyya", "Bektashiyya"], "Ordens sufis", "Global", "séculos XII–presente", "Ordens místicas", "sufi", "B03", coverage="Perfil de família")
add_many(["Alevismo", "Druzismo", "Alauísmo", "Yarsan/Ahl-e Haqq"], "Religiões relacionadas ao Islã/iranianas", "Oriente Médio/diáspora", "medieval–presente", "Religiões esotéricas/comunitárias", "islam", "B03; C06", coverage="Perfil de família", note="Não reduzir estas tradições a 'seitas islâmicas'; autodefinição varia.")
add("Fé Bahá'í", "Bahá'í", "Global", "1844/1863–presente", "Religião mundial", "bahai", "B05", coverage="Detalhado")
add("Babismo", "Bábí", "Irã", "1844–presente em pequenos grupos", "Religião histórica/viva", "bahai", "B05", status="Histórica/Viva", coverage="Perfil de família")

# Espiritualismos, esoterismos, neopaganismos e novos movimentos.
add("Espiritualismo moderno", "Espiritualista", "Europa/Américas", "1848–presente", "Movimento religioso", "spiritism", "C01; C06", coverage="Perfil de família", note="Não é idêntico ao Espiritismo kardecista.")
add("Espiritismo kardecista", "Espírita", "França/Brasil/Global", "1857–presente", "Religião/filosofia espiritualista", "spiritism", "C01", coverage="Detalhado")
add_many(["Vale do Amanhecer", "Legião da Boa Vontade/Religião de Deus", "Racionalismo Cristão", "Cultura Racional"], "Novas religiões brasileiras", "Brasil/diáspora", "séculos XX–XXI", "Novos movimentos religiosos", "spiritism", "C01; C06", coverage="Perfil de família")
add("Rastafari", "Afro-cristã", "Jamaica/Global", "década de 1930–presente", "Religião viva", "rastafari", "B02; C06", coverage="Detalhado")
add("Teosofia", "Esotérica moderna", "Global", "1875–presente", "Movimento esotérico", "theosophy", "C06", coverage="Detalhado", note="Esquemas de 'raças-raiz' têm história racialista e não são ciência.")
add("Antroposofia", "Esotérica cristã", "Global", "1912–presente", "Movimento esotérico", "theosophy", "C06", coverage="Perfil de família", note="Perfil teosófico aproximado; cristologia e prática antroposóficas exigem revisão própria.")
add_many(["Novo Pensamento", "Unity/New Thought", "Religious Science/Science of Mind", "Divine Science"], "Novo Pensamento", "América do Norte/Global", "século XIX–presente", "Movimentos metafísicos", "new_thought", "C06", coverage="Detalhado")
add_many(["Wicca Gardneriana", "Wicca Alexandrina", "Wicca Diânica", "Wicca eclética"], "Wicca", "Europa/Américas/Global", "década de 1950–presente", "Neopaganismo", "wicca", "C06", coverage="Perfil de família")
add("Wicca (perfil agregado)", "Wicca", "Global", "década de 1950–presente", "Neopaganismo", "wicca", "C06", coverage="Detalhado")
add_many(["Bruxaria Tradicional moderna", "Movimento da Deusa", "Feri Tradition", "Reclaiming Tradition"], "Neopagã/bruxaria moderna", "Global", "séculos XX–XXI", "Neopaganismos", "wicca", "C06", coverage="Perfil de família")
add("Thelema", "Esotérica moderna", "Global", "1904–presente", "Religião/filosofia esotérica", "thelema", "C04", coverage="Detalhado")
add_many(["Ordo Templi Orientis", "A∴A∴ e linhagens thelêmicas", "Typhonian Order"], "Thelêmica", "Global", "séculos XX–XXI", "Ordens esotéricas", "thelema", "C04", coverage="Perfil de família")
add_many(["Hermetic Order of the Golden Dawn", "Rosacrucianismos modernos", "Martinismo", "Fraternitas Saturni", "Magia do Caos"], "Esotérica/ocultista", "Europa/Américas/Global", "séculos XVII–XXI", "Ordens e correntes esotéricas", "esoteric_order", "C06", status="Viva/Histórica", coverage="Perfil de família")
add("Maçonaria (dimensão ritual/esotérica)", "Fraternal/esotérica", "Global", "1717–presente; antecedentes", "Fraternidade iniciática, não religião única", "esoteric_order", "C06", coverage="Parcial", note="Não classificar automaticamente como religião; membros podem pertencer a religiões distintas.")
add_many(["Druidismo moderno", "Heathenry/Ásatrú", "Forn Siðr", "Rodnovery", "Romuva", "Dievturība", "Helenismo reconstrucionista", "Religio Romana moderna", "Kemetismo moderno", "Neopaganismo celta"], "Neopagã/reconstrucionista", "Europa/Américas/Global", "séculos XIX–XXI", "Religiões de revitalização", "esoteric_order", "C06", coverage="Perfil de família", note="Reconstruções modernas não são continuidade documental simples da Antiguidade.")
add("Satanismo LaVeyano", "Satanista nontheísta", "Global", "1966–presente", "Nova religião", "satanism_nontheist", "C05", coverage="Detalhado")
add("The Satanic Temple", "Satanista nontheísta", "EUA/Global", "2013–presente", "Movimento religioso/ativista", "satanism_nontheist", "C06", coverage="Perfil de família")
add("Temple of Set", "Setiana/esotérica", "Global", "1975–presente", "Nova religião esotérica", "esoteric_order", "C06", coverage="Perfil de família")
add("Satanismos teístas (categoria descentralizada)", "Satanista teísta", "Global", "séculos XX–XXI", "Família descentralizada", "esoteric_order", "C06", coverage="Perfil de família", note="Não há cânone, fundador ou teologia comuns.")
add("Discordianismo", "Paródica/caosmística", "Global", "fins da década de 1950–presente", "Religião paródica/filosofia", "discordian", "C06", coverage="Detalhado")
add_many(["Church of the SubGenius", "Pastafarianismo", "Dudeísmo"], "Paródica", "Global", "séculos XX–XXI", "Religiões paródicas", "discordian", "C06", coverage="Perfil de família")
add("Nova Era/New Age (campo agregado)", "Esotérica descentralizada", "Global", "décadas de 1960–presente", "Campo espiritual plural", "new_age", "C06", coverage="Detalhado com ressalvas", note="Não é uma organização, doutrina ou panteão único.")
add_many(["Neoshamanismo", "Movimento do Potencial Humano", "Canalização de Mestres Ascensos", "Espiritualidade de cristais", "Astrologia psicológica moderna", "Perennialismo esotérico"], "Nova Era/esotérica", "Global", "séculos XX–XXI", "Campos espirituais", "new_age", "C06", coverage="Perfil de família")
add("Cientologia", "Nova religião", "Global", "1954–presente", "Nova religião", "scientology", "C06", coverage="Detalhado")
add_many(["Eckankar", "Movimento Rajneesh/Osho", "Sahaja Yoga", "Ananda Marga", "Meditação Transcendental", "Subud"], "Novos movimentos espirituais", "Global", "séculos XX–XXI", "Novos movimentos religiosos", "new_age", "C06", coverage="Perfil de família")
add_many(["Igreja da Unificação", "World Mission Society Church of God", "Shincheonji", "The Family International"], "Novos movimentos cristãos", "Global", "séculos XX–XXI", "Novos movimentos religiosos", "christian", "C06", coverage="Perfil de família")
add_many(["Raëlismo", "Aetherius Society", "Unarius Academy of Science", "Chen Tao", "Heaven's Gate", "Nation of Yahweh extraterrestre (categoria)"], "Religiões OVNI", "Global", "séculos XX–XXI", "Novos movimentos religiosos", "ufo_religion", "C06", status="Viva/Histórica", coverage="Perfil de família")
add("Santo Daime", "Cristã-enteogênica amazônica", "Brasil/Global", "década de 1930–presente", "Religião enteogênica", "entheogenic_christian", "M03; B02", coverage="Detalhado")
add("União do Vegetal", "Cristã-enteogênica amazônica", "Brasil/Global", "1961–presente", "Religião enteogênica", "entheogenic_christian", "M03; B02", coverage="Detalhado")
add("Barquinha", "Cristã-enteogênica amazônica", "Brasil", "1945–presente", "Religião enteogênica", "entheogenic_christian", "M03; B02", coverage="Detalhado")
add_many(["Igreja Nativa Americana (ver Peyotismo)", "Movimentos modernos de cogumelos sacramentais", "Igrejas de cannabis (categoria descentralizada)"], "Enteogênica moderna", "Américas/Global", "séculos XIX–XXI", "Movimentos religiosos", "native_american_church", "M03", coverage="Perfil de família", note="Legalidade, doutrina e continuidade indígena variam; não agrupar substâncias como tradição única.")
add("Humanismo secular", "Não teísta", "Global", "séculos XIX–XXI", "Cosmovisão ética não religiosa", "modern_nontheist", "C06", coverage="Detalhado", note="Incluído por comparação de cosmovisões, não como panteão.")
add_many(["Naturalismo religioso", "Religião da Humanidade/Positivismo religioso", "Ethical Culture", "Ateísmo religioso organizado"], "Não teísta/humanista", "Global", "séculos XIX–XXI", "Cosmovisões/movimentos éticos", "modern_nontheist", "C06", coverage="Perfil de família")


def finalize_tradition_ids() -> None:
    for index, tradition in enumerate(TRADITIONS, start=1):
        tradition.tradition_id = f"T{index:03d}"


finalize_tradition_ids()


def resolve_mapping(tradition: Tradition) -> dict[str, str]:
    mapping = PROFILES[tradition.profile] | tradition.overrides
    return {archetype.code: mapping.get(archetype.code, "— Sem correlato suficientemente documentado") for archetype in ARCHETYPES}


def mapping_counts(mapping: dict[str, str]) -> dict[str, int]:
    counts = {"●": 0, "≈": 0, "◇": 0, "?": 0, "—": 0}
    for value in mapping.values():
        marker = value[:1] if value else "—"
        counts[marker if marker in counts else "?"] += 1
    return counts


def apply_title(ws, title: str, subtitle: str, last_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color=COLORS["white"])
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["blue"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    ws.sheet_view.showGridLines = False


def style_header_row(ws, row: int, start_col: int, end_col: int, *, height: int = 52) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Aptos", size=9, bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.border = GRID
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def mapping_fill(value: str) -> PatternFill:
    marker = value[:1] if value else "—"
    color = {
        "●": COLORS["light_green"],
        "≈": COLORS["light_gold"],
        "◇": COLORS["light_blue"],
        "?": COLORS["light_red"],
        "—": COLORS["light_gray"],
    }.get(marker, COLORS["white"])
    return PatternFill("solid", fgColor=color)


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "LEIA-ME"
    apply_title(
        ws,
        "Atlas Comparativo de Religiões, Cosmovisões e Arquétipos",
        f"Reformulação acadêmico-editorial de {REVISION_DATE.strftime('%d/%m/%Y')} — matriz heurística, não tabela de equivalências",
        8,
    )
    widths = {"A": 22, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A4:H4")
    ws["A4"] = "VISÃO GERAL"
    ws["A4"].font = Font(size=12, bold=True, color=COLORS["white"])
    ws["A4"].fill = PatternFill("solid", fgColor=COLORS["gold"])
    ws["A4"].alignment = Alignment(horizontal="left")

    stats = [
        ("TRADIÇÕES/COSMOVISÕES", len(TRADITIONS), "linhas catalogadas"),
        ("ARQUÉTIPOS FUNCIONAIS", len(ARCHETYPES), "categorias comparativas"),
        ("PERFIS DE MAPEAMENTO", len(PROFILES), "bases reutilizadas com ressalvas"),
        ("FONTES-PORTA DE ENTRADA", len(SOURCES), "corpora e referências"),
    ]
    for index, (label, value, detail) in enumerate(stats):
        start = 1 + index * 2
        ws.merge_cells(start_row=5, start_column=start, end_row=5, end_column=start + 1)
        ws.merge_cells(start_row=6, start_column=start, end_row=6, end_column=start + 1)
        ws.merge_cells(start_row=7, start_column=start, end_row=7, end_column=start + 1)
        ws.cell(5, start, label)
        ws.cell(6, start, value)
        ws.cell(7, start, detail)
        for row in range(5, 8):
            cell = ws.cell(row, start)
            cell.fill = PatternFill("solid", fgColor=COLORS["cream"] if row != 6 else COLORS["light_blue"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=MEDIUM_NAVY, right=MEDIUM_NAVY, top=MEDIUM_NAVY if row == 5 else THIN_GRAY, bottom=MEDIUM_NAVY if row == 7 else THIN_GRAY)
        ws.cell(5, start).font = Font(size=9, bold=True, color=COLORS["navy"])
        ws.cell(6, start).font = Font(size=20, bold=True, color=COLORS["teal"])
        ws.cell(7, start).font = Font(size=9, italic=True, color=COLORS["dark"])
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 34
    ws.row_dimensions[7].height = 26

    ws.merge_cells("A9:H9")
    ws["A9"] = "COMO LER A MATRIZ"
    ws["A9"].font = Font(size=12, bold=True, color=COLORS["white"])
    ws["A9"].fill = PatternFill("solid", fgColor=COLORS["gold"])
    legend = [
        ("●", "correlação direta/central", COLORS["light_green"], "A função é claramente documentada no recorte indicado."),
        ("≈", "correlação parcial/variante", COLORS["light_gold"], "A função é regional, tardia, metafórica ou apenas uma parte do papel."),
        ("◇", "princípio impessoal/ético", COLORS["light_blue"], "Há uma função comparável, mas não uma entidade personificada."),
        ("?", "hipótese ou documentação insuficiente", COLORS["light_red"], "A leitura é debatida, fragmentária ou depende de fonte externa."),
        ("—", "sem correlato documentado", COLORS["light_gray"], "Não se preenche por semelhança superficial."),
    ]
    for row, (marker, meaning, color, explanation) in enumerate(legend, start=10):
        ws.cell(row, 1, marker)
        ws.cell(row, 2, meaning)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row, 3, explanation)
        for col in range(1, 9):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.border = GRID
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row, 1).font = Font(size=14, bold=True)
        ws.cell(row, 2).font = Font(bold=True)
        ws.row_dimensions[row].height = 28

    ws.merge_cells("A16:H16")
    ws["A16"] = "PRINCÍPIOS EDITORIAIS"
    ws["A16"].font = Font(size=12, bold=True, color=COLORS["white"])
    ws["A16"].fill = PatternFill("solid", fgColor=COLORS["gold"])
    principles = [
        "“Arquétipo” significa aqui função comparativa de indexação. Não afirma identidade entre deuses, origem comum, psicologia universal ou equivalência teológica.",
        "A data registra a atestação ou o recorte usado, não o nascimento de um povo. Tradições orais antigas não recebem datas paleolíticas sem evidência.",
        "Religiões vivas são internamente diversas. Um perfil de família é ponto de partida e deve ser revisado com especialistas e praticantes da tradição.",
        "Ciência, neurotransmissores e disciplinas acadêmicas foram retirados do panteão. Cosmovisões não teístas aparecem apenas quando a comparação funcional é útil.",
        "O catálogo é amplo, mas não literalmente exaustivo: existem milhares de religiões locais, linhagens, denominações e variantes sem corpus público ou nomenclatura estável.",
    ]
    for row, principle in enumerate(principles, start=17):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, f"• {principle}")
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS["cream"])
        ws.cell(row, 1).border = GRID
        ws.row_dimensions[row].height = 38

    ws.merge_cells("A23:H23")
    ws["A23"] = "NAVEGAÇÃO"
    ws["A23"].font = Font(size=12, bold=True, color=COLORS["white"])
    ws["A23"].fill = PatternFill("solid", fgColor=COLORS["gold"])
    navigation = [
        ("Matriz global", "Matriz global", "Compare 460 tradições × 44 funções e filtre por família, região ou cobertura."),
        ("Catálogo", "Catálogo", "Consulte metadados, notas de escopo e contagem de confiança."),
        ("Arquétipos", "Arquétipos", "Leia definição e exclusões de cada categoria."),
        ("Cronologia", "Cronologia", "Use uma periodização global com ressalvas regionais."),
        ("Fontes", "Fontes", "Abra corpora, referências acadêmicas e fontes confessionais identificadas."),
        ("Revisões", "Revisões", "Veja correções estruturais feitas em relação ao workbook original."),
        ("Aeons — autoral", "Aeons — autoral", "Consulte a hipótese esotérica original, agora separada da historiografia."),
    ]
    for row, (label, target, detail) in enumerate(navigation, start=24):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row, 1, label)
        ws.cell(row, 1).hyperlink = f"#'{target}'!A1"
        ws.cell(row, 1).style = "Hyperlink"
        ws.cell(row, 3, detail)
        for col in (1, 3):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=COLORS["light_blue"] if row % 2 == 0 else COLORS["cream"])
            ws.cell(row, col).border = GRID
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A4"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Matriz global")
    last_col = 9 + len(ARCHETYPES)
    apply_title(
        ws,
        "Matriz global de funções arquetípicas",
        "Filtre por tradição, família, região, status ou cobertura. Símbolos mostram força e natureza da correlação; “—” é um resultado válido.",
        last_col,
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws["A3"] = "● direto/central   |   ≈ parcial/variante   |   ◇ princípio impessoal   |   ? incerto   |   — sem correlato documentado"
    ws["A3"].font = Font(bold=True, color=COLORS["navy"])
    ws["A3"].fill = PatternFill("solid", fgColor=COLORS["sand"])
    ws["A3"].alignment = Alignment(horizontal="left")

    headers = ["ID", "Tradição / cosmovisão", "Família", "Região", "Período / atestação", "Tipo", "Status", "Cobertura", "Fontes"]
    headers.extend(f"{item.code} — {item.name}" for item in ARCHETYPES)
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, last_col, height=72)

    for index, archetype in enumerate(ARCHETYPES, start=10):
        ws.cell(header_row, index).comment = Comment(
            f"{archetype.definition}\n\nEvitar: {archetype.exclusion}",
            "Revisão metodológica",
        )

    for row, tradition in enumerate(TRADITIONS, start=header_row + 1):
        mapping = resolve_mapping(tradition)
        metadata = [
            tradition.tradition_id,
            tradition.name,
            tradition.family,
            tradition.region,
            tradition.period,
            tradition.kind,
            tradition.status,
            tradition.coverage,
            tradition.sources,
        ]
        for col, value in enumerate(metadata, start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 1).hyperlink = f"#'Catálogo'!A{row}"
        ws.cell(row, 1).style = "Hyperlink"
        if tradition.note:
            ws.cell(row, 2).comment = Comment(tradition.note, "Nota de escopo")
        for col, archetype in enumerate(ARCHETYPES, start=10):
            value = mapping[archetype.code]
            cell = ws.cell(row, col, value)
            cell.fill = mapping_fill(value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Aptos", size=8, color=COLORS["dark"])
        ws.row_dimensions[row].height = 62

    end_row = header_row + len(TRADITIONS)
    add_table(ws, "MatrizGlobal", f"A{header_row}:{get_column_letter(last_col)}{end_row}")
    widths = [11, 34, 24, 22, 23, 25, 16, 20, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(10, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.freeze_panes = "J6"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{end_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"1:{header_row}"


def build_catalog(wb: Workbook) -> None:
    ws = wb.create_sheet("Catálogo")
    headers = [
        "ID",
        "Tradição / cosmovisão",
        "Família",
        "Região",
        "Período / atestação",
        "Tipo",
        "Status",
        "Cobertura",
        "Perfil-base",
        "Fontes",
        "Nota de escopo",
        "●",
        "≈",
        "◇",
        "?",
        "—",
    ]
    apply_title(
        ws,
        "Catálogo global de tradições",
        "Inventário amplo para pesquisa. 'Perfil de família' indica mapeamento provisório; 'Fragmentário' exige fontes específicas antes de afirmar nomes ou funções.",
        len(headers),
    )
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, len(headers), height=56)

    for row, tradition in enumerate(TRADITIONS, start=header_row + 1):
        counts = mapping_counts(resolve_mapping(tradition))
        values = [
            tradition.tradition_id,
            tradition.name,
            tradition.family,
            tradition.region,
            tradition.period,
            tradition.kind,
            tradition.status,
            tradition.coverage,
            tradition.profile,
            tradition.sources,
            tradition.note,
            counts["●"],
            counts["≈"],
            counts["◇"],
            counts["?"],
            counts["—"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 1).hyperlink = f"#'Matriz global'!A{row}"
        ws.cell(row, 1).style = "Hyperlink"
        ws.row_dimensions[row].height = 48

    end_row = header_row + len(TRADITIONS)
    add_table(ws, "CatalogoGlobal", f"A{header_row}:P{end_row}")
    widths = [11, 36, 25, 23, 25, 27, 16, 20, 22, 16, 50, 7, 7, 7, 7, 7]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"A{header_row}:P{end_row}"
    ws.print_title_rows = f"1:{header_row}"


def build_archetypes(wb: Workbook) -> None:
    ws = wb.create_sheet("Arquétipos")
    headers = ["Código", "Função comparativa", "Critério de inclusão", "Evitar / não confundir", "●", "≈", "◇", "?", "—"]
    apply_title(
        ws,
        "Dicionário de arquétipos funcionais",
        "As categorias são instrumentos de indexação comparativa. Elas não afirmam universais psicológicos nem que entidades de tradições diferentes sejam a mesma.",
        len(headers),
    )
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, len(headers), height=52)

    aggregate = {item.code: {"●": 0, "≈": 0, "◇": 0, "?": 0, "—": 0} for item in ARCHETYPES}
    for tradition in TRADITIONS:
        for code, value in resolve_mapping(tradition).items():
            marker = value[:1] if value[:1] in aggregate[code] else "?"
            aggregate[code][marker] += 1

    for row, archetype in enumerate(ARCHETYPES, start=header_row + 1):
        counts = aggregate[archetype.code]
        values = [
            archetype.code,
            archetype.name,
            archetype.definition,
            archetype.exclusion,
            counts["●"],
            counts["≈"],
            counts["◇"],
            counts["?"],
            counts["—"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 56
    end_row = header_row + len(ARCHETYPES)
    add_table(ws, "DicionarioArquetipos", f"A{header_row}:I{end_row}")
    for col, width in enumerate([12, 34, 62, 58, 8, 8, 8, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def build_chronology(wb: Workbook) -> None:
    ws = wb.create_sheet("Cronologia")
    headers = ["Macroperíodo", "Intervalo aproximado", "Contextos úteis", "Mudanças religiosas documentáveis", "Limites da periodização"]
    apply_title(
        ws,
        "Cronologia comparativa e contextos históricos",
        "Períodos são ferramentas de navegação, não estágios universais de evolução religiosa. Datas e transições variam por região.",
        len(headers),
    )
    rows = [
        ("Pré-história e arqueologia", "antes de c. 3.200 a.C.", "Caçadores-coletores; aldeias; primeiros centros monumentais", "Sepultamentos, arte, depósitos e espaços rituais; crenças específicas raramente recuperáveis", "Não projetar 'xamanismo', 'Deusa-Mãe' ou enteogenia sem evidência contextual."),
        ("Primeiras sociedades letradas / Idade do Bronze", "c. 3.200–1.200 a.C.", "Mesopotâmia, Egito, Vale do Indo, China Shang, Egeu", "Templos, sacerdócios, realeza sagrada, hinos, adivinhação e panteões locais", "A escrita preserva sobretudo elites; tradições orais coexistem."),
        ("Idade do Ferro e formações clássicas", "c. 1.200–200 a.C.", "Impérios regionais, cidades-estado, redes mediterrâneas e asiáticas", "Canônicos, reformas rituais, filosofia, profecia e identidades imperiais", "“Era Axial” é uma hipótese comparativa debatida, não faixa universal."),
        ("Antiguidade tardia", "c. 200 a.C.–600 d.C.", "Impérios romano, parto/sassânida, Han e Gupta; rotas eurasiáticas", "Expansão de budismos, cristianismos, judaísmo rabínico, maniqueísmo e cultos sincréticos", "As fronteiras entre religião, etnia, filosofia e lei são modernas."),
        ("Períodos medievais regionais", "c. 600–1500", "Califados, cristandades, reinos africanos, Índia, China, Japão, Américas", "Islamização/cristianização, escolas budistas, bhakti, sufismo e religiões de corte", "“Idade Média” é eurocêntrico quando aplicado globalmente."),
        ("Primeira modernidade e colonialismos", "c. 1450–1800", "Expansões marítimas, escravidão atlântica, estados confessionais", "Reformas cristãs, diásporas africanas, sincretismos, missões e repressões", "Contato não significa substituição total; continuidades indígenas persistem."),
        ("Modernidade industrial", "c. 1800–1945", "Estados-nação, ciência moderna, colonialismo e impressão de massa", "Espiritualismo, Espiritismo, Teosofia, revivalismos e novas religiões", "Ciência não deve virar uma coluna de panteão; é outro regime de conhecimento."),
        ("Contemporâneo e global", "1945–presente", "Descolonização, migrações, mídia, direitos indígenas e internet", "Pentecostalismos, neopaganismos, New Age, religiões OVNI, revitalizações e novas diásporas", "Categorias mudam rapidamente; fontes devem registrar data e autodefinição."),
    ]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, len(headers), height=50)
    for row, values in enumerate(rows, start=header_row + 1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=COLORS["cream"] if row % 2 else COLORS["light_blue"])
        ws.row_dimensions[row].height = 82
    add_table(ws, "CronologiaComparativa", f"A{header_row}:E{header_row + len(rows)}")
    for col, width in enumerate([31, 24, 48, 62, 62], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def build_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Fontes")
    headers = ["Código", "Escopo", "Título / recurso", "Instituição", "URL", "Uso nesta revisão"]
    apply_title(
        ws,
        "Fontes e trilhas de verificação",
        "Mistura deliberada de corpora primários, referências acadêmicas e fontes confessionais identificadas. Uma fonte oficial descreve autocompreensão; não substitui história crítica.",
        len(headers),
    )
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, len(headers), height=52)
    for row, source in enumerate(SOURCES, start=header_row + 1):
        code, scope, title, institution, url, use = source
        for col, value in enumerate([code, scope, title, institution, url, use], start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if url:
            ws.cell(row, 5).hyperlink = url
            ws.cell(row, 5).style = "Hyperlink"
        ws.row_dimensions[row].height = 48
    end_row = header_row + len(SOURCES)
    add_table(ws, "FontesPesquisa", f"A{header_row}:F{end_row}")
    for col, width in enumerate([11, 24, 48, 38, 62, 62], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def build_revisions(wb: Workbook) -> None:
    ws = wb.create_sheet("Revisões")
    headers = ["Item original / problema", "Ação", "Justificativa historiográfica", "Onde consultar agora"]
    apply_title(
        ws,
        "Principais correções editoriais",
        "Registro transparente das mudanças de taxonomia, datação e método em relação a UNO.xlsx.",
        len(headers),
    )
    rows = [
        ("“Cosmovisão” como coluna misturava religião, ciência, disciplina e neurotransmissor.", "Substituída por Tradição / cosmovisão + Tipo.", "Biologia, química, neurociência e física quântica não são panteões religiosos.", "Catálogo; Matriz global"),
        ("“Potência” e eras universais vagas.", "Substituídas por família, região, atestação e cronologia crítica.", "Contextos políticos não avançam em um único eixo global.", "Cronologia"),
        ("“Subsaarianos”, “Indígena” e “Aborígene” como panteões únicos.", "Desagregados em povos e tradições; perfis agregados recebem ressalva.", "Os rótulos apagavam centenas de línguas, nações e sistemas.", "Catálogo"),
        ("Datas como Māori 4.000 a.C. ou 'Aborígene 10.000 a.C.'.", "Trocadas por atestação oral/documental sem data de origem inventada.", "Antiguidade cultural não autoriza datar conteúdos atuais sem evidência.", "Catálogo"),
        ("Cosmologia científica datada de 3 a.C. e Stephen Hawking como livro-base antigo.", "Removida do panteão.", "É anacronismo e mistura de regimes de conhecimento.", "Revisões"),
        ("Zoroastrismo repetido e datado no século XVIII.", "Unificado e datado com ressalva sobre as Gāthās e fontes pahlavi.", "A tradição é muito anterior e permanece viva.", "Matriz global; Fontes A05"),
        ("Stregheria/Aradia colocada em 700 a.C.", "Transferida para bruxaria/neopaganismo moderno.", "Aradia é publicação de 1899; continuidade antiga é reivindicação, não consenso.", "Catálogo"),
        ("Livre-arbítrio, determinismo, hedonismo e empirismo como panteões.", "Retirados da matriz religiosa.", "São problemas ou correntes filosóficas, não tradições com um panteão comum.", "LEIA-ME"),
        ("Amanita, ayahuasca e cogumelos atribuídos a povos/épocas sem evidência específica.", "Marcados como incertos ou limitados a religiões/contextos documentados.", "Arte rupestre e xamanismo não provam uso de uma substância.", "Fontes M02–M03"),
        ("Exu, Tawûsê Melek e deuses ctônicos aproximados ao Diabo.", "Equivalências recusadas explicitamente.", "Sincretização colonial e semelhança funcional não são identidade teológica.", "Matriz global A31"),
        ("Obras tardias tratadas como contemporâneas dos povos descritos.", "Atestação do corpus separada da antiguidade alegada do material.", "Eddas, Mabinogion, Popol Wuj e crônicas coloniais exigem crítica de transmissão.", "Catálogo"),
        ("Io Matua Kore como criador Māori universal.", "Marcado como restrito/contestado; perfil usa genealogias variáveis.", "A própria historiografia Māori registra diversidade e debate.", "Fontes I01"),
        ("“Livro-base” único para tradições orais.", "Substituído por fontes/corpora e notas de transmissão.", "Muitas tradições não têm cânone único e protegem conhecimento não público.", "Fontes; Catálogo"),
        ("Toda célula precisava de um nome.", "Introduzidos ●, ≈, ◇, ? e —.", "Ausência documentada é mais rigorosa que preenchimento especulativo.", "LEIA-ME; Matriz global"),
        ("Aeons astrológicos e thelêmicos misturados à cronologia histórica.", "Preservados em aba própria com rótulo de hipótese esotérica.", "São esquemas interpretativos modernos, não consenso historiográfico.", "Aeons — autoral"),
    ]
    header_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(header_row, col, header)
    style_header_row(ws, header_row, 1, len(headers), height=50)
    for row, values in enumerate(rows, start=header_row + 1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=COLORS["cream"] if row % 2 else COLORS["light_blue"])
        ws.row_dimensions[row].height = 72
    add_table(ws, "RegistroRevisoes", f"A{header_row}:D{header_row + len(rows)}")
    for col, width in enumerate([52, 46, 68, 30], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def build_aeons(wb: Workbook) -> None:
    ws = wb.create_sheet("Aeons — autoral")
    apply_title(
        ws,
        "Camada interpretativa original: aeons astrológicos e thelêmicos",
        "Preservada como hipótese esotérica/autoral. Não usar como periodização arqueológica ou consenso da história das religiões.",
        6,
    )
    ws.merge_cells("A4:F4")
    ws["A4"] = "STATUS EPISTÊMICO: interpretação moderna / esotérica"
    ws["A4"].font = Font(bold=True, color=COLORS["white"])
    ws["A4"].fill = PatternFill("solid", fgColor="9C0006")
    ws["A4"].alignment = Alignment(horizontal="center")

    start_row = 6
    copied_last_row = start_row
    if SOURCE.exists():
        original = load_workbook(SOURCE, data_only=False, read_only=True)
        if "Planilha1" in original.sheetnames:
            source_ws = original["Planilha1"]
            for source_row_index, source_row in enumerate(source_ws.iter_rows(), start=1):
                for source_col_index, cell in enumerate(source_row, start=1):
                    if cell.value is None:
                        continue
                    target_row = start_row + source_row_index - 1
                    ws.cell(target_row, source_col_index, cell.value)
                    copied_last_row = max(copied_last_row, target_row)
        original.close()
    if ws.cell(start_row, 1).value is None:
        fallback = [
            ["AEON", "QUANTITAS", "ERA ASTROLÓGICA", "AEON ASTROLÓGICO", "AEON THELÊMICO", "CORRESPONDÊNCIAS"],
            ["HÓRUS-AQUÁRIO", 22, "Era Moderna", "Aeon de Aquário", "Aeon de Hórus", "Hipótese interpretativa moderna."],
        ]
        for r_index, row in enumerate(fallback, start=start_row):
            for c_index, value in enumerate(row, start=1):
                ws.cell(r_index, c_index, value)
            copied_last_row = max(copied_last_row, r_index)

    last_row = max(start_row, copied_last_row)
    style_header_row(ws, start_row, 1, 6, height=48)
    for row in range(start_row + 1, last_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.border = GRID
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=COLORS["cream"] if row % 2 else COLORS["light_blue"])
        ws.row_dimensions[row].height = 62
    for col, width in enumerate([24, 13, 28, 28, 28, 80], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = f"A{start_row + 1}"


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.properties.title = "Atlas Comparativo de Religiões, Cosmovisões e Arquétipos"
    wb.properties.subject = "Religião comparada e história das religiões"
    wb.properties.creator = "Reformulação editorial assistida por IA a partir do estudo original"
    wb.properties.keywords = "religião comparada, panteões, arquétipos, cosmovisões, história"
    wb.properties.description = (
        f"Revisão de {REVISION_DATE.isoformat()} com {len(TRADITIONS)} tradições e "
        f"{len(ARCHETYPES)} categorias funcionais."
    )

    build_readme(wb)
    build_matrix(wb)
    build_catalog(wb)
    build_archetypes(wb)
    build_chronology(wb)
    build_sources(wb)
    build_revisions(wb)
    build_aeons(wb)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    return wb


def validate_data() -> None:
    if len({tradition.name for tradition in TRADITIONS}) != len(TRADITIONS):
        raise ValueError("Duplicate tradition names detected")
    if set(archetype.code for archetype in ARCHETYPES) != {f"A{i:02d}" for i in range(1, 45)}:
        raise ValueError("Archetype codes must be contiguous A01–A44")
    missing_profiles = sorted({tradition.profile for tradition in TRADITIONS} - PROFILES.keys())
    if missing_profiles:
        raise ValueError(f"Missing profiles: {missing_profiles}")
    valid_markers = {"●", "≈", "◇", "?", "—"}
    for profile_name, profile in PROFILES.items():
        unknown_codes = sorted(set(profile) - {item.code for item in ARCHETYPES})
        if unknown_codes:
            raise ValueError(f"Unknown archetype codes in {profile_name}: {unknown_codes}")
        for code, value in profile.items():
            if not value or value[0] not in valid_markers:
                raise ValueError(f"Invalid marker in {profile_name}/{code}: {value!r}")


def write_structured_csv(workbook: Workbook) -> None:
    """Write one AI-friendly long CSV plus a lossless semantic cell layer.

    The normalized records make common retrieval tasks trivial. The
    ``workbook_cell`` records retain every non-empty workbook value, comment,
    and hyperlink so information from narrative sheets is not lost.
    """
    fields = [
        "schema_version",
        "record_type",
        "sheet",
        "row",
        "column",
        "coordinate",
        "tradition_id",
        "tradition_name",
        "family",
        "region",
        "period",
        "kind",
        "status",
        "coverage",
        "profile",
        "source_codes",
        "note",
        "archetype_code",
        "archetype_name",
        "marker",
        "value",
        "definition",
        "exclusion",
        "source_title",
        "institution",
        "url",
        "hyperlink",
        "comment",
    ]

    def blank_record(record_type: str) -> dict[str, object]:
        record = {field: "" for field in fields}
        record["schema_version"] = "1.0"
        record["record_type"] = record_type
        return record

    with CSV_OUTPUT.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="raise")
        writer.writeheader()

        metadata = [
            ("dataset_title", "Atlas Comparativo de Religiões, Cosmovisões e Arquétipos"),
            ("revision_date", REVISION_DATE.isoformat()),
            ("tradition_count", str(len(TRADITIONS))),
            ("archetype_count", str(len(ARCHETYPES))),
            ("profile_count", str(len(PROFILES))),
            ("legend_●", "correlação direta/central"),
            ("legend_≈", "correlação parcial, regional, tardia ou metafórica"),
            ("legend_◇", "princípio impessoal, ético ou estrutural"),
            ("legend_?", "hipótese, disputa ou documentação insuficiente"),
            ("legend_—", "sem correlato suficientemente documentado"),
            ("primary_record_type", "Use tradition_archetype para a matriz normalizada pronta para consulta."),
            ("join_keys", "tradition_id liga tradition a tradition_archetype; archetype_code liga archetype_definition; source_codes referencia source."),
            ("workbook_cell_layer", "workbook_cell preserva toda célula não vazia, comentário e hyperlink das oito abas do XLSX."),
            (
                "method_warning",
                "Arquétipo é função de indexação; não prova identidade, origem comum ou universal psicológico.",
            ),
        ]
        for key, value in metadata:
            record = blank_record("dataset_metadata")
            record["coordinate"] = key
            record["value"] = value
            writer.writerow(record)

        for tradition in TRADITIONS:
            mapping = resolve_mapping(tradition)
            counts = mapping_counts(mapping)
            record = blank_record("tradition")
            record.update(
                tradition_id=tradition.tradition_id,
                tradition_name=tradition.name,
                family=tradition.family,
                region=tradition.region,
                period=tradition.period,
                kind=tradition.kind,
                status=tradition.status,
                coverage=tradition.coverage,
                profile=tradition.profile,
                source_codes=tradition.sources,
                note=tradition.note,
                value=tradition.name,
                comment=(
                    f"direct={counts['●']}; partial={counts['≈']}; conceptual={counts['◇']}; "
                    f"uncertain={counts['?']}; absent={counts['—']}"
                ),
            )
            writer.writerow(record)

            for archetype in ARCHETYPES:
                value = mapping[archetype.code]
                record = blank_record("tradition_archetype")
                record.update(
                    tradition_id=tradition.tradition_id,
                    tradition_name=tradition.name,
                    family=tradition.family,
                    region=tradition.region,
                    period=tradition.period,
                    kind=tradition.kind,
                    status=tradition.status,
                    coverage=tradition.coverage,
                    profile=tradition.profile,
                    source_codes=tradition.sources,
                    note=tradition.note,
                    archetype_code=archetype.code,
                    archetype_name=archetype.name,
                    marker=value[:1],
                    value=value[2:] if len(value) > 1 and value[1:2] == " " else value[1:],
                    definition=archetype.definition,
                    exclusion=archetype.exclusion,
                )
                writer.writerow(record)

        for archetype in ARCHETYPES:
            record = blank_record("archetype_definition")
            record.update(
                archetype_code=archetype.code,
                archetype_name=archetype.name,
                value=archetype.name,
                definition=archetype.definition,
                exclusion=archetype.exclusion,
            )
            writer.writerow(record)

        for code, scope, title, institution, url, use in SOURCES:
            record = blank_record("source")
            record.update(
                source_codes=code,
                family=scope,
                value=use,
                source_title=title,
                institution=institution,
                url=url,
            )
            writer.writerow(record)

        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None and cell.comment is None and cell.hyperlink is None:
                        continue
                    record = blank_record("workbook_cell")
                    record.update(
                        sheet=ws.title,
                        row=cell.row,
                        column=cell.column,
                        coordinate=cell.coordinate,
                        value=cell.value if cell.value is not None else "",
                        hyperlink=cell.hyperlink.target if cell.hyperlink else "",
                        comment=cell.comment.text if cell.comment else "",
                    )
                    writer.writerow(record)


def main() -> None:
    validate_data()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    CSV_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT)
    write_structured_csv(workbook)
    print(
        f"generated={OUTPUT.name} traditions={len(TRADITIONS)} "
        f"archetypes={len(ARCHETYPES)} profiles={len(PROFILES)} sheets={len(workbook.sheetnames)} "
        f"csv={CSV_OUTPUT.name}"
    )


if __name__ == "__main__":
    main()
